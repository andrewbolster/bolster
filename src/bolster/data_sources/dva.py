"""DVA (Driver & Vehicle Agency) Monthly Tests Statistics Module.

This module provides access to Northern Ireland's Driver & Vehicle Agency monthly
test statistics, including vehicle tests, driver tests, and theory tests.

Data is published monthly by the Department for Infrastructure (DfI) Northern Ireland.

Data Coverage:
    - Vehicle Tests (Full & Retests): April 2014 - Present
    - Driver Tests: April 2014 - Present
    - Theory Tests: April 2014 - Present
    - Test breakdowns by category and test centre

Data Source: Department for Infrastructure Northern Ireland provides Driver & Vehicle Agency
statistics through their publications portal at https://www.infrastructure-ni.gov.uk/publications?f%5B0%5D=type%3Astatisticalreports.
The DVA publishes monthly test statistics covering vehicle tests, driver tests, and theory tests
conducted across Northern Ireland, providing comprehensive data on driving and vehicle testing performance.

Update Frequency: Monthly publications are released covering the previous month's test statistics.
DVA data is published by the Department for Infrastructure Analytics Branch approximately 4-6 weeks
after the reference month ends, providing consistent monthly updates on driving test performance
and vehicle testing statistics across Northern Ireland.

Publication Details:
    - Published by: Department for Infrastructure (DfI) - Analytics Branch
    - Data Source: DVA Business & Regulatory Statistics

Example:
    >>> from bolster.data_sources import dva
    >>> # Get latest vehicle test statistics
    >>> df = dva.get_latest_vehicle_tests()
    >>> print(df.tail())

    >>> # Get latest driver test statistics
    >>> df = dva.get_latest_driver_tests()
    >>> print(f"Latest month: {df.iloc[-1]['month']}")

    >>> # Get latest theory test statistics
    >>> df = dva.get_latest_theory_tests()
    >>> print(f"Total tests: {df['tests_conducted'].sum():,}")

    >>> # Get all test types combined
    >>> data = dva.get_latest_all_tests()
    >>> print(data.keys())  # ['vehicle', 'driver', 'theory']
"""

import contextlib
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Optional, Union

import pandas as pd
from bs4 import BeautifulSoup

from bolster.utils.web import session

logger = logging.getLogger(__name__)

# Cache directory
CACHE_DIR = Path.home() / ".cache" / "bolster" / "dva"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

# Base URL for DVA statistics publications
DVA_PUBLICATIONS_URL = "https://www.infrastructure-ni.gov.uk/publications/type/statistics"
DVA_SEARCH_TERM = "driver-and-vehicle-agency-monthly-tests-conducted"


class DVADataError(Exception):
    """Base exception for DVA data errors."""

    pass


class DVADataNotFoundError(DVADataError):
    """Data file not available."""

    pass


def _hash_url(url: str) -> str:
    """Generate a safe filename from a URL."""
    import hashlib

    return hashlib.md5(url.encode()).hexdigest()


def _download_file(url: str, cache_ttl_hours: int = 24, force_refresh: bool = False) -> Path:
    """Download a file with caching support.

    Args:
        url: URL to download
        cache_ttl_hours: Cache validity in hours (default: 24)
        force_refresh: Force re-download even if cached

    Returns:
        Path to downloaded file

    Raises:
        DVADataNotFoundError: If download fails
    """
    url_hash = _hash_url(url)
    ext = Path(url).suffix or ".xlsx"
    cache_path = CACHE_DIR / f"{url_hash}{ext}"

    # Check cache
    if not force_refresh and cache_path.exists():
        age = datetime.now() - datetime.fromtimestamp(cache_path.stat().st_mtime)
        if age.total_seconds() < cache_ttl_hours * 3600:
            logger.info(f"Using cached file: {cache_path}")
            return cache_path

    # Download
    try:
        logger.info(f"Downloading {url}")
        response = session.get(url, timeout=60)
        response.raise_for_status()
        cache_path.write_bytes(response.content)
        logger.info(f"Saved to {cache_path}")
        return cache_path
    except Exception as e:
        raise DVADataNotFoundError(f"Failed to download {url}: {e}") from e


def get_latest_dva_publication_url() -> tuple[str, str, datetime]:
    """Get the URL of the latest DVA Monthly Tests publication.

    Attempts to find the most recent DVA monthly tests statistics publication
    by trying recent months in reverse order.

    Returns:
        Tuple of (excel_url, publication_title, publication_date)

    Raises:
        DVADataNotFoundError: If unable to find any recent publication

    Example:
        >>> url, title, pub_date = get_latest_dva_publication_url()
        >>> print(f"Latest: {title} (published {pub_date.strftime('%Y-%m-%d')})")
    """
    from dateutil.relativedelta import relativedelta

    # Try the last 6 months to find the most recent publication
    current_date = datetime.now()
    months_to_try = []

    for i in range(6):
        check_date = current_date - relativedelta(months=i)
        months_to_try.append(check_date)

    month_names = [
        "january",
        "february",
        "march",
        "april",
        "may",
        "june",
        "july",
        "august",
        "september",
        "october",
        "november",
        "december",
    ]

    for check_date in months_to_try:
        month_name = month_names[check_date.month - 1]
        year = check_date.year

        # Construct the publication URL
        pub_url = (
            f"https://www.infrastructure-ni.gov.uk/publications/"
            f"driver-and-vehicle-agency-monthly-tests-conducted-statistics-{month_name}-{year}"
        )

        logger.info(f"Trying publication URL: {pub_url}")

        try:
            response = session.get(pub_url, timeout=30)
            if response.status_code == 200:
                logger.info(f"Found publication for {month_name.title()} {year}")

                soup = BeautifulSoup(response.content, "html.parser")

                # Find Excel file link
                excel_url = None
                for file_link in soup.find_all("a", href=True):
                    href = file_link["href"]
                    if ".xlsx" in href.lower() and "tables" in href.lower():
                        if not href.startswith("http"):
                            href = f"https://www.infrastructure-ni.gov.uk{href}"
                        excel_url = href
                        break

                if excel_url:
                    pub_title = f"DVA Monthly Tests Conducted Statistics - {month_name.title()} {year}"
                    pub_date = datetime(year, check_date.month, 1)
                    logger.info(f"Found DVA Excel file: {excel_url}")
                    return excel_url, pub_title, pub_date

        except Exception as e:
            logger.debug(f"Failed to fetch {pub_url}: {e}")
            continue

    raise DVADataNotFoundError("Could not find any DVA monthly tests publications in the last 6 months") from e


def _parse_month_year(date_str: str) -> Optional[datetime]:
    """Parse a 'YYYY Month' string into a datetime.

    Args:
        date_str: String like '2024 December' or '2025 January'

    Returns:
        datetime object or None if parsing fails
    """
    if not date_str or not isinstance(date_str, str):
        return None

    date_str = date_str.strip()

    # Pattern: "YYYY Month"
    match = re.match(
        r"(\d{4})\s+(January|February|March|April|May|June|July|August|September|October|November|December)",
        date_str,
        re.IGNORECASE,
    )
    if match:
        year, month_name = match.groups()
        month_map = {
            "january": 1,
            "february": 2,
            "march": 3,
            "april": 4,
            "may": 5,
            "june": 6,
            "july": 7,
            "august": 8,
            "september": 9,
            "october": 10,
            "november": 11,
            "december": 12,
        }
        return datetime(int(year), month_map[month_name.lower()], 1)

    return None


def parse_vehicle_tests(file_path: Union[str, Path]) -> pd.DataFrame:
    """Parse DVA vehicle tests data from Excel file.

    Extracts full vehicle tests conducted from Table 1.1a.

    Args:
        file_path: Path to the DVA Excel file

    Returns:
        DataFrame with columns:
            - date: datetime (first day of month)
            - year: int
            - month: str (month name)
            - tests_conducted: int (full tests conducted)
            - rolling_12_month_total: int (optional, rolling 12-month sum)

    Example:
        >>> df = parse_vehicle_tests("dva-monthly-tests-december-2025.xlsx")
        >>> print(df.tail())
    """
    from openpyxl import load_workbook

    logger.info(f"Parsing vehicle tests from: {file_path}")

    wb = load_workbook(file_path, data_only=True)

    # Find the vehicle tests sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "1_1a" in name and "Veh" in name:
            sheet_name = name
            break

    if not sheet_name:
        raise DVADataError("Could not find vehicle tests sheet (1_1a)")

    sheet = wb[sheet_name]

    # Parse data rows (skip header rows 1-3)
    records = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        date_str = row[0]
        tests = row[1]
        rolling_total = row[3] if len(row) > 3 else None

        if not date_str or not tests:
            continue

        date = _parse_month_year(str(date_str))
        if not date:
            continue

        try:
            tests_int = int(tests)
        except (ValueError, TypeError):
            continue

        record = {
            "date": date,
            "year": date.year,
            "month": date.strftime("%B"),
            "tests_conducted": tests_int,
        }

        if rolling_total:
            with contextlib.suppress(ValueError, TypeError):
                record["rolling_12_month_total"] = int(rolling_total)

        records.append(record)

    df = pd.DataFrame(records)

    if df.empty:
        raise DVADataError("No vehicle test data found in file")

    logger.info(f"Parsed {len(df)} months of vehicle test data ({df['year'].min()}-{df['year'].max()})")
    return df


def parse_driver_tests(file_path: Union[str, Path]) -> pd.DataFrame:
    """Parse DVA driver tests data from Excel file.

    Extracts driver tests conducted from Table 2.1.

    Args:
        file_path: Path to the DVA Excel file

    Returns:
        DataFrame with columns:
            - date: datetime (first day of month)
            - year: int
            - month: str (month name)
            - tests_conducted: int (driver tests conducted)
            - rolling_12_month_total: int (optional, rolling 12-month sum)

    Example:
        >>> df = parse_driver_tests("dva-monthly-tests-december-2025.xlsx")
        >>> print(df.tail())
    """
    from openpyxl import load_workbook

    logger.info(f"Parsing driver tests from: {file_path}")

    wb = load_workbook(file_path, data_only=True)

    # Find the driver tests sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "2_1" in name and "Driver" in name:
            sheet_name = name
            break

    if not sheet_name:
        raise DVADataError("Could not find driver tests sheet (2_1)")

    sheet = wb[sheet_name]

    # Parse data rows (skip header rows 1-3)
    records = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        date_str = row[0]
        tests = row[1]
        rolling_total = row[3] if len(row) > 3 else None

        if not date_str or not tests:
            continue

        date = _parse_month_year(str(date_str))
        if not date:
            continue

        try:
            tests_int = int(tests)
        except (ValueError, TypeError):
            continue

        record = {
            "date": date,
            "year": date.year,
            "month": date.strftime("%B"),
            "tests_conducted": tests_int,
        }

        if rolling_total:
            with contextlib.suppress(ValueError, TypeError):
                record["rolling_12_month_total"] = int(rolling_total)

        records.append(record)

    df = pd.DataFrame(records)

    if df.empty:
        raise DVADataError("No driver test data found in file")

    logger.info(f"Parsed {len(df)} months of driver test data ({df['year'].min()}-{df['year'].max()})")
    return df


def parse_theory_tests(file_path: Union[str, Path]) -> pd.DataFrame:
    """Parse DVA theory tests data from Excel file.

    Extracts theory tests conducted from Table 3.1.

    Args:
        file_path: Path to the DVA Excel file

    Returns:
        DataFrame with columns:
            - date: datetime (first day of month)
            - year: int
            - month: str (month name)
            - tests_conducted: int (theory tests conducted)
            - rolling_12_month_total: int (optional, rolling 12-month sum)

    Example:
        >>> df = parse_theory_tests("dva-monthly-tests-december-2025.xlsx")
        >>> print(df.tail())
    """
    from openpyxl import load_workbook

    logger.info(f"Parsing theory tests from: {file_path}")

    wb = load_workbook(file_path, data_only=True)

    # Find the theory tests sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "3_1" in name and "Theory" in name:
            sheet_name = name
            break

    if not sheet_name:
        raise DVADataError("Could not find theory tests sheet (3_1)")

    sheet = wb[sheet_name]

    # Parse data rows (skip header rows 1-3)
    records = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        date_str = row[0]
        tests = row[1]
        rolling_total = row[3] if len(row) > 3 else None

        if not date_str or not tests:
            continue

        date = _parse_month_year(str(date_str))
        if not date:
            continue

        try:
            tests_int = int(tests)
        except (ValueError, TypeError):
            continue

        record = {
            "date": date,
            "year": date.year,
            "month": date.strftime("%B"),
            "tests_conducted": tests_int,
        }

        if rolling_total:
            with contextlib.suppress(ValueError, TypeError):
                record["rolling_12_month_total"] = int(rolling_total)

        records.append(record)

    df = pd.DataFrame(records)

    if df.empty:
        raise DVADataError("No theory test data found in file")

    logger.info(f"Parsed {len(df)} months of theory test data ({df['year'].min()}-{df['year'].max()})")
    return df


# ============================================================================
# Main API Functions
# ============================================================================


def get_latest_vehicle_tests(force_refresh: bool = False) -> pd.DataFrame:
    """Get the latest vehicle test statistics.

    Downloads and parses the most recent DVA monthly tests publication.
    Results are cached for 7 days unless force_refresh=True.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with monthly vehicle test data

    Example:
        >>> df = get_latest_vehicle_tests()
        >>> print(f"Latest month: {df.iloc[-1]['month']} {df.iloc[-1]['year']}")
        >>> print(f"Tests conducted: {df.iloc[-1]['tests_conducted']:,}")
    """
    excel_url, _, _ = get_latest_dva_publication_url()
    file_path = _download_file(excel_url, cache_ttl_hours=168, force_refresh=force_refresh)
    return parse_vehicle_tests(file_path)


def get_latest_driver_tests(force_refresh: bool = False) -> pd.DataFrame:
    """Get the latest driver test statistics.

    Downloads and parses the most recent DVA monthly tests publication.
    Results are cached for 7 days unless force_refresh=True.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with monthly driver test data

    Example:
        >>> df = get_latest_driver_tests()
        >>> print(f"Latest month: {df.iloc[-1]['month']} {df.iloc[-1]['year']}")
        >>> print(f"Tests conducted: {df.iloc[-1]['tests_conducted']:,}")
    """
    excel_url, _, _ = get_latest_dva_publication_url()
    file_path = _download_file(excel_url, cache_ttl_hours=168, force_refresh=force_refresh)
    return parse_driver_tests(file_path)


def get_latest_theory_tests(force_refresh: bool = False) -> pd.DataFrame:
    """Get the latest theory test statistics.

    Downloads and parses the most recent DVA monthly tests publication.
    Results are cached for 7 days unless force_refresh=True.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with monthly theory test data

    Example:
        >>> df = get_latest_theory_tests()
        >>> print(f"Latest month: {df.iloc[-1]['month']} {df.iloc[-1]['year']}")
        >>> print(f"Tests conducted: {df.iloc[-1]['tests_conducted']:,}")
    """
    excel_url, _, _ = get_latest_dva_publication_url()
    file_path = _download_file(excel_url, cache_ttl_hours=168, force_refresh=force_refresh)
    return parse_theory_tests(file_path)


def get_latest_all_tests(force_refresh: bool = False) -> dict[str, pd.DataFrame]:
    """Get all test types (vehicle, driver, theory) from the latest publication.

    Downloads the file once and parses all three test types.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        Dictionary with keys 'vehicle', 'driver', 'theory' containing DataFrames

    Example:
        >>> data = get_latest_all_tests()
        >>> print(f"Vehicle tests: {len(data['vehicle'])} months")
        >>> print(f"Driver tests: {len(data['driver'])} months")
        >>> print(f"Theory tests: {len(data['theory'])} months")
    """
    excel_url, _, _ = get_latest_dva_publication_url()
    file_path = _download_file(excel_url, cache_ttl_hours=168, force_refresh=force_refresh)

    return {
        "vehicle": parse_vehicle_tests(file_path),
        "driver": parse_driver_tests(file_path),
        "theory": parse_theory_tests(file_path),
    }


# ============================================================================
# Helper Functions for Analysis
# ============================================================================


def get_tests_by_year(df: pd.DataFrame, year: int) -> pd.DataFrame:
    """Filter test data for a specific year.

    Args:
        df: Test statistics DataFrame
        year: Year to filter for

    Returns:
        DataFrame with only the specified year's data

    Example:
        >>> df = get_latest_vehicle_tests()
        >>> df_2024 = get_tests_by_year(df, 2024)
        >>> print(f"2024 total: {df_2024['tests_conducted'].sum():,}")
    """
    return df[df["year"] == year].reset_index(drop=True)


def get_tests_by_month(df: pd.DataFrame, month: str, year: int) -> pd.DataFrame:
    """Get test data for a specific month and year.

    Args:
        df: Test statistics DataFrame
        month: Month name (e.g., 'January', 'December')
        year: Year

    Returns:
        DataFrame with single row for the specified month

    Example:
        >>> df = get_latest_vehicle_tests()
        >>> dec_2025 = get_tests_by_month(df, 'December', 2025)
        >>> print(f"December 2025: {dec_2025['tests_conducted'].values[0]:,}")
    """
    return df[(df["month"] == month) & (df["year"] == year)].reset_index(drop=True)


def calculate_growth_rates(df: pd.DataFrame, periods: int = 12) -> pd.DataFrame:
    """Calculate year-on-year growth rates for test statistics.

    Args:
        df: Test statistics DataFrame
        periods: Number of months for comparison (default: 12 for YoY)

    Returns:
        DataFrame with additional column:
            - yoy_growth: Percentage change vs same month previous year

    Example:
        >>> df = get_latest_vehicle_tests()
        >>> df_growth = calculate_growth_rates(df)
        >>> print(df_growth[['date', 'tests_conducted', 'yoy_growth']].tail())
    """
    result = df.copy()
    result["yoy_growth"] = result["tests_conducted"].pct_change(periods=periods) * 100
    return result


def get_summary_statistics(df: pd.DataFrame, start_year: Optional[int] = None, end_year: Optional[int] = None) -> dict:
    """Calculate summary statistics for test data.

    Args:
        df: Test statistics DataFrame
        start_year: Optional start year for summary
        end_year: Optional end year for summary

    Returns:
        Dictionary with summary statistics:
            - period: Time period covered
            - total_tests: Total tests in period
            - monthly_mean: Average monthly tests
            - monthly_min: Minimum monthly tests
            - monthly_max: Maximum monthly tests
            - months_count: Number of months included

    Example:
        >>> df = get_latest_vehicle_tests()
        >>> stats = get_summary_statistics(df, start_year=2020)
        >>> print(f"Average monthly tests since 2020: {stats['monthly_mean']:,.0f}")
    """
    filtered = df.copy()

    if start_year:
        filtered = filtered[filtered["year"] >= start_year]
    if end_year:
        filtered = filtered[filtered["year"] <= end_year]

    return {
        "period": f"{filtered['year'].min()}-{filtered['year'].max()}",
        "total_tests": int(filtered["tests_conducted"].sum()),
        "monthly_mean": float(filtered["tests_conducted"].mean()),
        "monthly_min": int(filtered["tests_conducted"].min()),
        "monthly_max": int(filtered["tests_conducted"].max()),
        "months_count": len(filtered),
    }


def validate_dva_test_data(df: pd.DataFrame) -> bool:  # pragma: no cover
    """Validate DVA test statistics data integrity.

    Args:
        df: DataFrame from DVA test functions (vehicle, driver, or theory tests)

    Returns:
        True if validation passes, False otherwise
    """
    if df.empty:
        logging.warning("DVA test data is empty")
        return False

    required_cols = {"month", "tests_conducted"}
    if not required_cols.issubset(df.columns):
        missing = required_cols - set(df.columns)
        logging.warning(f"Missing required DVA columns: {missing}")
        return False

    # Check for non-negative test counts
    if (df["tests_conducted"] < 0).any():
        logging.warning("Found negative test counts in DVA data")
        return False

    # Check for reasonable monthly test volumes
    # Vehicle tests typically range from 40,000 to 100,000 per month in NI
    if df["tests_conducted"].max() > 200000:  # Allow for variation but catch obvious errors
        logging.warning("Unreasonably high test counts found")
        return False

    return True
