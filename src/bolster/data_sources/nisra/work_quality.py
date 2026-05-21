"""NISRA Work Quality in Northern Ireland Module.

This module provides access to Work Quality statistics for Northern Ireland,
covering seventeen indicators of job quality for employees aged 18 and over.

The publication draws on two main sources:
- Labour Force Survey (LFS): subjective and contractual indicators
- Annual Survey of Hours and Earnings (ASHE): earnings-based indicator

Work quality indicators covered:
    1.  Real Living Wage: proportion earning above the Real Living Wage
    2.  Zero-hours contracts: proportion not on zero-hours contracts
    3.  Secure employment: permanent or preferred temporary contracts
    4.  Job satisfaction: satisfied or very satisfied with job
    5.  Career progression: opportunities for career progression
    6.  Decision making: involvement in workplace decisions
    7.  Meaningful work: performing work perceived as meaningful
    8.  Flexible work: access to flexible working arrangements
    9.  Line manager support: positive line manager relationship
    10. Bullying & harassment: not bullied or harassed at work
    11. Skills match: has the skills required for current duties
    12. Workplace accidents: no accident reported at work
    13. Underemployment: employees who are underemployed
    14. Overemployment: employees who are overemployed
    15. Trade union membership: member of a trade union
    16. Training participation: participated in training (last 13 weeks)
    17. Overtime working: reported working overtime

Data Source: NISRA Labour Market and Social Welfare statistics.
    Index page: https://www.nisra.gov.uk/statistics/labour-market-and-social-welfare/work-quality

Update Frequency: Annual (published in spring for the preceding calendar year).

Data Coverage:
    - Most indicators: 2020 to present (calendar year)
    - Line manager support, bullying & harassment, skills match: 2022 to present
      (reported as rolling 12-month periods, e.g., "July 2024 to June 2025")

Breakdowns available per indicator:
    - Overall (NI)
    - By sex (Male / Female)
    - By age group (18-39 / 40 and over)
    - By deprivation quintile (most deprived / least deprived)
    - By skill level (low skilled / high skilled)

Examples:
    >>> from bolster.data_sources.nisra import work_quality
    >>> df = work_quality.get_latest_work_quality()
    >>> 'indicator' in df.columns
    True
    >>> 'value' in df.columns
    True
    >>> url, year = work_quality.get_work_quality_publication_url()
    >>> url.startswith('https://')
    True
    >>> year >= 2025
    True

Publication Details:
    - Frequency: Annual
    - Published by: NISRA
    - Coverage: Employees aged 18 and over in Northern Ireland
    - Source surveys: Labour Force Survey (LFS) and ASHE
"""

import logging
import re
from pathlib import Path

import openpyxl
import pandas as pd

from bolster.utils.web import session

from ._base import NISRADataNotFoundError, NISRAValidationError, download_file

logger = logging.getLogger(__name__)

# Stable index page listing all work quality publications
WORK_QUALITY_INDEX_URL = "https://www.nisra.gov.uk/statistics/labour-market-and-social-welfare/work-quality"
NISRA_BASE_URL = "https://www.nisra.gov.uk"

# All 17 indicator sheet names and their canonical indicator names
_INDICATOR_SHEETS: dict[str, str] = {
    "Table_1.1": "real_living_wage",
    "Table_1.2": "no_zero_hours_contract",
    "Table_1.3": "secure_employment",
    "Table_1.4": "job_satisfaction",
    "Table_1.5": "career_progression",
    "Table_1.6": "decision_making",
    "Table_1.7": "meaningful_work",
    "Table_1.8": "flexible_work",
    "Table_1.9": "line_manager_support",
    "Table_1.10": "no_bullying_harassment",
    "Table_1.11": "skills_match",
    "Table_1.12": "no_workplace_accident",
    "Table_1.13": "underemployed",
    "Table_1.14": "overemployed",
    "Table_1.15": "trade_union_member",
    "Table_1.16": "recent_training",
    "Table_1.17": "working_overtime",
}

# Row-type label for the overall NI figure (sub-table 'a')
_OVERALL_TABLE_SUFFIX = "a"


def get_work_quality_publication_url() -> tuple[str, int]:
    """Get the URL of the latest Work Quality Excel file and publication year.

    Scrapes the NISRA work quality index page to discover the most recent
    publication link, then follows to the publication page to find the Excel file.

    Returns:
        Tuple of (excel_url, year) e.g.
        ("https://www.nisra.gov.uk/system/files/.../work-quality-2025.xlsx", 2025)

    Raises:
        NISRADataNotFoundError: If unable to find the publication or Excel file

    Example:
        >>> url, year = get_work_quality_publication_url()
        >>> url.startswith('https://')
        True
        >>> year >= 2025
        True
    """
    from bs4 import BeautifulSoup

    logger.info("Fetching Work Quality publication URL from index page...")

    try:
        response = session.get(WORK_QUALITY_INDEX_URL, timeout=30)
        response.raise_for_status()
    except Exception as e:
        raise NISRADataNotFoundError(f"Failed to fetch Work Quality index page: {e}") from e

    soup = BeautifulSoup(response.content, "html.parser")

    # Collect all links whose text matches "Work Quality in Northern Ireland YYYY"
    publications: list[tuple[int, str]] = []
    for a_tag in soup.find_all("a", href=True):
        text = a_tag.get_text(strip=True)
        href = a_tag["href"]
        match = re.search(r"Work Quality in Northern Ireland\s+(\d{4})", text, re.IGNORECASE)
        if match:
            year = int(match.group(1))
            pub_url = href if href.startswith("http") else f"{NISRA_BASE_URL}{href}"
            publications.append((year, pub_url))
            logger.debug(f"Found publication: {year} → {pub_url}")

    if not publications:
        raise NISRADataNotFoundError(
            f"Could not find any Work Quality publication links on the index page. Check: {WORK_QUALITY_INDEX_URL}"
        )

    # Use the most recent publication
    publications.sort(key=lambda x: x[0], reverse=True)
    latest_year, pub_url = publications[0]
    logger.info(f"Latest Work Quality publication: {latest_year} at {pub_url}")

    # Follow to the publication page and find the Excel file
    try:
        pub_response = session.get(pub_url, timeout=30)
        pub_response.raise_for_status()
    except Exception as e:
        raise NISRADataNotFoundError(f"Failed to fetch Work Quality publication page {pub_url}: {e}") from e

    pub_soup = BeautifulSoup(pub_response.content, "html.parser")

    excel_url: str | None = None
    for a_tag in pub_soup.find_all("a", href=True):
        href = a_tag["href"]
        if ".xlsx" in href.lower() and "work-quality" in href.lower():
            excel_url = href if href.startswith("http") else f"{NISRA_BASE_URL}{href}"
            break

    if excel_url is None:
        # Fallback: any .xlsx link on the publication page
        for a_tag in pub_soup.find_all("a", href=True):
            href = a_tag["href"]
            if ".xlsx" in href.lower():
                excel_url = href if href.startswith("http") else f"{NISRA_BASE_URL}{href}"
                logger.warning(f"Using fallback Excel link (no 'work-quality' in path): {excel_url}")
                break

    if excel_url is None:
        raise NISRADataNotFoundError(f"Could not find Excel file on Work Quality publication page: {pub_url}")

    logger.info(f"Found Work Quality Excel URL: {excel_url}")
    return excel_url, latest_year


def _normalise_year_label(raw: str | int | float) -> str | None:
    """Normalise a raw year cell value to a compact string label.

    Accepts integer years (2020), float years (2025.0), and period strings
    ("July 2024 to June 2025").  Returns None for blank / unsupported values.

    Args:
        raw: The raw cell value from the Excel sheet.

    Returns:
        Normalised label (e.g. "2025" or "July 2024 to June 2025") or None.

    Example:
        >>> _normalise_year_label(2025)
        '2025'
        >>> _normalise_year_label(2025.0)
        '2025'
        >>> _normalise_year_label('July 2024 to June 2025')
        'July 2024 to June 2025'
        >>> _normalise_year_label(None) is None
        True
    """
    if raw is None:
        return None
    if isinstance(raw, int | float):
        try:
            return str(int(raw))
        except (ValueError, TypeError):
            return None
    raw_str = str(raw).strip()
    if not raw_str or raw_str.lower() in ("none", "nan", "year"):
        return None
    return raw_str


def _extract_year_int(label: str) -> int:
    """Extract the latest calendar year from a year label.

    For plain year strings ("2025") returns 2025.
    For period strings ("July 2024 to June 2025") returns 2025 (end year).

    Args:
        label: Normalised year label as returned by _normalise_year_label.

    Returns:
        Integer calendar year.

    Example:
        >>> _extract_year_int('2025')
        2025
        >>> _extract_year_int('July 2024 to June 2025')
        2025
    """
    # Find all 4-digit years
    years_found = [int(m) for m in re.findall(r"\b(\d{4})\b", label)]
    if not years_found:
        raise ValueError(f"No year found in label: {label!r}")
    return max(years_found)


def _parse_indicator_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    indicator: str,
) -> list[dict]:
    """Parse the overall NI row (sub-table 'a') from one indicator sheet.

    Each sheet contains multiple sub-tables stacked vertically.  Sub-table 'a'
    always holds the overall NI proportions; we only extract that for the
    canonical long-format output.

    Args:
        wb: Open openpyxl workbook (read_only=True, data_only=True).
        sheet_name: Name of the worksheet (e.g. "Table_1.1").
        indicator: Canonical snake_case indicator name.

    Returns:
        List of dicts with keys: indicator, year_label, year, value, geography.
    """
    ws = wb[sheet_name]
    rows: list[dict] = []
    in_overall_table = False

    for row_vals in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row_vals):
            continue

        cell0 = str(row_vals[0]).strip() if row_vals[0] is not None else ""

        # Detect the start of sub-table 'a' (overall NI)
        # Pattern: "Table X.Xa: ..."  (letter 'a' suffix)
        if re.match(r"Table\s+\d+\.\d+a\s*:", cell0, re.IGNORECASE):
            in_overall_table = True
            continue

        # Detect start of any OTHER sub-table (b, c, d, e) — stop reading
        if re.match(r"Table\s+\d+\.\d+[b-zA-Z]\s*:", cell0, re.IGNORECASE):
            if in_overall_table:
                break  # done with sub-table 'a'
            continue

        if not in_overall_table:
            continue

        # Skip the header row ("Year", "Proportion …")
        if cell0.lower() == "year":
            continue

        # Try to parse a data row: col0 = year, col1 = value
        year_label = _normalise_year_label(row_vals[0])
        if year_label is None:
            continue

        raw_val = row_vals[1] if len(row_vals) > 1 else None
        # Skip suppressed values marked [x]
        if isinstance(raw_val, str) and "[x]" in raw_val:
            logger.debug(f"{indicator} {year_label}: suppressed value [x]")
            continue
        if raw_val is None:
            continue

        try:
            value = float(raw_val)
        except (ValueError, TypeError):
            logger.debug(f"{indicator} {year_label}: could not convert {raw_val!r} to float")
            continue

        try:
            year_int = _extract_year_int(year_label)
        except ValueError:
            logger.debug(f"{indicator}: could not extract year from {year_label!r}")
            continue

        rows.append(
            {
                "indicator": indicator,
                "year_label": year_label,
                "year": year_int,
                "value": value,
                "geography": "Northern Ireland",
            }
        )

    return rows


def parse_work_quality_file(file_path: str | Path, year: int) -> pd.DataFrame:
    """Parse a Work Quality Excel file into a tidy long-format DataFrame.

    Extracts the overall NI proportion for each of the 17 work quality
    indicators.  One row per (indicator, year_label) combination.

    Args:
        file_path: Path to the Work Quality Excel (.xlsx) file.
        year: Publication year (e.g. 2025 for the 2025 publication).

    Returns:
        DataFrame with columns:
            - indicator: str  — snake_case indicator name (e.g. "job_satisfaction")
            - year_label: str — year as printed in the source (e.g. "2025" or
              "July 2024 to June 2025" for rolling-period indicators)
            - year: int       — latest calendar year in the period
            - value: float    — proportion (percentage, 0-100)
            - geography: str  — always "Northern Ireland" for this dataset

    Raises:
        NISRADataNotFoundError: If the file cannot be opened or no data is found.

    Example:
        >>> url, yr = get_work_quality_publication_url()
        >>> path = download_file(url, cache_ttl_hours=90 * 24)
        >>> df = parse_work_quality_file(path, yr)
        >>> sorted(df.columns.tolist())
        ['geography', 'indicator', 'value', 'year', 'year_label']
        >>> len(df) > 0
        True
    """
    logger.info(f"Parsing Work Quality file: {file_path} (publication year {year})")
    file_path = Path(file_path)

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        raise NISRADataNotFoundError(f"Cannot open Work Quality file {file_path}: {e}") from e

    all_rows: list[dict] = []
    for sheet_name, indicator in _INDICATOR_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet {sheet_name!r} not found in workbook — skipping {indicator}")
            continue
        rows = _parse_indicator_sheet(wb, sheet_name, indicator)
        if not rows:
            logger.warning(f"No data rows extracted from {sheet_name} ({indicator})")
        all_rows.extend(rows)

    wb.close()

    if not all_rows:
        raise NISRADataNotFoundError(f"No data extracted from Work Quality file: {file_path}")

    df = pd.DataFrame(all_rows)
    df = df.sort_values(["indicator", "year", "year_label"]).reset_index(drop=True)
    logger.info(
        f"Parsed {len(df)} rows across {df['indicator'].nunique()} indicators "
        f"(years {df['year'].min()}–{df['year'].max()})"
    )
    return df


def get_latest_work_quality(force_refresh: bool = False) -> pd.DataFrame:
    """Download and parse the latest Work Quality publication.

    Automatically discovers the current publication URL by scraping the NISRA
    index page, downloads the Excel file (caching for 90 days), and returns a
    tidy long-format DataFrame.

    Args:
        force_refresh: If True, bypass cache and re-download the file.

    Returns:
        DataFrame with columns:
            - indicator: str  (e.g. "job_satisfaction")
            - year_label: str (e.g. "2025" or "July 2024 to June 2025")
            - year: int
            - value: float    (percentage, 0-100)
            - geography: str  ("Northern Ireland")

    Raises:
        NISRADataNotFoundError: If the publication or file cannot be found.

    Example:
        >>> df = get_latest_work_quality()
        >>> 'indicator' in df.columns
        True
        >>> df['year'].max() >= 2025
        True
    """
    excel_url, year = get_work_quality_publication_url()
    file_path = download_file(excel_url, cache_ttl_hours=90 * 24, force_refresh=force_refresh)
    return parse_work_quality_file(file_path, year)


def validate_work_quality_data(df: pd.DataFrame) -> bool:
    """Validate a Work Quality DataFrame for expected structure and values.

    Args:
        df: DataFrame returned by parse_work_quality_file or get_latest_work_quality.

    Returns:
        True if all checks pass.

    Raises:
        NISRAValidationError: If any validation check fails.

    Example:
        >>> df = get_latest_work_quality()
        >>> validate_work_quality_data(df)
        True
    """
    if df.empty:
        raise NISRAValidationError("DataFrame is empty")

    required_columns = {"indicator", "year_label", "year", "value", "geography"}
    missing = required_columns - set(df.columns)
    if missing:
        raise NISRAValidationError(f"Missing required columns: {missing}")

    if df["indicator"].nunique() < 10:
        raise NISRAValidationError(f"Expected at least 10 indicators, found {df['indicator'].nunique()}")

    if df["year"].max() < 2024:
        raise NISRAValidationError(f"No recent data found — latest year is {df['year'].max()}")

    out_of_range = df[(df["value"] < 0) | (df["value"] > 100)]
    if not out_of_range.empty:
        raise NISRAValidationError(f"Values outside valid percentage range [0, 100]:\n{out_of_range}")

    if df["geography"].nunique() != 1 or df["geography"].iloc[0] != "Northern Ireland":
        raise NISRAValidationError(f"Unexpected geography values: {df['geography'].unique().tolist()}")

    logger.info(
        f"Work Quality data validation passed: {len(df)} rows, "
        f"{df['indicator'].nunique()} indicators, years {df['year'].min()}–{df['year'].max()}"
    )
    return True
