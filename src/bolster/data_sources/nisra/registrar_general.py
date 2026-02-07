"""NISRA Registrar General Quarterly Tables Data Source.

Provides access to quarterly vital statistics for Northern Ireland, including:
- Quarterly births and stillbirths (from Q1 2009)
- Quarterly deaths, marriages, and civil partnerships
- LGD-level breakdowns for the current quarter
- Birth and death rates per 1,000 population

These quarterly tables provide higher-level aggregated statistics compared to the
monthly data, with additional metrics like stillbirths, infant deaths, and rates.

Data Source:
    **Publications page**: https://www.nisra.gov.uk/statistics/births-deaths-and-marriages/registrar-general-quarterly-report

    The quarterly tables are published approximately 6 weeks after the end of each quarter.
    Historical data is available from Q1 2009.

Update Frequency: Quarterly (February, May, August, November)
Geographic Coverage: Northern Ireland (with LGD breakdowns)

Example:
    >>> from bolster.data_sources.nisra import registrar_general
    >>> # Get all quarterly vital statistics
    >>> data = registrar_general.get_quarterly_vital_statistics()
    >>> print(data['births'].tail())

    >>> # Get quarterly births only
    >>> births = registrar_general.get_quarterly_births()
    >>> print(f"Q1 2024 births: {births[(births['year']==2024) & (births['quarter']==1)]['total_births'].values[0]}")

    >>> # Get LGD-level statistics
    >>> lgd_df = registrar_general.get_lgd_statistics()
    >>> print(lgd_df)

    >>> # Cross-validate with monthly data
    >>> report = registrar_general.get_validation_report()
    >>> print(report['births_validation'])
"""

import logging
import re
from pathlib import Path
from typing import Dict, Optional, Tuple, Union

import pandas as pd
from openpyxl import load_workbook

from bolster.utils.web import session

from ._base import (
    NISRADataNotFoundError,
    NISRAValidationError,
    download_file,
    safe_float,
    safe_int,
)

logger = logging.getLogger(__name__)

# Base URL for registrar general statistics
REGISTRAR_GENERAL_BASE_URL = (
    "https://www.nisra.gov.uk/statistics/births-deaths-and-marriages/registrar-general-quarterly-report"
)

# List of 11 Local Government Districts in Northern Ireland
NI_LGDS = [
    "Antrim and Newtownabbey",
    "Ards and North Down",
    "Armagh City, Banbridge and Craigavon",
    "Belfast",
    "Causeway Coast and Glens",
    "Derry City and Strabane",
    "Fermanagh and Omagh",
    "Lisburn and Castlereagh",
    "Mid and East Antrim",
    "Mid Ulster",
    "Newry, Mourne and Down",
]


def get_latest_publication_url() -> Tuple[str, str, int, int]:
    """Scrape NISRA to find the latest Registrar General Quarterly Tables file.

    Navigates the publication structure:
    1. Scrapes the Registrar General Quarterly Report page
    2. Finds the latest quarterly tables publication link
    3. Extracts the Excel file URL

    Returns:
        Tuple of (excel_url, publication_page_url, year, quarter)

    Raises:
        NISRADataNotFoundError: If publication or file not found
    """
    from bs4 import BeautifulSoup

    mother_page = REGISTRAR_GENERAL_BASE_URL

    try:
        response = session.get(mother_page, timeout=30)
        response.raise_for_status()
    except Exception as e:
        raise NISRADataNotFoundError(f"Failed to fetch Registrar General page: {e}")

    soup = BeautifulSoup(response.content, "html.parser")

    # Find the latest quarterly tables publication link
    # Pattern: "Registrar General Quarterly Tables, Quarter X YYYY"
    pub_link = None
    year = None
    quarter = None

    for link in soup.find_all("a", href=True):
        link_text = link.get_text(strip=True)

        # Match pattern like "Registrar General Quarterly Tables, Quarter 3 2025"
        # or "Registrar General Quarterly Tables Quarter 3 2025"
        if "Quarterly Tables" in link_text and "Quarter" in link_text:
            href = link["href"]

            if href.startswith("/"):
                href = f"https://www.nisra.gov.uk{href}"

            # Extract quarter and year from link text
            match = re.search(r"Quarter\s*(\d)\s*(\d{4})", link_text)
            if match:
                quarter = int(match.group(1))
                year = int(match.group(2))

            pub_link = href
            logger.info(f"Found Registrar General Quarterly Tables: Q{quarter} {year}")
            break

    if not pub_link:
        raise NISRADataNotFoundError("Could not find Registrar General Quarterly Tables publication")

    # Scrape the publication page for Excel file
    try:
        pub_response = session.get(pub_link, timeout=30)
        pub_response.raise_for_status()
    except Exception as e:
        raise NISRADataNotFoundError(f"Failed to fetch publication page: {e}")

    pub_soup = BeautifulSoup(pub_response.content, "html.parser")

    # Find Excel file link
    # Pattern: "Quarter 3 2025 Tables.xlsx" or similar
    excel_url = None

    for link in pub_soup.find_all("a", href=True):
        href = link["href"]
        link_text = link.get_text(strip=True)

        if href.endswith(".xlsx") and "Tables" in href:
            if href.startswith("/"):
                href = f"https://www.nisra.gov.uk{href}"

            excel_url = href
            logger.info(f"Found Excel file: {href}")
            break

    if not excel_url:
        raise NISRADataNotFoundError("Could not find Excel file on publication page")

    return excel_url, pub_link, year, quarter


def parse_quarterly_births(sheet) -> pd.DataFrame:
    """Parse Table 1a - Quarterly Births from the Excel sheet.

    The births table contains quarterly data from Q1 2009 with:
    - Total births, birth rate, stillbirths
    - Births outside marriage (count and %)
    - Teenage births (count and %)
    - Births to mothers 30+ (count and %)

    Args:
        sheet: openpyxl worksheet object for Table 1a

    Returns:
        DataFrame with quarterly births data
    """
    records = []

    # Find header row - look for "Registration Year" or similar
    header_row = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        row_str = [str(cell).lower() if cell else "" for cell in row]
        # Look for registration year in the row
        if any("registration year" in cell for cell in row_str):
            header_row = row_idx
            break

    if not header_row:
        logger.warning("Could not find header row in births table")
        return pd.DataFrame()

    logger.debug(f"Found births header row at row {header_row}")

    # Parse data rows - columns are fixed positions based on actual file structure
    # Col 0: Registration Year
    # Col 1: Registration Quarter
    # Col 2: Total Live Births
    # Col 3: Live Births Rate
    # Col 4: Births Outside Marriage
    # Col 5: % Births Outside Marriage
    # Col 6: Births To Teenage Mothers
    # Col 7: % Births To Teenage Mothers
    # Col 8: Births to Mothers Aged 30+
    # Col 9: % Births to Mothers Aged 30+
    # Col 10: Stillbirths

    for row in sheet.iter_rows(min_row=header_row + 1, max_row=200, values_only=True):
        year = safe_int(row[0])
        quarter = safe_int(row[1])

        if year is None or quarter is None:
            continue

        # Skip if year is before 2009
        if year < 2009:
            continue

        total_births = safe_int(row[2])
        if total_births is None:
            continue

        record = {
            "year": year,
            "quarter": quarter,
            "total_births": total_births,
            "birth_rate": safe_float(row[3]),
            "births_outside_marriage": safe_int(row[4]),
            "pct_outside_marriage": safe_float(row[5]),
            "teenage_births": safe_int(row[6]),
            "pct_teenage": safe_float(row[7]),
            "births_30_plus": safe_int(row[8]),
            "pct_30_plus": safe_float(row[9]),
            "stillbirths": safe_int(row[10]) if len(row) > 10 else None,
        }

        records.append(record)

    df = pd.DataFrame(records)

    if len(df) > 0:
        # Ensure year and quarter are integers
        df["year"] = df["year"].astype(int)
        df["quarter"] = df["quarter"].astype(int)

        # Create date column (first day of quarter)
        df["date"] = pd.to_datetime(
            df.apply(
                lambda r: f"{int(r['year'])}-{(int(r['quarter']) - 1) * 3 + 1:02d}-01",
                axis=1,
            )
        )
        df = df.sort_values(["year", "quarter"]).reset_index(drop=True)

    logger.info(f"Parsed {len(df)} quarterly birth records")
    return df


def parse_quarterly_deaths(sheet) -> pd.DataFrame:
    """Parse Table 1b - Quarterly Deaths from the Excel sheet.

    The deaths table contains quarterly data with:
    - Total deaths and death rate
    - Marriages and civil partnerships
    - Infant deaths

    Args:
        sheet: openpyxl worksheet object for Table 1b

    Returns:
        DataFrame with quarterly deaths/marriages data
    """
    records = []

    # Find header row - look for "Registration Year" or similar
    header_row = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        row_str = [str(cell).lower() if cell else "" for cell in row]
        if any("registration year" in cell for cell in row_str):
            header_row = row_idx
            break

    if not header_row:
        logger.warning("Could not find header row in deaths table")
        return pd.DataFrame()

    logger.debug(f"Found deaths header row at row {header_row}")

    # Parse data rows - columns are fixed positions based on actual file structure
    # Col 0: Registration Year
    # Col 1: Registration Quarter
    # Col 2: Number of Deaths
    # Col 3: Death Rate
    # Col 4: Infant Deaths
    # Col 5: Total Marriages
    # Col 6: Civil Marriages (% of total)
    # Col 7: Same-Sex Marriage (% of total)
    # Col 8: Civil Partnerships

    for row in sheet.iter_rows(min_row=header_row + 1, max_row=200, values_only=True):
        year = safe_int(row[0])
        quarter = safe_int(row[1])

        if year is None or quarter is None:
            continue

        if year < 2009:
            continue

        deaths = safe_int(row[2])
        if deaths is None:
            continue

        record = {
            "year": year,
            "quarter": quarter,
            "deaths": deaths,
            "death_rate": safe_float(row[3]),
            "infant_deaths": safe_int(row[4]),
            "marriages": safe_int(row[5]),
            "civil_marriage_pct": safe_float(row[6]),
            "same_sex_pct": safe_float(row[7]) if row[7] != "-" else None,
            "civil_partnerships": safe_int(row[8]) if len(row) > 8 else None,
        }

        records.append(record)

    df = pd.DataFrame(records)

    if len(df) > 0:
        # Ensure year and quarter are integers
        df["year"] = df["year"].astype(int)
        df["quarter"] = df["quarter"].astype(int)

        df["date"] = pd.to_datetime(
            df.apply(
                lambda r: f"{int(r['year'])}-{(int(r['quarter']) - 1) * 3 + 1:02d}-01",
                axis=1,
            )
        )
        df = df.sort_values(["year", "quarter"]).reset_index(drop=True)

    logger.info(f"Parsed {len(df)} quarterly deaths/marriages records")
    return df


def parse_lgd_statistics(sheet) -> pd.DataFrame:
    """Parse Table 3b - LGD-level statistics from the Excel sheet.

    The LGD table contains current quarter statistics by Local Government District:
    - Population estimate
    - Births and birth rate
    - Deaths and death rate
    - Marriages

    Args:
        sheet: openpyxl worksheet object for Table 3b

    Returns:
        DataFrame with LGD-level statistics
    """
    records = []

    # Find header row - look for "Area" column (first column header)
    header_row = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        row_str = [str(cell).lower() if cell else "" for cell in row]
        # The header row has "area" as first column and "mid-year population" in second
        if "area" in row_str or any("population" in cell for cell in row_str):
            header_row = row_idx
            break

    if not header_row:
        logger.warning("Could not find header row in LGD table")
        return pd.DataFrame()

    logger.debug(f"Found LGD header row at row {header_row}")

    # Parse data rows - columns are fixed positions based on actual file structure
    # Col 0: Area (LGD name)
    # Col 1: Mid-Year population estimate
    # Col 2: Total Births
    # Col 3: Birth Rate
    # Col 4: Number of births outside marriage
    # Col 5: % of births outside marriage
    # Col 6: Stillbirths
    # Col 7: All Deaths
    # Col 8: Death Rate
    # Col 9: Deaths of children under 1 Year
    # Col 10: Deaths due to Cancer
    # Col 11: Deaths due to Covid-19
    # Col 12: Deaths due to Ischaemic Heart Disease
    # Col 13: Marriages

    for row in sheet.iter_rows(min_row=header_row + 1, max_row=30, values_only=True):
        lgd_name = str(row[0]).strip() if row[0] else ""

        # Check if this is a valid LGD row
        is_lgd = any(lgd.lower() in lgd_name.lower() or lgd_name.lower() in lgd.lower() for lgd in NI_LGDS)

        if not is_lgd:
            continue

        record = {
            "lgd": lgd_name,
            "population": safe_int(row[1]),
            "births": safe_int(row[2]),
            "birth_rate": safe_float(row[3]),
            "births_outside_marriage": safe_int(row[4]),
            "pct_outside_marriage": safe_float(row[5]),
            "stillbirths": safe_int(row[6]),
            "deaths": safe_int(row[7]),
            "death_rate": safe_float(row[8]),
            "infant_deaths": safe_int(row[9]),
            "cancer_deaths": safe_int(row[10]),
            "covid_deaths": safe_int(row[11]),
            "heart_disease_deaths": safe_int(row[12]),
            "marriages": safe_int(row[13]) if len(row) > 13 else None,
        }

        records.append(record)

    df = pd.DataFrame(records)

    # Remove any rows with empty LGD names
    if not df.empty and "lgd" in df.columns:
        df = df[df["lgd"].str.strip() != ""]
        df = df.reset_index(drop=True)

    logger.info(f"Parsed {len(df)} LGD records")
    return df


def parse_quarterly_tables(file_path: Union[str, Path]) -> Dict[str, pd.DataFrame]:
    """Parse the Registrar General Quarterly Tables Excel file.

    The file contains multiple tables:
    - Table 1a: Quarterly births and stillbirths
    - Table 1b: Quarterly deaths, marriages, civil partnerships
    - Table 3b: Current quarter by LGD

    Args:
        file_path: Path to the quarterly tables Excel file

    Returns:
        Dict with keys 'births', 'deaths', 'lgd' containing DataFrames
    """
    file_path = Path(file_path)
    wb = load_workbook(file_path, data_only=True)

    results = {}

    # Parse Table 1a (births) - look for exact "Table 1a" match
    births_sheet = None
    for name in wb.sheetnames:
        # Must start with "Table" to avoid matching "Figure 1a"
        if name.lower().startswith("table") and "1a" in name.lower():
            births_sheet = wb[name]
            logger.debug(f"Found births sheet: {name}")
            break

    if births_sheet:
        results["births"] = parse_quarterly_births(births_sheet)
    else:
        logger.warning("Could not find births table (Table 1a)")
        results["births"] = pd.DataFrame()

    # Parse Table 1b (deaths/marriages) - look for exact "Table 1b" match
    deaths_sheet = None
    for name in wb.sheetnames:
        if name.lower().startswith("table") and "1b" in name.lower():
            deaths_sheet = wb[name]
            logger.debug(f"Found deaths sheet: {name}")
            break

    if deaths_sheet:
        results["deaths"] = parse_quarterly_deaths(deaths_sheet)
    else:
        logger.warning("Could not find deaths table (Table 1b)")
        results["deaths"] = pd.DataFrame()

    # Parse Table 3b (LGD statistics) - look for exact "Table 3b" match
    lgd_sheet = None
    for name in wb.sheetnames:
        if name.lower().startswith("table") and "3b" in name.lower():
            lgd_sheet = wb[name]
            logger.debug(f"Found LGD sheet: {name}")
            break

    if lgd_sheet:
        results["lgd"] = parse_lgd_statistics(lgd_sheet)
    else:
        logger.warning("Could not find LGD table (Table 3b)")
        results["lgd"] = pd.DataFrame()

    return results


def get_quarterly_vital_statistics(
    force_refresh: bool = False,
) -> Dict[str, pd.DataFrame]:
    """Get all quarterly vital statistics from the Registrar General Tables.

    Automatically discovers and downloads the most recent quarterly tables
    publication from NISRA.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        Dict with keys:
            - 'births': DataFrame with quarterly births data
            - 'deaths': DataFrame with quarterly deaths/marriages data
            - 'lgd': DataFrame with LGD-level breakdowns

    Raises:
        NISRADataNotFoundError: If latest publication cannot be found

    Example:
        >>> data = get_quarterly_vital_statistics()
        >>> print(data['births'].columns)
        >>> print(data['births'].tail())
    """
    excel_url, pub_url, year, quarter = get_latest_publication_url()

    logger.info(f"Downloading Q{quarter} {year} quarterly tables from: {excel_url}")

    # Cache for 90 days (quarterly data)
    cache_ttl_hours = 90 * 24
    file_path = download_file(excel_url, cache_ttl_hours=cache_ttl_hours, force_refresh=force_refresh)

    return parse_quarterly_tables(file_path)


def get_quarterly_births(force_refresh: bool = False) -> pd.DataFrame:
    """Get quarterly births data.

    Convenience function to get only the births table.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns:
            - year: int (2009+)
            - quarter: int (1-4)
            - date: datetime (first day of quarter)
            - total_births: int
            - birth_rate: float (per 1,000 population)
            - stillbirths: int (if available)

    Example:
        >>> births = get_quarterly_births()
        >>> q1_2024 = births[(births['year']==2024) & (births['quarter']==1)]
        >>> print(f"Q1 2024 births: {q1_2024['total_births'].values[0]}")
    """
    data = get_quarterly_vital_statistics(force_refresh=force_refresh)
    return data["births"]


def get_quarterly_deaths(force_refresh: bool = False) -> pd.DataFrame:
    """Get quarterly deaths and marriages data.

    Convenience function to get only the deaths table.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns:
            - year: int (2009+)
            - quarter: int (1-4)
            - date: datetime
            - deaths: int
            - death_rate: float (per 1,000 population)
            - marriages: int
            - civil_partnerships: int

    Example:
        >>> deaths = get_quarterly_deaths()
        >>> print(deaths[deaths['year'] == 2024])
    """
    data = get_quarterly_vital_statistics(force_refresh=force_refresh)
    return data["deaths"]


def get_lgd_statistics(force_refresh: bool = False) -> pd.DataFrame:
    """Get current quarter LGD-level statistics.

    Convenience function to get the LGD breakdown table.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns:
            - lgd: str (Local Government District name)
            - population: int (mid-year estimate)
            - births: int
            - birth_rate: float
            - deaths: int
            - death_rate: float
            - marriages: int

    Example:
        >>> lgd = get_lgd_statistics()
        >>> print(lgd.sort_values('births', ascending=False))
    """
    data = get_quarterly_vital_statistics(force_refresh=force_refresh)
    return data["lgd"]


def validate_against_monthly_births(
    quarterly_df: pd.DataFrame,
    monthly_df: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """Compare quarterly births totals against aggregated monthly births.

    Cross-validates quarterly data against monthly data to verify consistency.
    Some differences are expected due to timing of registrations.

    Args:
        quarterly_df: Quarterly births DataFrame from get_quarterly_births()
        monthly_df: Monthly births DataFrame from births module (auto-loaded if None)

    Returns:
        DataFrame with columns:
            - year: int
            - quarter: int
            - quarterly_total: int
            - monthly_sum: int
            - difference: int
            - pct_diff: float (percentage difference)

    Example:
        >>> births_q = get_quarterly_births()
        >>> validation = validate_against_monthly_births(births_q)
        >>> print(validation[validation['pct_diff'].abs() > 1])
    """
    if monthly_df is None:
        from . import births

        monthly_data = births.get_latest_births(event_type="registration")
        # Get Persons total for each month
        monthly_df = monthly_data[monthly_data["sex"] == "Persons"].copy()
        monthly_df["year"] = monthly_df["month"].dt.year
        monthly_df["month_num"] = monthly_df["month"].dt.month
        monthly_df["quarter"] = ((monthly_df["month_num"] - 1) // 3) + 1

    # Aggregate monthly to quarterly
    monthly_quarterly = monthly_df.groupby(["year", "quarter"]).agg(monthly_sum=("births", "sum")).reset_index()

    # Merge with quarterly data
    comparison = quarterly_df[["year", "quarter", "total_births"]].merge(
        monthly_quarterly,
        on=["year", "quarter"],
        how="inner",
    )

    comparison = comparison.rename(columns={"total_births": "quarterly_total"})
    comparison["difference"] = comparison["quarterly_total"] - comparison["monthly_sum"]
    comparison["pct_diff"] = (comparison["difference"] / comparison["monthly_sum"] * 100).round(2)

    return comparison.sort_values(["year", "quarter"])


def validate_against_monthly_marriages(
    quarterly_df: pd.DataFrame,
    monthly_df: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """Compare quarterly marriages totals against aggregated monthly marriages.

    Cross-validates quarterly data against monthly data to verify consistency.

    Args:
        quarterly_df: Quarterly deaths DataFrame from get_quarterly_deaths()
        monthly_df: Monthly marriages DataFrame from marriages module

    Returns:
        DataFrame with comparison columns

    Example:
        >>> deaths_q = get_quarterly_deaths()
        >>> validation = validate_against_monthly_marriages(deaths_q)
        >>> print(validation)
    """
    if monthly_df is None:
        from . import marriages

        monthly_df = marriages.get_latest_marriages()
        monthly_df["quarter"] = ((monthly_df["date"].dt.month - 1) // 3) + 1

    # Aggregate monthly to quarterly
    monthly_quarterly = monthly_df.groupby(["year", "quarter"]).agg(monthly_sum=("marriages", "sum")).reset_index()

    # Merge with quarterly data
    comparison = quarterly_df[["year", "quarter", "marriages"]].merge(
        monthly_quarterly,
        on=["year", "quarter"],
        how="inner",
    )

    comparison = comparison.rename(columns={"marriages": "quarterly_total"})
    comparison["difference"] = comparison["quarterly_total"] - comparison["monthly_sum"]
    comparison["pct_diff"] = (comparison["difference"] / comparison["monthly_sum"] * 100).round(2)

    return comparison.sort_values(["year", "quarter"])


def get_validation_report(
    force_refresh: bool = False,
) -> Dict[str, pd.DataFrame]:
    """Run all cross-validations and return comprehensive report.

    Compares quarterly data against monthly data sources to verify consistency.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        Dict with keys:
            - 'births_validation': Quarterly vs monthly births comparison
            - 'marriages_validation': Quarterly vs monthly marriages comparison
            - 'summary': Overall validation statistics

    Example:
        >>> report = get_validation_report()
        >>> print(report['summary'])
        >>> print(report['births_validation'])
    """
    data = get_quarterly_vital_statistics(force_refresh=force_refresh)

    results = {}

    # Births validation
    if not data["births"].empty:
        try:
            results["births_validation"] = validate_against_monthly_births(data["births"])
        except Exception as e:
            logger.warning(f"Could not validate births: {e}")
            results["births_validation"] = pd.DataFrame()

    # Marriages validation
    if not data["deaths"].empty and "marriages" in data["deaths"].columns:
        try:
            results["marriages_validation"] = validate_against_monthly_marriages(data["deaths"])
        except Exception as e:
            logger.warning(f"Could not validate marriages: {e}")
            results["marriages_validation"] = pd.DataFrame()

    # Create summary
    summary_data = []

    for key, validation_df in results.items():
        if not validation_df.empty and "pct_diff" in validation_df.columns:
            avg_diff = validation_df["pct_diff"].abs().mean()
            max_diff = validation_df["pct_diff"].abs().max()
            quarters_validated = len(validation_df)

            summary_data.append(
                {
                    "validation": key,
                    "quarters_compared": quarters_validated,
                    "avg_pct_diff": round(avg_diff, 2),
                    "max_pct_diff": round(max_diff, 2),
                    "within_2pct": (validation_df["pct_diff"].abs() <= 2).sum(),
                }
            )

    results["summary"] = pd.DataFrame(summary_data)

    return results


def validate_data(df: pd.DataFrame, data_type: str = "births") -> bool:
    """Validate quarterly data for consistency and reasonable values.

    Args:
        df: DataFrame to validate
        data_type: Type of data ('births', 'deaths', or 'lgd')

    Returns:
        True if validation passes

    Raises:
        NISRAValidationError: If validation fails
    """
    if df.empty:
        raise NISRAValidationError(f"Empty DataFrame for {data_type}")

    if data_type == "births":
        required_cols = ["year", "quarter", "total_births"]
        for col in required_cols:
            if col not in df.columns:
                raise NISRAValidationError(f"Missing required column: {col}")

        # Check no negative values
        if (df["total_births"] < 0).any():
            raise NISRAValidationError("Negative birth counts found")

        # Check reasonable range (NI typically has 4,000-6,000 births per quarter)
        if df["total_births"].max() > 10000:
            raise NISRAValidationError("Unreasonably high quarterly birth count")

        # Check we have historical data
        if df["year"].min() > 2010:
            raise NISRAValidationError(f"Expected data from 2009, earliest is {df['year'].min()}")

    elif data_type == "deaths":
        required_cols = ["year", "quarter", "deaths"]
        for col in required_cols:
            if col not in df.columns:
                raise NISRAValidationError(f"Missing required column: {col}")

        if (df["deaths"] < 0).any():
            raise NISRAValidationError("Negative death counts found")

    elif data_type == "lgd":
        if len(df) < 11:
            raise NISRAValidationError(f"Expected 11 LGDs, found {len(df)}")

    logger.info(f"Validation passed for {data_type} data")
    return True
