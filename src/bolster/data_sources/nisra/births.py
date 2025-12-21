"""NISRA Monthly Birth Registrations Data Source.

Provides access to monthly birth registration statistics for Northern Ireland with breakdowns by:
- Sex (Persons, Male, Female)
- Event type (Registration date vs Birth/Occurrence date)

Births data are based on residence of mother at time of birth. Data includes both:
- Births by month of registration: When the birth was officially registered
- Births by month of occurrence: When the birth actually occurred

Most births are registered within 42 days in Northern Ireland, so registration data lags
occurrence data slightly.

Data Source:
    **Mother Page**: https://www.nisra.gov.uk/statistics/births-deaths-and-marriages/births

    This page lists all births statistics publications in reverse chronological order
    (newest first). The module automatically scrapes this page to find the latest
    "Monthly Births" publication, then downloads the Excel file from that publication's
    detail page.

    The monthly files contain complete time series from 2006 to present, updated monthly.
    This ensures the module always retrieves the most recent data without hardcoding dates.

Update Frequency: Monthly (published 12th of following month at 9:30 AM)
Geographic Coverage: Northern Ireland (based on mother's residence)

Example:
    >>> from bolster.data_sources.nisra import births
    >>> # Get latest births by registration date
    >>> df = births.get_latest_births(event_type='registration')
    >>> print(df.head())

    >>> # Get latest births by occurrence date
    >>> df = births.get_latest_births(event_type='occurrence')
"""

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Literal

import pandas as pd
from openpyxl import load_workbook

from ._base import (
    NISRADataNotFoundError,
    NISRAValidationError,
    download_file,
    safe_int,
)

logger = logging.getLogger(__name__)

# Base URL for births statistics
BIRTHS_BASE_URL = "https://www.nisra.gov.uk/statistics/births-deaths-and-marriages/births"


def get_latest_births_publication_url() -> str:
    """Scrape NISRA births mother page to find the latest monthly births file.

    Navigates the publication structure:
    1. Scrapes mother page for "Monthly Births" publication
    2. Follows link to publication detail page
    3. Finds latest Excel file

    Returns:
        URL of the latest monthly births Excel file

    Raises:
        NISRADataNotFoundError: If publication or file not found
    """
    import requests
    from bs4 import BeautifulSoup

    mother_page = BIRTHS_BASE_URL

    try:
        response = requests.get(mother_page, timeout=30)
        response.raise_for_status()
    except requests.RequestException as e:
        raise NISRADataNotFoundError(f"Failed to fetch births mother page: {e}")

    soup = BeautifulSoup(response.content, "html.parser")

    # Find "Monthly Births" publication link
    pub_link = None
    for link in soup.find_all("a", href=True):
        link_text = link.get_text(strip=True)
        if "Monthly Births" in link_text and "publications" in link["href"]:
            href = link["href"]
            if href.startswith("/"):
                href = f"https://www.nisra.gov.uk{href}"
            pub_link = href
            logger.info(f"Found Monthly Births publication: {link_text}")
            break

    if not pub_link:
        raise NISRADataNotFoundError("Could not find 'Monthly Births' publication on mother page")

    # Scrape the publication page for Excel file
    try:
        pub_response = requests.get(pub_link, timeout=30)
        pub_response.raise_for_status()
    except requests.RequestException as e:
        raise NISRADataNotFoundError(f"Failed to fetch publication page: {e}")

    pub_soup = BeautifulSoup(pub_response.content, "html.parser")

    # Find Excel file link
    excel_url = None
    for link in pub_soup.find_all("a", href=True):
        href = link["href"]
        if href.endswith(".xlsx") and "Monthly" in href and "Births" in href:
            if href.startswith("/"):
                href = f"https://www.nisra.gov.uk{href}"
            excel_url = href

            # Extract month/year from filename for logging
            # Pattern: "Monthly Births November 2025.xlsx"
            match = re.search(r"Monthly\s+Births\s+([A-Za-z]+)\s+(\d{4})", href)
            if match:
                month = match.group(1)
                year = match.group(2)
                logger.info(f"Found latest births file: {month} {year}")
            break

    if not excel_url:
        raise NISRADataNotFoundError("Could not find Excel file on publication page")

    return excel_url


def parse_births_file(
    file_path: str | Path,
    event_type: Literal["registration", "occurrence", "both"] = "both",
) -> pd.DataFrame | dict[str, pd.DataFrame]:
    """Parse NISRA monthly births Excel file into long-format DataFrames.

    The births file contains two main sheets:
    - Births_Month of Registration: Births by when registered
    - Births_Month of Birth: Births by when occurred

    Each sheet has three stacked tables (Persons, Males, Females) in wide format
    (months as rows, years as columns). This function converts to long format.

    Args:
        file_path: Path to the births Excel file
        event_type: Which event type to parse:
            - "registration": Births by registration date only
            - "occurrence": Births by birth/occurrence date only
            - "both": Return dict with both types

    Returns:
        If event_type is "both": Dict with keys "registration" and "occurrence"
        Otherwise: Single DataFrame with columns:
            - month: datetime (first day of month)
            - sex: str (Persons, Male, Female)
            - births: int (number of births)

    Raises:
        NISRAValidationError: If file structure is unexpected
    """
    file_path = Path(file_path)
    wb = load_workbook(file_path, data_only=True)

    results = {}

    # Sheet mapping
    sheet_mapping = {
        "registration": "Births_Month of Registration",
        "occurrence": "Births_Month of Birth",
    }

    for event_key, sheet_name in sheet_mapping.items():
        if event_type != "both" and event_type != event_key:
            continue

        if sheet_name not in wb.sheetnames:
            raise NISRAValidationError(f"Expected sheet '{sheet_name}' not found in file")

        sheet = wb[sheet_name]

        # Parse the three stacked tables
        dfs = []

        # Table positions (adjusted based on actual data)
        # Persons: rows 4-17
        # Males: rows 19-32
        # Females: rows 34-47
        sex_tables = [
            ("Persons", 4, 17),
            ("Male", 19, 32),
            ("Female", 34, 47),
        ]

        for sex, header_row, end_row in sex_tables:
            df = _parse_births_table(sheet, header_row, end_row, sex, event_key)
            dfs.append(df)

        # Combine all sex breakdowns
        combined = pd.concat(dfs, ignore_index=True)

        # Sort by month and sex
        combined = combined.sort_values(["month", "sex"])

        results[event_key] = combined

    # Return based on event_type
    if event_type == "both":
        return results
    else:
        return results[event_type]


def _parse_births_table(
    sheet,
    header_row: int,
    end_row: int,
    sex: str,
    event_type: str,
) -> pd.DataFrame:
    """Parse a single births table (one sex) from the Excel sheet.

    Args:
        sheet: openpyxl worksheet
        header_row: Row number with year headers (1-indexed)
        end_row: Last row of data (1-indexed)
        sex: Sex category (Persons, Male, Female)
        event_type: registration or occurrence

    Returns:
        Long-format DataFrame with columns: month, sex, births
    """
    # Read header row to get years
    header_cells = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]

    # First column is "Month of Registration" or "Month of Birth"
    # Remaining columns are years
    years = []
    for cell in header_cells[1:]:  # Skip first column
        if cell is not None:
            # Clean year (might have notes like "2025\n[Note 1]")
            year_str = str(cell).split("\n")[0].strip()
            try:
                year = int(year_str)
                years.append(year)
            except ValueError:
                break  # Stop at first non-year column

    if not years:
        raise NISRAValidationError(f"Could not parse years from header row {header_row}")

    # Parse data rows (months)
    month_names = [
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    ]

    data_rows = []

    for row in sheet.iter_rows(min_row=header_row + 1, max_row=end_row, values_only=True):
        month_label = str(row[0]).strip() if row[0] else None

        # Skip "Total" row
        if month_label == "Total":
            continue

        # Match against expected month names
        if month_label not in month_names:
            continue

        month_index = month_names.index(month_label) + 1  # 1-12

        # Extract birth counts for each year
        for i, year in enumerate(years):
            value = row[i + 1]  # +1 to skip month label column

            # Handle missing data (dash or None)
            if value == "-" or value is None:
                continue

            births = safe_int(value)

            if births is None:
                continue

            # Create month datetime (first day of month)
            try:
                month_date = datetime(year, month_index, 1)
            except ValueError:
                continue

            data_rows.append(
                {
                    "month": month_date,
                    "sex": sex,
                    "births": births,
                }
            )

    return pd.DataFrame(data_rows)


def get_latest_births(
    event_type: Literal["registration", "occurrence", "both"] = "both",
    force_refresh: bool = False,
) -> pd.DataFrame | dict[str, pd.DataFrame]:
    """Get the latest monthly births data.

    Automatically discovers and downloads the most recent monthly births publication
    from the NISRA website.

    Args:
        event_type: Which event type to retrieve:
            - "registration": Births by month registered
            - "occurrence": Births by month of birth (occurrence)
            - "both": Return dict with both types
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        If event_type is "both": Dict with keys "registration" and "occurrence"
        Otherwise: Single DataFrame with columns:
            - month: datetime (first day of month)
            - sex: str (Persons, Male, Female)
            - births: int

    Raises:
        NISRADataNotFoundError: If latest publication cannot be found
        NISRAValidationError: If file structure is unexpected

    Example:
        >>> # Get births by registration date
        >>> df = get_latest_births(event_type='registration')
        >>> print(df[df['sex'] == 'Male'].head())

        >>> # Get both types
        >>> data = get_latest_births(event_type='both')
        >>> print(data['registration'].head())
        >>> print(data['occurrence'].head())
    """
    # Discover latest publication
    excel_url = get_latest_births_publication_url()

    logger.info(f"Downloading latest births data from: {excel_url}")

    # Cache for 30 days (monthly data)
    cache_ttl_hours = 30 * 24
    file_path = download_file(excel_url, cache_ttl_hours=cache_ttl_hours, force_refresh=force_refresh)

    # Parse the file
    return parse_births_file(file_path, event_type=event_type)


def validate_births_totals(df: pd.DataFrame) -> bool:
    """Validate that Male + Female births equal Persons births for each month.

    Args:
        df: DataFrame from parse_births_file or get_latest_births

    Returns:
        True if validation passes

    Raises:
        NISRAValidationError: If validation fails
    """
    months = df["month"].unique()

    for month in months:
        month_data = df[df["month"] == month]

        persons = month_data[month_data["sex"] == "Persons"]["births"].sum()
        male = month_data[month_data["sex"] == "Male"]["births"].sum()
        female = month_data[month_data["sex"] == "Female"]["births"].sum()

        if persons != male + female:
            raise NISRAValidationError(
                f"Month {month.date()}: Persons ({persons}) != Male ({male}) + Female ({female})"
            )

    logger.info(f"Validation passed: Male + Female = Persons for {len(months)} months")
    return True
