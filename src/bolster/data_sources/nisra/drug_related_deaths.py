"""NISRA Drug-Related and Drug Misuse Deaths Data Source.

Provides access to NISRA's annual statistics on drug-related deaths and deaths
due to drug misuse in Northern Ireland, by registration year (2014 onwards in the
current release). The data are accredited official statistics, published annually
(typically in May).

NISRA distinguishes two related measures:

- **Drug-related deaths**: All deaths where a drug was implicated, including
  prescription medicines, controlled drugs, and accidental/intentional poisonings.
- **Drug misuse deaths**: A narrower subset where the underlying cause is drug
  abuse/dependence or the death involved a controlled (illegal) substance.

Three dimensions are exposed:

- ``summary``: Annual totals, crude rates, and age-standardised rates by year
  and measure (from Table 1).
- ``age``: Counts by age band, gender, and year (from Tables 3a and 3c).
- ``substances``: Counts of deaths mentioning selected substances by year
  (from Table 4a, drug-related deaths only).

Data Source:
    **Mother Page**: https://www.nisra.gov.uk/statistics/cause-death/drug-related-deaths

    This page lists each annual publication (a rolling 11-year window such as
    "2014-2024"). The module scrapes the page, selects the publication covering
    the most recent end year, follows it, and downloads the single Excel workbook
    containing all tables.

Update Frequency: Annual (typically May)
Geographic Coverage: Northern Ireland

Example:
    >>> from bolster.data_sources.nisra import drug_related_deaths as drd
    >>> df = drd.get_latest_data(dimension="summary")
    >>> sorted(df.columns.tolist())
    ['deaths', 'measure', 'metric', 'value', 'year']
    >>> len(df) > 0
    True
"""

import logging
import re
from pathlib import Path
from typing import Literal

import pandas as pd
from openpyxl import load_workbook

from bolster.utils.web import session

from ._base import (
    NISRADataNotFoundError,
    download_file,
    make_absolute_url,
    safe_float,
    safe_int,
)

logger = logging.getLogger(__name__)

NISRA_BASE_URL = "https://www.nisra.gov.uk"
DRD_LANDING_PAGE = "https://www.nisra.gov.uk/statistics/cause-death/drug-related-deaths"

DimensionType = Literal["summary", "age", "substances", "all"]


def get_latest_publication_url() -> str:
    """Return the URL of the most recent drug-related deaths Excel workbook.

    Scrapes the NISRA mother page, identifies the publication covering the most
    recent end year (e.g. "2014-2024"), follows it, and returns the ``.xlsx``
    download URL found on the publication detail page.

    Returns:
        Absolute URL of the latest Excel workbook.

    Raises:
        NISRADataNotFoundError: If no publication or Excel link can be located.

    Example:
        >>> url = get_latest_publication_url()
        >>> url.endswith(".xlsx")
        True
    """
    from bs4 import BeautifulSoup

    logger.info("Fetching drug-related deaths landing page: %s", DRD_LANDING_PAGE)
    try:
        resp = session.get(DRD_LANDING_PAGE, timeout=30)
        resp.raise_for_status()
    except Exception as exc:  # pragma: no cover - network failure path
        raise NISRADataNotFoundError(f"Failed to fetch drug-related deaths landing page: {exc}") from exc

    soup = BeautifulSoup(resp.content, "html.parser")

    # Publication slugs look like ".../drug-related-and-drug-misuse-deaths-2014-2024".
    # Pick the one with the highest end year.
    best_url: str | None = None
    best_end_year = -1
    slug_re = re.compile(r"drug-related-and-drug-misuse-deaths-(\d{4})-(\d{4})")
    for a in soup.find_all("a", href=True):
        match = slug_re.search(a["href"])
        if match:
            end_year = int(match.group(2))
            if end_year > best_end_year:
                best_end_year = end_year
                best_url = make_absolute_url(a["href"], NISRA_BASE_URL)

    if best_url is None:
        raise NISRADataNotFoundError(f"Could not find a drug-related deaths publication link on {DRD_LANDING_PAGE}")

    logger.info("Selected drug-related deaths publication (end year %s): %s", best_end_year, best_url)
    try:
        pub_resp = session.get(best_url, timeout=30)
        pub_resp.raise_for_status()
    except Exception as exc:  # pragma: no cover - network failure path
        raise NISRADataNotFoundError(f"Failed to fetch publication page {best_url}: {exc}") from exc

    pub_soup = BeautifulSoup(pub_resp.content, "html.parser")
    for a in pub_soup.find_all("a", href=True):
        href = a["href"]
        if href.lower().endswith(".xlsx"):
            return make_absolute_url(href, NISRA_BASE_URL)

    raise NISRADataNotFoundError(f"Could not find an .xlsx link on publication page {best_url}")


def _year_columns(header_row: tuple) -> dict[int, int]:
    """Map column index -> year for a header row whose cells are year labels.

    The total/combined column (e.g. "Total (2014-2024)") is excluded.
    """
    cols: dict[int, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        text = str(cell).strip()
        if text.isdigit() and len(text) == 4:
            cols[idx] = int(text)
    return cols


def parse_summary(file_path: str | Path) -> pd.DataFrame:
    """Parse annual totals and rates (Table 1) into long format.

    Table 1 stacks three blocks vertically (Persons, Males, Females). Each block
    contains both measures (drug-related deaths and drug misuse deaths) with their
    counts, crude rates, and age-standardised rates.

    Args:
        file_path: Path to the drug deaths Excel workbook.

    Returns:
        Long-format DataFrame with columns:

        - year: Registration year (int)
        - gender: 'Persons', 'Males', or 'Females'
        - measure: 'drug_related' or 'drug_misuse'
        - metric: 'deaths', 'crude_rate', or 'asmr'
        - value: Numeric value for the metric

    Raises:
        NISRADataNotFoundError: If Table 1 is missing from the workbook.
    """
    wb = load_workbook(file_path, data_only=True)
    if "Table_1" not in wb.sheetnames:
        wb.close()
        raise NISRADataNotFoundError("Table_1 (summary) not found in workbook")
    sheet = wb["Table_1"]

    records: list[dict] = []
    current_gender: str | None = None
    year_cols: dict[int, int] = {}
    current_measure: str | None = None

    for row in sheet.iter_rows(values_only=True):
        first = row[0]
        if first is None:
            continue
        label = str(first).strip()
        label_lower = label.lower()

        # A header row whose label is a gender begins a new block.
        if label in ("Persons", "Males", "Females"):
            current_gender = label
            year_cols = _year_columns(row)
            current_measure = None
            continue

        if not year_cols or current_gender is None:
            continue

        # Determine measure/metric from the row label.
        if label_lower.startswith("drug-related deaths"):
            current_measure, metric = "drug_related", "deaths"
        elif label_lower.startswith("deaths due to drug misuse"):
            current_measure, metric = "drug_misuse", "deaths"
        elif label_lower.startswith("crude rate"):
            metric = "crude_rate"
        elif label_lower.startswith("age-standardised mortality rate"):
            metric = "asmr"
        else:
            # Rolling averages, percentages, all-cause totals: skip.
            continue

        if current_measure is None:
            continue

        for col_idx, year in year_cols.items():
            value = safe_float(row[col_idx]) if metric != "deaths" else safe_int(row[col_idx])
            if value is None:
                continue
            records.append(
                {
                    "year": year,
                    "gender": current_gender,
                    "measure": current_measure,
                    "metric": metric,
                    "value": value,
                }
            )

    wb.close()

    df = pd.DataFrame(records)
    # `deaths` is a convenience alias for value where metric == 'deaths'.
    df["deaths"] = df.apply(lambda r: int(r["value"]) if r["metric"] == "deaths" else pd.NA, axis=1)
    df = df.sort_values(["measure", "gender", "metric", "year"]).reset_index(drop=True)
    logger.info("Parsed summary: %d rows across %d years", len(df), df["year"].nunique())
    return df


def parse_age(file_path: str | Path) -> pd.DataFrame:
    """Parse counts by age band, gender, and year (Tables 3a and 3c).

    Table 3a covers drug-related deaths; Table 3c covers drug misuse deaths. Each
    stacks Persons/Males/Females blocks vertically.

    Args:
        file_path: Path to the drug deaths Excel workbook.

    Returns:
        Long-format DataFrame with columns:

        - year: Registration year (int)
        - measure: 'drug_related' or 'drug_misuse'
        - gender: 'Persons', 'Males', or 'Females'
        - age_band: Age band label (e.g. 'Under 25', '25-34', '65 and over')
        - deaths: Count of deaths

    Raises:
        NISRADataNotFoundError: If neither age table is present.
    """
    wb = load_workbook(file_path, data_only=True)
    sheet_measures = [("Table_3a", "drug_related"), ("Table_3c", "drug_misuse")]
    if not any(name in wb.sheetnames for name, _ in sheet_measures):
        wb.close()
        raise NISRADataNotFoundError("No age breakdown tables (Table_3a/Table_3c) found in workbook")

    records: list[dict] = []
    for sheet_name, measure in sheet_measures:
        if sheet_name not in wb.sheetnames:
            continue
        sheet = wb[sheet_name]
        current_gender: str | None = None
        year_cols: dict[int, int] = {}

        for row in sheet.iter_rows(values_only=True):
            first = row[0]
            if first is None:
                continue
            label = str(first).strip()

            if label in ("Persons", "Males", "Females"):
                current_gender = label
                year_cols = _year_columns(row)
                continue

            if not year_cols or current_gender is None or label == "All":
                continue

            for col_idx, year in year_cols.items():
                deaths = safe_int(row[col_idx])
                if deaths is None:
                    continue
                records.append(
                    {
                        "year": year,
                        "measure": measure,
                        "gender": current_gender,
                        "age_band": label,
                        "deaths": deaths,
                    }
                )

    wb.close()
    df = pd.DataFrame(records).sort_values(["measure", "gender", "age_band", "year"]).reset_index(drop=True)
    logger.info("Parsed age breakdown: %d rows", len(df))
    return df


def parse_substances(file_path: str | Path) -> pd.DataFrame:
    """Parse counts of drug-related deaths mentioning selected substances (Table 4a).

    Args:
        file_path: Path to the drug deaths Excel workbook.

    Returns:
        Long-format DataFrame with columns:

        - year: Registration year (int)
        - substance: Substance label (e.g. 'Heroin/Morphine', 'Cocaine')
        - deaths: Count of drug-related deaths mentioning the substance

    Raises:
        NISRADataNotFoundError: If Table 4a is missing from the workbook.
    """
    wb = load_workbook(file_path, data_only=True)
    if "Table_4a" not in wb.sheetnames:
        wb.close()
        raise NISRADataNotFoundError("Table_4a (substances) not found in workbook")
    sheet = wb["Table_4a"]

    records: list[dict] = []
    year_cols: dict[int, int] = {}
    for row in sheet.iter_rows(values_only=True):
        first = row[0]
        if first is None:
            continue
        label = str(first).strip()

        if not year_cols:
            candidate = _year_columns(row)
            if candidate:
                year_cols = candidate
            continue

        # Skip the total row that repeats the all-deaths figure.
        if label.lower().startswith("drug related"):
            continue

        for col_idx, year in year_cols.items():
            deaths = safe_int(row[col_idx])
            if deaths is None:
                continue
            records.append({"year": year, "substance": label, "deaths": deaths})

    wb.close()
    df = pd.DataFrame(records).sort_values(["substance", "year"]).reset_index(drop=True)
    logger.info("Parsed substances: %d rows across %d substances", len(df), df["substance"].nunique())
    return df


def parse_data(file_path: str | Path, dimension: DimensionType = "summary") -> pd.DataFrame | dict[str, pd.DataFrame]:
    """Parse the drug deaths workbook for one or all dimensions.

    Args:
        file_path: Path to the drug deaths Excel workbook.
        dimension: Which dimension(s) to parse:

            - 'summary': Annual totals and rates (Table 1)
            - 'age': Counts by age band, gender, year (Tables 3a/3c)
            - 'substances': Counts by selected substance (Table 4a)
            - 'all': All dimensions (returns dict)

    Returns:
        DataFrame for a single dimension, or dict of DataFrames for 'all'.

    Example:
        >>> url = get_latest_publication_url()
        >>> path = download_file(url, cache_ttl_hours=24 * 30)
        >>> data = parse_data(path, dimension="all")
        >>> sorted(data.keys())
        ['age', 'substances', 'summary']
    """
    if dimension == "all":
        return {
            "summary": parse_summary(file_path),
            "age": parse_age(file_path),
            "substances": parse_substances(file_path),
        }
    if dimension == "summary":
        return parse_summary(file_path)
    if dimension == "age":
        return parse_age(file_path)
    if dimension == "substances":
        return parse_substances(file_path)
    raise ValueError(f"Invalid dimension: {dimension}. Must be one of: summary, age, substances, all")


def get_latest_data(
    dimension: DimensionType = "summary", force_refresh: bool = False
) -> pd.DataFrame | dict[str, pd.DataFrame]:
    """Get the latest drug-related deaths data.

    Args:
        dimension: Which dimension(s) to retrieve (see :func:`parse_data`).
        force_refresh: Force re-download even if cached.

    Returns:
        DataFrame or dict of DataFrames depending on dimension.

    Example:
        >>> df = get_latest_data(dimension="summary")
        >>> sorted(df.columns.tolist())
        ['deaths', 'measure', 'metric', 'value', 'year']
        >>> len(df) > 0
        True
    """
    url = get_latest_publication_url()
    file_path = download_file(url, cache_ttl_hours=24 * 30, force_refresh=force_refresh)
    return parse_data(file_path, dimension=dimension)


def validate_data(df: pd.DataFrame) -> bool:
    """Validate a parsed drug-related deaths summary DataFrame.

    Args:
        df: DataFrame from :func:`get_latest_data` with ``dimension='summary'``.

    Returns:
        True if validation passes, False otherwise.

    Example:
        >>> import pandas as pd
        >>> validate_data(pd.DataFrame())
        False
    """
    if df is None or df.empty:
        logger.warning("Drug deaths data is empty")
        return False

    required_cols = {"year", "gender", "measure", "metric", "value"}
    if not required_cols.issubset(df.columns):
        missing = required_cols - set(df.columns)
        logger.warning("Missing required columns: %s", missing)
        return False

    if (df["value"] < 0).any():
        logger.warning("Found negative values in drug deaths data")
        return False

    # Need at least a few years of data to be a useful time series.
    if df["year"].nunique() < 5:
        logger.warning("Too few years of data: %d", df["year"].nunique())
        return False

    return True
