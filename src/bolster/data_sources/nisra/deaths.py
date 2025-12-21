"""NISRA Weekly Death Registrations Data Source.

Provides access to weekly death registration statistics for Northern Ireland with breakdowns by:
- Demographics (age, sex)
- Geography (Local Government Districts)
- Place of death (hospital, home, care home, etc.)

Data is based on registration date, not death occurrence date. Most deaths are registered
within 5 days in Northern Ireland.

Update Frequency: Weekly
Geographic Coverage: Northern Ireland
Source: https://www.nisra.gov.uk/statistics/death-statistics/weekly-death-registrations-northern-ireland

Example:
    >>> from bolster.data_sources.nisra import deaths
    >>> # Get latest demographics breakdown
    >>> df = deaths.get_latest_deaths(dimension='demographics')
    >>> print(df.head())

    >>> # Get specific week
    >>> df = deaths.parse_deaths_file('/path/to/file.xlsx', dimension='geography')
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Literal, Optional, Union

import pandas as pd
from openpyxl import load_workbook

from ._base import (
    NISRADataNotFoundError,
    NISRAValidationError,
    download_file,
    safe_float,
    safe_int,
    scrape_download_links,
)

logger = logging.getLogger(__name__)

# URL patterns
WEEKLY_DEATHS_LANDING_PAGE = "https://www.nisra.gov.uk/publications/weekly-death-registrations-northern-ireland-{year}"
WEEKLY_DEATHS_BASE_URL = (
    "https://www.nisra.gov.uk/statistics/death-statistics/weekly-death-registrations-northern-ireland"
)
HISTORICAL_DEATHS_URL = "https://www.nisra.gov.uk/publications/historical-final-weekly-deaths-data"

DimensionType = Literal["demographics", "geography", "place", "totals", "all"]


def get_latest_weekly_deaths_url() -> str:
    """Scrape NISRA page to find the latest weekly deaths file.

    Returns:
        URL of the latest weekly deaths Excel file

    Raises:
        NISRADataNotFoundError: If no file found
    """
    current_year = datetime.now().year
    page_url = WEEKLY_DEATHS_LANDING_PAGE.format(year=current_year)

    try:
        links = scrape_download_links(page_url, file_extension=".xlsx")
    except Exception:
        # Try the main landing page if current year page doesn't exist
        links = scrape_download_links(WEEKLY_DEATHS_BASE_URL, file_extension=".xlsx")

    if not links:
        raise NISRADataNotFoundError("No weekly deaths Excel files found on NISRA website")

    # Filter for weekly deaths files (exclude historical series, etc.)
    weekly_files = [
        link for link in links if "weekly_deaths" in link["url"].lower() and "historical" not in link["url"].lower()
    ]

    if not weekly_files:
        # Fall back to any xlsx link if no explicit weekly files found
        weekly_files = links

    # Return the first one (usually the most recent)
    latest = weekly_files[0]
    logger.info(f"Found latest weekly deaths file: {latest['text']}")
    return latest["url"]


def parse_deaths_totals(file_path: Union[str, Path]) -> pd.DataFrame:
    """Parse weekly totals with COVID-19, flu/pneumonia, and excess deaths from weekly deaths file.

    Extracts Table 1a and creates a flat table with columns:
    - week_ending: Friday of the reporting week
    - week_number: Week number in the year
    - observed_deaths: Total deaths registered in the week
    - deaths_same_week_2024: Deaths in corresponding week in 2024
    - expected_deaths_5yr: Average deaths over previous 5 years (2020-2024)
    - expected_deaths_ons: Average deaths using ONS methodology (2019, 2021-2024)
    - excess_deaths_5yr: Observed minus expected (5yr method)
    - expected_deaths_current: Expected deaths using current methodology
    - excess_deaths_current: Observed minus expected (current method)
    - flu_pneumonia_deaths: Deaths mentioning flu or pneumonia
    - covid_deaths: Deaths mentioning COVID-19

    Args:
        file_path: Path to the weekly deaths Excel file

    Returns:
        DataFrame with weekly totals and disease-specific deaths

    Raises:
        NISRAValidationError: If data validation fails
    """
    wb = load_workbook(file_path, data_only=True)

    if "Table 1a" not in wb.sheetnames:
        raise NISRADataNotFoundError("Table 1a (totals) not found in file")

    sheet = wb["Table 1a"]

    # Find header row (row 4 based on inspection)
    # Header row contains: "Registration Week", "Week Ending (Friday)", etc.
    header_row = 4

    records = []

    # Read data rows (start from row 5)
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        week_number = row[0]
        week_ending = row[1]

        # Skip if no week ending date (end of data)
        if not week_ending or not isinstance(week_ending, datetime):
            break

        # Extract the key columns
        # Column indices based on inspection:
        # A(0): Week Number
        # B(1): Week Ending
        # C(2): Observed Deaths
        # D(3): Deaths in 2024
        # G(6): Expected Deaths (5yr, 2020-2024)
        # H(7): Expected Deaths (ONS method)
        # I(8): Excess Deaths (old method)
        # J(9): Expected Deaths (current method)
        # M(12): Excess Deaths (current method)
        # P(15): Flu/Pneumonia Deaths
        # Q(16): COVID-19 Deaths

        records.append(
            {
                "week_ending": week_ending,
                "week_number": int(week_number) if isinstance(week_number, (int, float)) else week_number,
                "observed_deaths": int(row[2]) if row[2] is not None else None,
                "deaths_same_week_2024": int(row[3]) if row[3] is not None else None,
                "expected_deaths_5yr": float(row[6]) if row[6] is not None else None,
                "expected_deaths_ons": float(row[7]) if row[7] is not None else None,
                "excess_deaths_5yr": float(row[8]) if row[8] is not None else None,
                "expected_deaths_current": float(row[9]) if row[9] is not None else None,
                "excess_deaths_current": float(row[12]) if row[12] is not None else None,
                "flu_pneumonia_deaths": int(row[15]) if row[15] is not None else None,
                "covid_deaths": int(row[16]) if row[16] is not None else None,
            }
        )

    wb.close()

    df = pd.DataFrame(records)
    df["week_ending"] = pd.to_datetime(df["week_ending"])

    # Validate: observed deaths should be positive
    if (df["observed_deaths"] <= 0).any():
        raise NISRAValidationError("Found zero or negative observed deaths")

    logger.info(f"Totals parsed successfully ({len(df)} weeks)")
    logger.info(f"Total COVID-19 deaths in period: {df['covid_deaths'].sum()}")
    logger.info(f"Total Flu/Pneumonia deaths in period: {df['flu_pneumonia_deaths'].sum()}")

    return df


def parse_deaths_demographics(file_path: Union[str, Path]) -> pd.DataFrame:
    """Parse demographics dimension (age, sex) from weekly deaths file.

    Extracts Table 2 and creates a flat table with columns:
    - week_ending: Friday of the reporting week
    - sex: Total, Male, or Female
    - age_range: All, 0-14, 15-44, 45-64, 65-74, 75-84, 85+
    - deaths: Count of deaths

    Args:
        file_path: Path to the weekly deaths Excel file

    Returns:
        DataFrame with demographics breakdown

    Raises:
        NISRAValidationError: If data validation fails
    """
    wb = load_workbook(file_path, data_only=True)

    if "Table 2" not in wb.sheetnames:
        raise NISRADataNotFoundError("Table 2 (demographics) not found in file")

    sheet = wb["Table 2"]

    # Find the header row (contains "Sex", "Age", and week columns)
    header_row = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if row[0] == "Sex" and row[1] == "Age":
            header_row = row_idx
            break

    if header_row is None:
        raise NISRADataNotFoundError("Could not find header row in Table 2")

    # Extract headers
    headers = [cell for cell in sheet[header_row]]

    # Find week columns (start from column index 2, format: "Week N\n DD Month YYYY")
    # Skip "to Date" cumulative column
    week_cols = []
    for col_idx, cell in enumerate(headers[2:], start=2):
        if cell.value and "Week" in str(cell.value) and "to Date" not in str(cell.value):
            week_cols.append(col_idx)

    # Build flat table
    records = []
    current_sex = None  # Track sex across rows for merged cells

    # Read data rows
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        sex = row[0]
        age_range = row[1]

        # Skip if no sex or age (footer rows, etc.)
        if not sex and not age_range:
            continue

        # Skip footer notes
        if sex and "Note" in str(sex):
            break

        # If sex is None but age_range exists, inherit sex from previous row
        if sex is None:
            sex = current_sex
        else:
            current_sex = sex

        # Parse each week column
        for col_idx in week_cols:
            deaths = row[col_idx]

            if deaths is None or deaths == "":
                continue

            # Parse week ending date from header
            week_header = headers[col_idx].value
            week_ending = _parse_week_ending(week_header)

            records.append(
                {
                    "week_ending": week_ending,
                    "sex": str(sex).strip(),
                    "age_range": str(age_range).strip(),
                    "deaths": int(deaths),
                }
            )

    wb.close()

    df = pd.DataFrame(records)
    df["week_ending"] = pd.to_datetime(df["week_ending"])

    # Validate
    _validate_demographics(df)

    return df


def parse_deaths_geography(file_path: Union[str, Path]) -> pd.DataFrame:
    """Parse geography dimension (Local Government Districts) from weekly deaths file.

    Extracts Table 3 and creates a flat table with columns:
    - week_ending: Friday of the reporting week
    - lgd: Local Government District name
    - deaths: Count of deaths

    Args:
        file_path: Path to the weekly deaths Excel file

    Returns:
        DataFrame with geography breakdown

    Raises:
        NISRAValidationError: If data validation fails
    """
    wb = load_workbook(file_path, data_only=True)

    if "Table 3" not in wb.sheetnames:
        raise NISRADataNotFoundError("Table 3 (geography) not found in file")

    sheet = wb["Table 3"]

    # Find header row
    header_row = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if "Week Ending" in str(row[1] or ""):
            header_row = row_idx
            break

    if header_row is None:
        raise NISRADataNotFoundError("Could not find header row in Table 3")

    # Extract LGD names from header (columns 2 onwards)
    # Exclude 'Total' column to avoid double-counting
    headers = [cell.value for cell in sheet[header_row]]
    lgd_columns = {
        idx: headers[idx] for idx in range(2, len(headers)) if headers[idx] and "Total" not in str(headers[idx])
    }

    records = []

    # Read data rows
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        _week_num = row[0]  # Week number column exists but not used
        week_ending = row[1]

        # Skip if no week ending date
        if not week_ending or not isinstance(week_ending, datetime):
            continue

        # Parse each LGD column
        for col_idx, lgd_name in lgd_columns.items():
            deaths = row[col_idx]

            if deaths is not None and deaths != "":
                records.append({"week_ending": week_ending, "lgd": str(lgd_name).strip(), "deaths": int(deaths)})

    wb.close()

    df = pd.DataFrame(records)
    df["week_ending"] = pd.to_datetime(df["week_ending"])

    # Validate
    _validate_geography(df)

    return df


def parse_deaths_place(file_path: Union[str, Path]) -> pd.DataFrame:
    """Parse place of death dimension from weekly deaths file.

    Extracts Table 4 and creates a flat table with columns:
    - week_ending: Friday of the reporting week
    - place_of_death: Hospital, Care/Nursing Home, Hospice, Home, Other
    - deaths: Count of deaths

    Args:
        file_path: Path to the weekly deaths Excel file

    Returns:
        DataFrame with place of death breakdown

    Raises:
        NISRAValidationError: If data validation fails
    """
    wb = load_workbook(file_path, data_only=True)

    if "Table 4" not in wb.sheetnames:
        raise NISRADataNotFoundError("Table 4 (place) not found in file")

    sheet = wb["Table 4"]

    # Find header row
    header_row = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if "Week Ending" in str(row[1] or ""):
            header_row = row_idx
            break

    if header_row is None:
        raise NISRADataNotFoundError("Could not find header row in Table 4")

    # Extract place names from header
    headers = [cell.value for cell in sheet[header_row]]

    # Clean place names (remove newlines, etc.)
    place_columns = {}
    for idx in range(2, len(headers)):
        if headers[idx] and headers[idx] != "Total":
            place_name = str(headers[idx]).replace("\n", " ").strip()
            place_columns[idx] = place_name

    records = []

    # Read data rows
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        week_ending = row[1]

        # Skip if no week ending date
        if not week_ending or not isinstance(week_ending, datetime):
            continue

        # Parse each place column
        for col_idx, place_name in place_columns.items():
            deaths = row[col_idx]

            if deaths is not None and deaths != "":
                records.append({"week_ending": week_ending, "place_of_death": place_name, "deaths": int(deaths)})

    wb.close()

    df = pd.DataFrame(records)
    df["week_ending"] = pd.to_datetime(df["week_ending"])

    # Validate
    _validate_place(df)

    return df


def parse_deaths_file(
    file_path: Union[str, Path], dimension: DimensionType = "all"
) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
    """Parse weekly deaths file for one or all dimensions.

    Args:
        file_path: Path to the weekly deaths Excel file
        dimension: Which dimension(s) to parse:
            - 'totals': Weekly totals with COVID, flu/pneumonia, excess deaths
            - 'demographics': Age and sex breakdown
            - 'geography': Local Government Districts
            - 'place': Place of death
            - 'all': All dimensions (returns dict)

    Returns:
        DataFrame for single dimension, or dict of DataFrames for 'all'

    Example:
        >>> df = parse_deaths_file('deaths.xlsx', dimension='totals')
        >>> data = parse_deaths_file('deaths.xlsx', dimension='all')
        >>> print(data['totals'].head())
    """
    if dimension == "all":
        return {
            "totals": parse_deaths_totals(file_path),
            "demographics": parse_deaths_demographics(file_path),
            "geography": parse_deaths_geography(file_path),
            "place": parse_deaths_place(file_path),
        }
    elif dimension == "totals":
        return parse_deaths_totals(file_path)
    elif dimension == "demographics":
        return parse_deaths_demographics(file_path)
    elif dimension == "geography":
        return parse_deaths_geography(file_path)
    elif dimension == "place":
        return parse_deaths_place(file_path)
    else:
        raise ValueError(f"Invalid dimension: {dimension}. Must be one of: totals, demographics, geography, place, all")


def get_latest_deaths(
    dimension: DimensionType = "all", force_refresh: bool = False
) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
    """Get the latest weekly deaths data.

    Args:
        dimension: Which dimension(s) to retrieve
        force_refresh: Force re-download even if cached

    Returns:
        DataFrame or dict of DataFrames depending on dimension

    Example:
        >>> df = get_latest_deaths(dimension='demographics')
        >>> print(f"Latest week: {df['week_ending'].max()}")
    """
    url = get_latest_weekly_deaths_url()
    file_path = download_file(url, cache_ttl_hours=7 * 24, force_refresh=force_refresh)
    return parse_deaths_file(file_path, dimension=dimension)


def get_historical_deaths(
    years: Optional[List[int]] = None, force_refresh: bool = False, include_age_breakdowns: bool = False
) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
    """Get historical weekly deaths data (2011-2024).

    Downloads and parses the NISRA historical weekly deaths file which contains
    final (not provisional) data for multiple years. Includes total deaths,
    respiratory deaths, COVID-19 deaths, flu/pneumonia deaths, and age breakdowns.

    Args:
        years: List of years to include (default: all available years)
        force_refresh: Force re-download even if cached
        include_age_breakdowns: If True, returns dict with 'totals' and 'age_breakdowns' DataFrames
                               If False (default), returns only totals DataFrame

    Returns:
        If include_age_breakdowns=False:
            DataFrame with columns:
                - year: Year
                - week_number: Week number in year
                - week_ending: Friday of the reporting week
                - total_deaths: Total deaths registered in week
                - expected_deaths_5yr: Average deaths over previous 5 years
                - excess_deaths: Observed minus expected deaths
                - respiratory_deaths_involving: Deaths involving respiratory diseases
                - flu_pneumonia_deaths_involving: Deaths involving flu/pneumonia
                - covid_deaths_involving: Deaths involving COVID-19
                - covid_deaths_due_to: Deaths due to COVID-19 (underlying cause)

        If include_age_breakdowns=True:
            Dict with:
                'totals': DataFrame as above
                'age_breakdowns': DataFrame in long format with columns:
                    - year: Year
                    - week_ending: Friday of the reporting week
                    - age_range: Age range label (e.g., '0-7 days', '15-44', '85+')
                    - deaths: Death count for this age range

    Example:
        >>> # Get totals only
        >>> df = get_historical_deaths()
        >>> print(f"Data spans {df['year'].min()} to {df['year'].max()}")

        >>> # Get totals with age breakdowns
        >>> data = get_historical_deaths(years=[2020, 2021], include_age_breakdowns=True)
        >>> totals = data['totals']
        >>> age_data = data['age_breakdowns']
        >>> # Analyze age distribution
        >>> age_summary = age_data.groupby('age_range')['deaths'].sum()
    """
    # Download the historical file
    links = scrape_download_links(HISTORICAL_DEATHS_URL, file_extension=".xlsx")
    if not links:
        raise NISRADataNotFoundError("No historical deaths Excel file found")

    url = links[0]["url"]
    file_path = download_file(url, cache_ttl_hours=30 * 24, force_refresh=force_refresh)  # Cache for 30 days

    wb = load_workbook(file_path, data_only=True)

    # Determine which years to parse
    available_years = []
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("Weekly Deaths_"):
            year = int(sheet_name.split("_")[1])
            available_years.append(year)

    if years is None:
        years = sorted(available_years)
    else:
        # Validate requested years
        invalid_years = set(years) - set(available_years)
        if invalid_years:
            raise ValueError(f"Years not available: {invalid_years}. Available: {available_years}")

    logger.info(f"Parsing historical deaths for years: {years}")

    all_records = []
    age_breakdown_records = []

    # Define age column mapping (consistent across years, but this makes it flexible)
    # These are standard column indices for the historical file format
    AGE_COLUMNS = {
        "0-7 days": 20,
        "7 days-1 year": 21,
        "1-14": 22,
        "15-44": 23,
        "45-64": 24,
        "65-74": 25,
        "75-84": 26,
        "85+": 27,
    }

    for year in years:
        sheet = wb[f"Weekly Deaths_{year}"]

        # Header is in row 4, data starts at row 5
        # Using standardized column indices for the historical format
        # (these are stable across years in the historical file)

        for row in sheet.iter_rows(min_row=5, values_only=True):
            week_number = row[0]
            week_ending = row[2]  # Week Ends (Friday)

            if not week_ending or not isinstance(week_ending, datetime):
                break

            # Parse main totals record
            all_records.append(
                {
                    "year": year,
                    "week_number": week_number,
                    "week_ending": week_ending,
                    "total_deaths": safe_int(row[3]),
                    "expected_deaths_5yr": safe_float(row[4]),
                    "expected_deaths_current": safe_float(row[8]),
                    "excess_deaths": safe_float(row[11]),
                    "respiratory_deaths_involving": safe_int(row[14]),
                    "flu_pneumonia_deaths_involving": safe_int(row[16]),
                    "covid_deaths_involving": safe_int(row[18]) or 0,
                    "covid_deaths_due_to": safe_int(row[19]) or 0,
                }
            )

            # Parse age breakdowns if requested
            if include_age_breakdowns:
                for age_range, col_idx in AGE_COLUMNS.items():
                    deaths = safe_int(row[col_idx])
                    if deaths is not None:  # Only include if we have data
                        age_breakdown_records.append(
                            {"year": year, "week_ending": week_ending, "age_range": age_range, "deaths": deaths}
                        )

    wb.close()

    # Create totals DataFrame
    df_totals = pd.DataFrame(all_records)
    df_totals["week_ending"] = pd.to_datetime(df_totals["week_ending"])

    logger.info(
        f"Historical deaths parsed: {len(df_totals)} weeks from {df_totals['year'].min()} to {df_totals['year'].max()}"
    )
    logger.info(f"Total COVID deaths (involving): {df_totals['covid_deaths_involving'].sum()}")

    if include_age_breakdowns:
        df_age = pd.DataFrame(age_breakdown_records)
        df_age["week_ending"] = pd.to_datetime(df_age["week_ending"])
        logger.info(f"Age breakdowns parsed: {len(df_age)} records across {df_age['age_range'].nunique()} age ranges")

        return {"totals": df_totals, "age_breakdowns": df_age}
    else:
        return df_totals


def get_combined_deaths(
    years: Optional[List[int]] = None, include_current_year: bool = True, force_refresh: bool = False
) -> pd.DataFrame:
    """Get combined historical and current year deaths data.

    Combines the historical deaths file (2011-2024) with the current year's
    provisional data to give a complete time series including the most recent weeks.

    Args:
        years: List of years to include from historical data (default: last 5 years + current)
        include_current_year: Include current year provisional data (default: True)
        force_refresh: Force re-download of both historical and current files

    Returns:
        DataFrame with columns:
            - year: Year
            - week_number: Week number in year
            - week_ending: Friday of the reporting week
            - total_deaths: Total deaths registered in week
            - covid_deaths: COVID-19 deaths (from current year) or covid_deaths_involving (historical)
            - flu_pneumonia_deaths: Flu/pneumonia deaths
            - excess_deaths: Excess deaths (observed - expected)
            - data_source: 'historical' or 'current'

    Example:
        >>> # Get past 5 years plus 2025 YTD
        >>> df = get_combined_deaths()
        >>> print(f"Data from {df['week_ending'].min().date()} to {df['week_ending'].max().date()}")

        >>> # Create multi-year visualizations
        >>> import plotly.express as px
        >>> fig = px.line(df, x='week_ending', y='total_deaths', color='year')
    """
    current_year = datetime.now().year

    # Default to last 5 complete years if not specified
    if years is None:
        years = list(range(current_year - 5, current_year))

    # Get historical data for specified years
    df_historical = get_historical_deaths(years=years, force_refresh=force_refresh)

    # Standardize column names for combining
    df_historical = df_historical.rename(
        columns={"covid_deaths_involving": "covid_deaths", "flu_pneumonia_deaths_involving": "flu_pneumonia_deaths"}
    )
    df_historical["data_source"] = "historical"

    # Add current year if requested
    if include_current_year:
        try:
            # Get current year totals (includes COVID, flu/pneumonia, excess deaths)
            df_current = get_latest_deaths(dimension="totals", force_refresh=force_refresh)

            # Add year and week_of_year
            df_current["year"] = df_current["week_ending"].dt.year
            df_current["data_source"] = "current"

            # Standardize columns to match historical
            df_current = df_current.rename(
                columns={"observed_deaths": "total_deaths", "excess_deaths_current": "excess_deaths"}
            )

            # Select matching columns
            common_cols = [
                "year",
                "week_number",
                "week_ending",
                "total_deaths",
                "covid_deaths",
                "flu_pneumonia_deaths",
                "excess_deaths",
                "data_source",
            ]

            # Combine
            df_combined = pd.concat([df_historical[common_cols], df_current[common_cols]], ignore_index=True)

            logger.info(f"Combined data: {len(df_historical)} historical weeks + {len(df_current)} current year weeks")

        except Exception as e:
            logger.warning(f"Could not get current year data: {e}. Returning historical only.")
            df_combined = df_historical
    else:
        df_combined = df_historical

    # Sort by date
    df_combined = df_combined.sort_values("week_ending").reset_index(drop=True)

    return df_combined


# Helper functions


def _parse_week_ending(week_header: str) -> datetime:
    """Parse week ending date from Excel header.

    Example headers:
    - "Week 1\n 3 January 2025"
    - "Week 2\n 10 January 2025"
    """
    import re

    from dateutil import parser

    # Extract date part (after newline)
    if "\n" in week_header:
        date_str = week_header.split("\n")[1].strip()
    else:
        date_str = week_header

    # Try to parse with dateutil
    try:
        return parser.parse(date_str)
    except (ValueError, TypeError):
        # Fallback: extract with regex
        match = re.search(r"(\d{1,2})\s+(\w+)\s+(\d{4})", date_str)
        if match:
            day, month, year = match.groups()
            date_str = f"{day} {month} {year}"
            return parser.parse(date_str)

    raise ValueError(f"Could not parse date from: {week_header}")


def _validate_demographics(df: pd.DataFrame):
    """Validate demographics data against expected totals."""
    # For each week, Male + Female should equal Total
    for week in df["week_ending"].unique():
        week_data = df[(df["week_ending"] == week) & (df["age_range"] == "All")]

        total_deaths = week_data[week_data["sex"].str.contains("Total", na=False)]["deaths"].sum()
        male_deaths = week_data[
            week_data["sex"].str.contains("Male", na=False) & ~week_data["sex"].str.contains("Female", na=False)
        ]["deaths"].sum()
        female_deaths = week_data[week_data["sex"].str.contains("Female", na=False)]["deaths"].sum()

        if total_deaths != (male_deaths + female_deaths):
            raise NISRAValidationError(
                f"Week {week}: Demographics validation failed. "
                f"Total={total_deaths}, Male+Female={male_deaths + female_deaths}"
            )

    logger.info("Demographics validation passed")


def _validate_geography(df: pd.DataFrame):
    """Validate geography data."""
    # Basic validation: ensure we have expected LGDs
    lgds = df["lgd"].unique()

    expected_min_lgds = 10  # At least 10 LGDs expected

    if len(lgds) < expected_min_lgds:
        logger.warning(f"Only found {len(lgds)} LGDs, expected at least {expected_min_lgds}")

    logger.info(f"Geography validation passed ({len(lgds)} LGDs found)")


def _validate_place(df: pd.DataFrame):
    """Validate place of death data."""
    # Basic validation: ensure we have expected places
    places = df["place_of_death"].unique()

    expected_places = {"Hospital", "Home"}  # At minimum

    missing = expected_places - set(places)
    if missing:
        logger.warning(f"Missing expected places: {missing}")

    logger.info(f"Place validation passed ({len(places)} places found)")
