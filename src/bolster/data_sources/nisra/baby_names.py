"""NISRA Baby Names Northern Ireland Data Source.

Provides access to baby name statistics for Northern Ireland from the Northern Ireland
Statistics and Research Agency (NISRA), including:
- Full historical list of all first forenames given to babies registered in NI (1997–present)
- Annual rank and count for every name, by sex (Boys/Girls)

The module uses the Full Name List file which contains all registered names with their
rank and count for each year from 1997 to the most recent publication year.

Data Source:
    **Statistics Page**: https://www.nisra.gov.uk/statistics/births/baby-names

    The statistics page lists all Baby Names publications in reverse chronological order
    (newest first). The module automatically scrapes this page to find the latest
    Baby Names publication, then downloads the Full Names List Excel file from that
    publication's detail page.

    The full names list files contain complete time series from 1997 to the most recent
    year, updated annually in April.

Update Frequency: Annual (published April each year)
Geographic Coverage: Northern Ireland (births registered in NI)

Example:
    >>> from bolster.data_sources.nisra import baby_names
    >>> df = baby_names.get_baby_names()
    >>> sorted(df.columns.tolist())
    ['count', 'name', 'rank', 'sex', 'year']
    >>> sorted(df['sex'].unique().tolist())
    ['Boys', 'Girls']
    >>> df['year'].min() >= 1997
    True
"""

import logging
import re
from pathlib import Path

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from bolster.utils.web import session

from ._base import (
    NISRADataNotFoundError,
    NISRAValidationError,
    download_file,
    safe_int,
)

logger = logging.getLogger(__name__)

# Statistics page listing all baby names publications
BABY_NAMES_STATS_URL = "https://www.nisra.gov.uk/statistics/births/baby-names"
NISRA_BASE_URL = "https://www.nisra.gov.uk"

# Header row in the Full Name List Tables (1-indexed)
_FULL_LIST_HEADER_ROW = 6
# First data row (1-indexed)
_FULL_LIST_DATA_START_ROW = 7
# Columns per year block (Name, Number of Babies, Rank)
_COLS_PER_YEAR = 3


def get_baby_names_publication_url() -> str:
    """Scrape NISRA to find the latest Baby Names Full Name List Excel URL.

    Navigates the publication structure:
    1. Scrapes the baby names statistics page for the latest publication link
    2. Follows link to the publication detail page
    3. Finds the Full Names List Excel file link

    Returns:
        URL of the latest Full Name List Excel file

    Raises:
        NISRADataNotFoundError: If publication or file not found
    """
    try:
        response = session.get(BABY_NAMES_STATS_URL, timeout=30)
        response.raise_for_status()
    except Exception as e:
        raise NISRADataNotFoundError(f"Failed to fetch baby names statistics page: {e}") from e

    soup = BeautifulSoup(response.content, "html.parser")

    # Find the most recent "Baby Names YYYY" publication link (first match = latest)
    pub_link = None
    pub_year = None
    for link in soup.find_all("a", href=True):
        href = link["href"]
        text = link.get_text(strip=True)
        # Match links like /publications/baby-names-2025
        if re.search(r"/publications/baby-names-\d{4}", href):
            # Extract year from href
            m = re.search(r"baby-names-(\d{4})", href)
            year = int(m.group(1)) if m else 0
            if pub_year is None or year > pub_year:
                pub_year = year
                pub_link = href
                logger.info(f"Found Baby Names publication: {text.strip()!r} ({href})")

    if not pub_link:
        raise NISRADataNotFoundError("Could not find Baby Names publication link on statistics page")

    # Make absolute URL
    if pub_link.startswith("/"):
        pub_link = f"{NISRA_BASE_URL}{pub_link}"

    # Fetch the publication detail page
    try:
        pub_response = session.get(pub_link, timeout=30)
        pub_response.raise_for_status()
    except Exception as e:
        raise NISRADataNotFoundError(f"Failed to fetch publication page {pub_link}: {e}") from e

    pub_soup = BeautifulSoup(pub_response.content, "html.parser")

    # Find the "Full Names List" Excel file link
    # Link text is like "Full Names List, 1997 to 2025 TablesMicrosoft Excel ..."
    excel_url = None
    for link in pub_soup.find_all("a", href=True):
        href = link["href"]
        text = link.get_text(strip=True)
        if href.lower().endswith(".xlsx") and "full" in text.lower() and "name" in text.lower():
            excel_url = href
            logger.info(f"Found Full Name List file: {text!r} -> {href}")
            break

    # Fallback: any xlsx link with "Full_Name_List" in the path
    if not excel_url:
        for link in pub_soup.find_all("a", href=True):
            href = link["href"]
            if href.lower().endswith(".xlsx") and "full_name_list" in href.lower():
                excel_url = href
                logger.info(f"Found Full Name List file (fallback): {href}")
                break

    if not excel_url:
        raise NISRADataNotFoundError(f"Could not find Full Name List Excel file on publication page: {pub_link}")

    # Make absolute URL
    if excel_url.startswith("/"):
        excel_url = f"{NISRA_BASE_URL}{excel_url}"

    return excel_url


def parse_baby_names_file(file_path: str | Path) -> pd.DataFrame:
    """Parse NISRA Full Name List Excel file into long-format DataFrame.

    The Full Name List Excel file contains two sheets:
    - Table 1: Boys' names (1997 to present), wide format with 3 columns per year
                (Name, Number of Babies, Rank), 29+ year blocks across the row
    - Table 2: Girls' names, same structure as Table 1

    Names with suppressed counts (shown as '..') are excluded (names with fewer
    than 3 occurrences are suppressed for disclosure control).

    Args:
        file_path: Path to the Full Name List Excel file

    Returns:
        Long-format DataFrame with columns:

        - year: int — registration year
        - name: str — first forename (title case as registered)
        - sex: str — "Boys" or "Girls"
        - rank: int — rank within that sex and year (1 = most popular)
        - count: int — number of babies registered with that name

    Raises:
        NISRAValidationError: If the file structure is unexpected or no data parsed
    """
    file_path = Path(file_path)
    wb = load_workbook(file_path, data_only=True, read_only=True)

    sheet_sex_map = {
        "Table 1": "Boys",
        "Table 2": "Girls",
    }

    all_records: list[dict] = []

    for sheet_name, sex in sheet_sex_map.items():
        if sheet_name not in wb.sheetnames:
            raise NISRAValidationError(f"Expected sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

        ws = wb[sheet_name]
        records = _parse_full_name_list_sheet(ws, sex)
        all_records.extend(records)
        logger.info(f"Parsed {len(records)} records from {sheet_name} ({sex})")

    if not all_records:
        raise NISRAValidationError("No records parsed from file — check file structure")

    df = pd.DataFrame(all_records)
    df["year"] = df["year"].astype(int)
    df["rank"] = df["rank"].astype(int)
    df["count"] = df["count"].astype(int)
    df = df.sort_values(["year", "sex", "rank"]).reset_index(drop=True)

    logger.info(f"Parsed {len(df)} total records covering {df['year'].min()}–{df['year'].max()}")
    return df


def _parse_full_name_list_sheet(ws, sex: str) -> list[dict]:
    """Parse a single sex table (Table 1 or Table 2) from the Full Name List workbook.

    The sheet layout is:
    - Rows 1–5: Cover notes / metadata
    - Row 6: Headers — "YYYY Name", "Number of Babies", "Rank", repeated for each year
    - Row 7+: Data — one name per row position, repeated for each year block

    Each year occupies exactly 3 columns. Years run left-to-right in chronological order.
    Names within each year block are sorted by rank (ascending).

    Args:
        ws: openpyxl worksheet (read_only)
        sex: "Boys" or "Girls"

    Returns:
        List of record dicts with keys: year, name, sex, rank, count
    """
    # Read header row to extract years
    headers = list(ws.iter_rows(min_row=_FULL_LIST_HEADER_ROW, max_row=_FULL_LIST_HEADER_ROW, values_only=True))[0]

    # Build list of (col_index, year) for each year block
    year_cols: list[tuple[int, int]] = []
    for col_idx in range(0, len(headers), _COLS_PER_YEAR):
        cell_val = headers[col_idx]
        if cell_val is None:
            break
        # Header format: "1997 Name", "1998 Name", etc.
        m = re.match(r"(\d{4})\s+Name", str(cell_val).strip())
        if m:
            year = int(m.group(1))
            year_cols.append((col_idx, year))

    if not year_cols:
        raise NISRAValidationError(f"Could not parse year columns from header row in sheet for {sex}")

    logger.debug(f"{sex}: Found {len(year_cols)} year columns, {year_cols[0][1]}–{year_cols[-1][1]}")

    records: list[dict] = []

    # Read all data rows at once
    for row in ws.iter_rows(min_row=_FULL_LIST_DATA_START_ROW, values_only=True):
        # Skip completely empty rows
        if all(v is None for v in row):
            continue

        for col_idx, year in year_cols:
            # Each year block: [Name, Number of Babies, Rank]
            if col_idx + 2 >= len(row):
                continue

            name = row[col_idx]
            count_val = row[col_idx + 1]
            rank_val = row[col_idx + 2]

            # Skip suppressed ('..' placeholder) or missing values
            if name is None or str(name).strip() in ("", "..", "-", "Contents"):
                continue
            if count_val is None or str(count_val).strip() in ("", "..", "-"):
                continue
            if rank_val is None or str(rank_val).strip() in ("", "..", "-"):
                continue

            count = safe_int(count_val)
            rank = safe_int(rank_val)

            if count is None or rank is None:
                continue

            records.append(
                {
                    "year": year,
                    "name": str(name).strip(),
                    "sex": sex,
                    "rank": rank,
                    "count": count,
                }
            )

    return records


def get_baby_names(force_refresh: bool = False) -> pd.DataFrame:
    """Get the full historical Baby Names series for Northern Ireland (1997–present).

    Automatically discovers and downloads the most recent Full Name List publication
    from the NISRA website, which contains the complete historical series from 1997.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        Long-format DataFrame with columns:

        - year: int — registration year (1997–present)
        - name: str — first forename as registered
        - sex: str — "Boys" or "Girls"
        - rank: int — rank within sex and year (1 = most popular)
        - count: int — number of babies with that name

    Raises:
        NISRADataNotFoundError: If the latest publication cannot be found
        NISRAValidationError: If the file structure is unexpected

    Example:
        >>> df = get_baby_names()
        >>> sorted(df.columns.tolist())
        ['count', 'name', 'rank', 'sex', 'year']
        >>> df['year'].min() >= 1997
        True
        >>> sorted(df['sex'].unique().tolist())
        ['Boys', 'Girls']
        >>> df[df['year'] == df['year'].max()].nsmallest(1, 'rank')['name'].iloc[0] is not None
        True
    """
    excel_url = get_baby_names_publication_url()
    logger.info(f"Downloading baby names data from: {excel_url}")

    # Cache for 180 days — published once per year in April
    file_path = download_file(excel_url, cache_ttl_hours=180 * 24, force_refresh=force_refresh)

    return parse_baby_names_file(file_path)


def validate_baby_names(df: pd.DataFrame) -> bool:
    """Validate a baby names DataFrame for structural and data integrity.

    Checks:
    - Required columns are present
    - No null values in any column
    - Both sexes present ("Boys" and "Girls")
    - Year range starts at or before 1999 (data should go back to 1997)
    - Rank starts at 1 for at least one year/sex combination
    - All counts are positive (> 0)
    - No negative ranks or counts

    Args:
        df: DataFrame to validate (from parse_baby_names_file or get_baby_names)

    Returns:
        True if validation passes

    Raises:
        NISRAValidationError: If any validation check fails, with descriptive message

    Example:
        >>> import pandas as pd
        >>> valid_df = pd.DataFrame({
        ...     'year': [2020, 2020], 'name': ['Noah', 'Jack'],
        ...     'sex': ['Boys', 'Boys'], 'rank': [1, 2], 'count': [100, 90]
        ... })
        >>> validate_baby_names(valid_df)
        True
    """
    required_columns = {"year", "name", "sex", "rank", "count"}

    # Check empty DataFrame
    if df is None or df.empty:
        raise NISRAValidationError("DataFrame is empty")

    # Check required columns
    missing_cols = required_columns - set(df.columns)
    if missing_cols:
        raise NISRAValidationError(f"Missing required columns: {missing_cols}")

    # Check for null values
    null_counts = df[list(required_columns)].isnull().sum()
    cols_with_nulls = null_counts[null_counts > 0]
    if not cols_with_nulls.empty:
        raise NISRAValidationError(f"Columns with null values: {cols_with_nulls.to_dict()}")

    # Check both sexes present
    actual_sexes = set(df["sex"].unique())
    if "Boys" not in actual_sexes:
        raise NISRAValidationError(f"Missing 'Boys' sex category. Found: {actual_sexes}")
    if "Girls" not in actual_sexes:
        raise NISRAValidationError(f"Missing 'Girls' sex category. Found: {actual_sexes}")

    # Check year range
    min_year = df["year"].min()
    if min_year > 1999:
        raise NISRAValidationError(
            f"Year range starts at {min_year} — expected data back to at least 1999 (ideally 1997)"
        )

    # Check no non-positive counts (before checking rank range)
    bad_counts = (df["count"] <= 0).sum()
    if bad_counts > 0:
        raise NISRAValidationError(f"{bad_counts} rows have non-positive count values")

    # Check no non-positive ranks (before checking min_rank == 1)
    bad_ranks = (df["rank"] <= 0).sum()
    if bad_ranks > 0:
        raise NISRAValidationError(f"{bad_ranks} rows have non-positive rank values")

    # Check rank starts at 1
    min_rank = df["rank"].min()
    if min_rank != 1:
        raise NISRAValidationError(f"Minimum rank is {min_rank} — expected 1")

    logger.info(
        f"Validation passed: {len(df):,} records, years {df['year'].min()}–{df['year'].max()}, "
        f"sexes: {sorted(actual_sexes)}"
    )
    return True
