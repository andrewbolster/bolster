"""NISRA Annual Survey of Hours and Earnings (ASHE) Module.

This module provides access to Northern Ireland's earnings statistics:
- Median weekly, hourly, and annual earnings
- Breakdowns by employment type, sector, geography, occupation, industry
- Gender pay gap analysis
- Historical timeseries from 1997 to present

Data is published annually in October by NISRA's Economic & Labour Market Statistics Branch.

Data Source: Northern Ireland Statistics and Research Agency provides Annual Survey of Hours
and Earnings statistics through their Work, Pay and Benefits section at
https://www.nisra.gov.uk/statistics/work-pay-and-benefits/annual-survey-hours-and-earnings.
ASHE data covers employee earnings across all sectors based on a sample survey of payroll
records from HMRC PAYE data, providing comprehensive earnings statistics for Northern Ireland.

Update Frequency: Annual publications released in October each year, covering earnings data
for the reference period of April. The dataset provides the most comprehensive and official
source of employee earnings statistics for Northern Ireland, updated once per year with
historical revisions as necessary.

Data Coverage:
    - Weekly Earnings: 1997 - Present (annual, full-time/part-time/all)
    - Hourly Earnings: 1997 - Present (annual, excluding overtime)
    - Annual Earnings: 1999 - Present (annual, full-time/part-time/all)
    - Geographic: 11 Local Government Districts (workplace vs residence basis)
    - Sector: Public vs Private sector comparison (2005 - Present)

Examples:
    >>> from bolster.data_sources.nisra import ashe
    >>> # Get latest weekly earnings timeseries
    >>> df = ashe.get_latest_ashe_timeseries(metric='weekly')
    >>> sorted(df.columns.tolist())
    ['median_weekly_earnings', 'work_pattern', 'year']

    >>> # Get geographic earnings by workplace
    >>> df_geo = ashe.get_latest_ashe_geography(basis='workplace')
    >>> 'median_weekly_earnings' in df_geo.columns
    True

    >>> # Get public vs private sector comparison
    >>> df_sector = ashe.get_latest_ashe_sector()
    >>> 'location' in df_sector.columns
    True

Publication Details:
    - Frequency: Annual (October publication)
    - Reference period: April of each year
    - Published by: NISRA Economic & Labour Market Statistics Branch
    - Contact: economicstats@nisra.gov.uk
    - Base: Employee jobs in Northern Ireland (not self-employed)
"""

import logging
import re
from datetime import datetime
from pathlib import Path

import pandas as pd

from bolster.utils.web import session

from ._base import NISRADataNotFoundError, download_file

logger = logging.getLogger(__name__)

# Base URL for NISRA ASHE statistics
ASHE_BASE_URL = "https://www.nisra.gov.uk/statistics/work-pay-and-benefits/annual-survey-hours-and-earnings"


def get_latest_ashe_publication_url() -> tuple[str, int]:
    """Get the URL of the latest ASHE publication and its year.

    Scrapes the NISRA ASHE page to find the most recent publication.

    Returns:
        Tuple of (publication_url, year)

    Raises:
        NISRADataNotFoundError: If unable to find the latest publication

    Example:
        >>> url, year = get_latest_ashe_publication_url()
        >>> url.startswith('https://')
        True
    """
    from bs4 import BeautifulSoup

    logger.info("Fetching latest ASHE publication URL...")

    try:
        response = session.get(ASHE_BASE_URL, timeout=30)
        response.raise_for_status()
    except Exception as e:
        raise NISRADataNotFoundError(f"Failed to fetch ASHE page: {e}") from e

    soup = BeautifulSoup(response.content, "html.parser")

    # Find the link to the latest publication
    # Pattern: "Employee earnings in NI 2025" or "Employee earnings in Northern Ireland 2025"
    publication_links = soup.find_all("a", href=True)

    for link in publication_links:
        link_text = link.get_text(strip=True)
        # Match "Employee earnings in Northern Ireland YYYY" or "Employee earnings in NI YYYY"
        match = re.search(r"Employee earnings in (?:Northern Ireland|NI)\s+(\d{4})", link_text)
        if match:
            year = int(match.group(1))
            pub_url = link["href"]
            if not pub_url.startswith("http"):
                pub_url = f"https://www.nisra.gov.uk{pub_url}"

            logger.info(f"Found latest ASHE publication: {year} at {pub_url}")
            return pub_url, year

    raise NISRADataNotFoundError("Could not find latest ASHE publication")


def get_ashe_file_url(year: int, file_type: str = "timeseries") -> str:
    """Construct URL for ASHE file based on year and file type.

    Args:
        year: Publication year (e.g., 2025)
        file_type: Type of file - 'timeseries' or 'linked'

    Returns:
        URL to the Excel file

    Example:
        >>> url = get_ashe_file_url(2025, 'timeseries')
        >>> url.startswith('https://')
        True
    """
    # ASHE is published in October
    month = 10

    if file_type == "timeseries":
        # Pattern: ASHE-1997-{year}-headline-timeseries.xlsx
        filename = f"ASHE-1997-{year}-headline-timeseries.xlsx"
    elif file_type == "linked":
        # Pattern: ASHE-{year}-linked.xlsx
        filename = f"ASHE-{year}-linked.xlsx"
    else:
        raise ValueError(f"Unknown file_type: {file_type}. Use 'timeseries' or 'linked'")

    url = f"https://www.nisra.gov.uk/system/files/statistics/{year}-{month:02d}/{filename}"
    logger.info(f"Constructed ASHE file URL: {url}")
    return url


def parse_ashe_timeseries_weekly(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE weekly earnings timeseries.

    Extracts the weekly earnings data from the timeseries Excel file.

    Args:
        file_path: Path to the ASHE timeseries Excel file

    Returns:
        DataFrame with columns:
            - year: int
            - work_pattern: str ('Full-time', 'Part-time', 'All')
            - median_weekly_earnings: float (£)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'timeseries'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_timeseries_weekly(path)
        >>> sorted(df.columns.tolist())
        ['median_weekly_earnings', 'work_pattern', 'year']
    """
    logger.info(f"Parsing ASHE weekly earnings from: {file_path}")

    # Read Weekly sheet, skip first 4 rows to get to header
    df = pd.read_excel(file_path, sheet_name="Weekly", skiprows=4)

    # Convert to long format
    records = []
    for _, row in df.iterrows():
        year = int(row["Year"])
        for work_pattern in ["Full-time", "Part-time", "All"]:
            records.append(
                {"year": year, "work_pattern": work_pattern, "median_weekly_earnings": float(row[work_pattern])}
            )

    result = pd.DataFrame(records)
    logger.info(f"Parsed {len(result)} weekly earnings records ({result['year'].min()}-{result['year'].max()})")
    return result


def parse_ashe_timeseries_hourly(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE hourly earnings timeseries.

    Extracts the hourly earnings data (excluding overtime) from the timeseries Excel file.

    Args:
        file_path: Path to the ASHE timeseries Excel file

    Returns:
        DataFrame with columns:
            - year: int
            - work_pattern: str ('Full-time', 'Part-time', 'All')
            - median_hourly_earnings: float (£)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'timeseries'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_timeseries_hourly(path)
        >>> sorted(df.columns.tolist())
        ['median_hourly_earnings', 'work_pattern', 'year']
    """
    logger.info(f"Parsing ASHE hourly earnings from: {file_path}")

    # Read Hourly sheet, skip first 4 rows to get to header
    df = pd.read_excel(file_path, sheet_name="Hourly", skiprows=4)

    # Convert to long format
    records = []
    for _, row in df.iterrows():
        year = int(row["Year"])
        for work_pattern in ["Full-time", "Part-time", "All"]:
            records.append(
                {"year": year, "work_pattern": work_pattern, "median_hourly_earnings": float(row[work_pattern])}
            )

    result = pd.DataFrame(records)
    logger.info(f"Parsed {len(result)} hourly earnings records ({result['year'].min()}-{result['year'].max()})")
    return result


def parse_ashe_timeseries_annual(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE annual earnings timeseries.

    Extracts the annual earnings data from the timeseries Excel file.

    Args:
        file_path: Path to the ASHE timeseries Excel file

    Returns:
        DataFrame with columns:
            - year: int
            - work_pattern: str ('Full-time', 'Part-time', 'All')
            - median_annual_earnings: float (£)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'timeseries'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_timeseries_annual(path)
        >>> sorted(df.columns.tolist())
        ['median_annual_earnings', 'work_pattern', 'year']
    """
    logger.info(f"Parsing ASHE annual earnings from: {file_path}")

    # Read Annual sheet, skip first 4 rows to get to header
    df = pd.read_excel(file_path, sheet_name="Annual", skiprows=4)

    # Convert to long format
    records = []
    for _, row in df.iterrows():
        year = int(row["Year"])
        for work_pattern in ["Full-time", "Part-time", "All"]:
            records.append(
                {"year": year, "work_pattern": work_pattern, "median_annual_earnings": float(row[work_pattern])}
            )

    result = pd.DataFrame(records)
    logger.info(f"Parsed {len(result)} annual earnings records ({result['year'].min()}-{result['year'].max()})")
    return result


def parse_ashe_geography(file_path: str | Path, basis: str = "workplace", year: int = None) -> pd.DataFrame:
    """Parse ASHE geographic earnings data.

    Extracts earnings by Local Government District from the linked tables file.

    Args:
        file_path: Path to the ASHE linked tables Excel file
        basis: 'workplace' (MapA) or 'residence' (MapB)
        year: Year of the data (if not provided, will be extracted from file)

    Returns:
        DataFrame with columns:
            - year: int
            - lgd: str (Local Government District name)
            - basis: str ('workplace' or 'residence')
            - median_weekly_earnings: float (£)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_geography(path, basis='workplace', year=year)
        >>> 'median_weekly_earnings' in df.columns
        True
    """
    logger.info(f"Parsing ASHE geography ({basis}) from: {file_path}")

    # Select the correct sheet
    sheet_name = "MapA" if basis == "workplace" else "MapB"

    # Read the sheet, skip first 2 rows to get to data
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=2)

    # The first column has LGD names, second column has earnings
    df.columns = ["lgd", "median_weekly_earnings"]

    # Remove any NaN rows
    df = df.dropna()

    # Extract year from data if not provided
    if year is None:
        # Try to extract from filename
        match = re.search(r"ASHE-(\d{4})-linked", str(file_path))
        year = int(match.group(1)) if match else datetime.now().year

    # Add metadata columns
    df["year"] = year
    df["basis"] = basis

    # Reorder columns
    df = df[["year", "lgd", "basis", "median_weekly_earnings"]]

    # Convert earnings to float
    df["median_weekly_earnings"] = df["median_weekly_earnings"].astype(float)

    logger.info(f"Parsed {len(df)} LGD earnings records for {year} ({basis} basis)")
    return df


def parse_ashe_sector(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE public vs private sector earnings.

    Extracts public and private sector earnings timeseries from the linked tables file.

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:
            - year: int
            - location: str ('Northern Ireland' or 'United Kingdom')
            - sector: str ('Public' or 'Private')
            - median_weekly_earnings: float (£)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_sector(path)
        >>> 'location' in df.columns
        True
    """
    logger.info(f"Parsing ASHE sector data from: {file_path}")

    # Read Figure5 sheet, skip first 3 rows to get to header
    df = pd.read_excel(file_path, sheet_name="Figure5", skiprows=3)

    # Columns should be: Year, NI Public, NI Private, UK Public, UK Private
    df.columns = ["Year", "NI Public", "NI Private", "UK Public", "UK Private"]

    # Convert to long format
    records = []
    for _, row in df.iterrows():
        year = int(row["Year"])
        records.append(
            {
                "year": year,
                "location": "Northern Ireland",
                "sector": "Public",
                "median_weekly_earnings": float(row["NI Public"]),
            }
        )
        records.append(
            {
                "year": year,
                "location": "Northern Ireland",
                "sector": "Private",
                "median_weekly_earnings": float(row["NI Private"]),
            }
        )
        records.append(
            {
                "year": year,
                "location": "United Kingdom",
                "sector": "Public",
                "median_weekly_earnings": float(row["UK Public"]),
            }
        )
        records.append(
            {
                "year": year,
                "location": "United Kingdom",
                "sector": "Private",
                "median_weekly_earnings": float(row["UK Private"]),
            }
        )

    result = pd.DataFrame(records)
    logger.info(f"Parsed {len(result)} sector earnings records ({result['year'].min()}-{result['year'].max()})")
    return result


def get_latest_ashe_timeseries(metric: str = "weekly", force_refresh: bool = False) -> pd.DataFrame:
    """Get the latest ASHE timeseries data.

    Downloads and parses the most recent ASHE timeseries publication.
    Results are cached for 90 days unless force_refresh=True.

    Args:
        metric: Type of earnings - 'weekly', 'hourly', or 'annual'
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with timeseries earnings data (1997-present for weekly/hourly, 1999-present for annual)

    Example:
        >>> df = get_latest_ashe_timeseries(metric='weekly')
        >>> sorted(df.columns.tolist())
        ['median_weekly_earnings', 'work_pattern', 'year']
    """
    _, year = get_latest_ashe_publication_url()
    file_url = get_ashe_file_url(year, file_type="timeseries")

    # Cache for 90 days (published annually)
    file_path = download_file(file_url, cache_ttl_hours=90 * 24, force_refresh=force_refresh)

    if metric == "weekly":
        return parse_ashe_timeseries_weekly(file_path)
    if metric == "hourly":
        return parse_ashe_timeseries_hourly(file_path)
    if metric == "annual":
        return parse_ashe_timeseries_annual(file_path)
    raise ValueError(f"Unknown metric: {metric}. Use 'weekly', 'hourly', or 'annual'")


def get_latest_ashe_geography(basis: str = "workplace", force_refresh: bool = False) -> pd.DataFrame:
    """Get the latest ASHE geographic earnings data.

    Downloads and parses the most recent ASHE linked tables publication.
    Results are cached for 90 days unless force_refresh=True.

    Args:
        basis: 'workplace' (where employees work) or 'residence' (where employees live)
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with earnings by Local Government District

    Example:
        >>> df = get_latest_ashe_geography(basis='workplace')
        >>> 'median_weekly_earnings' in df.columns
        True
    """
    _, year = get_latest_ashe_publication_url()
    file_url = get_ashe_file_url(year, file_type="linked")

    # Cache for 90 days (published annually)
    file_path = download_file(file_url, cache_ttl_hours=90 * 24, force_refresh=force_refresh)

    return parse_ashe_geography(file_path, basis=basis, year=year)


def get_latest_ashe_sector(force_refresh: bool = False) -> pd.DataFrame:
    """Get the latest ASHE public vs private sector earnings data.

    Downloads and parses the most recent ASHE linked tables publication.
    Results are cached for 90 days unless force_refresh=True.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with public and private sector earnings timeseries (2005-present)

    Example:
        >>> df = get_latest_ashe_sector()
        >>> 'location' in df.columns
        True
    """
    _, year = get_latest_ashe_publication_url()
    file_url = get_ashe_file_url(year, file_type="linked")

    # Cache for 90 days (published annually)
    file_path = download_file(file_url, cache_ttl_hours=90 * 24, force_refresh=force_refresh)

    return parse_ashe_sector(file_path)


# ============================================================================
# Helper Functions for Analysis
# ============================================================================


def get_earnings_by_year(df: pd.DataFrame, year: int) -> pd.DataFrame:
    """Filter earnings data for a specific year.

    Args:
        df: ASHE DataFrame
        year: Year to filter for

    Returns:
        DataFrame with only the specified year's data

    Example:
        >>> df = get_latest_ashe_timeseries('weekly')
        >>> df_2025 = get_earnings_by_year(df, 2025)
        >>> 'median_weekly_earnings' in df_2025.columns
        True
    """
    return df[df["year"] == year].reset_index(drop=True)


def calculate_growth_rates(df: pd.DataFrame, periods: int = 1) -> pd.DataFrame:
    """Calculate year-on-year growth rates for earnings.

    Args:
        df: ASHE DataFrame with 'year' and earnings column
        periods: Number of years for comparison (default: 1 for YoY)

    Returns:
        DataFrame with additional growth rate column

    Example:
        >>> df = get_latest_ashe_timeseries('weekly')
        >>> df_growth = calculate_growth_rates(df)
        >>> 'earnings_yoy_growth' in df_growth.columns
        True
    """
    result = df.copy()

    # Identify the earnings column
    earnings_col = None
    for col in ["median_weekly_earnings", "median_hourly_earnings", "median_annual_earnings"]:
        if col in result.columns:
            earnings_col = col
            break

    if earnings_col is None:
        raise ValueError("No earnings column found in DataFrame")

    # Calculate growth rate for each work pattern/sector/geography
    # Group by non-year, non-earnings columns
    group_cols = [col for col in result.columns if col not in ["year", earnings_col]]

    if group_cols:
        result["earnings_yoy_growth"] = result.groupby(group_cols)[earnings_col].pct_change(periods=periods) * 100
    else:
        result["earnings_yoy_growth"] = result[earnings_col].pct_change(periods=periods) * 100

    return result


# ---------------------------------------------------------------------------
# Content-fingerprint sheet scanner for the ASHE linked tables file
# ---------------------------------------------------------------------------
# NISRA reassigns figure numbers each year based on editorial focus.
# E.g. "hours distribution" was Figure 3 in 2022-23, Figure 9 in 2024-25.
# We identify sheets by the column signature of their header row and optional
# keywords from the subtitle, making parsers robust across publication years.

_SHEET_SIGNATURES: dict[str, dict] = {
    # Timeseries: Year | UK | NI  — three variants share this column shape;
    # disambiguated by subtitle keyword
    "ni_uk_weekly_earnings": {
        "cols": ("Year", "UK", "NI"),
        "subtitle_keywords": ["weekly", "full-time"],
        "excludes": ["gender pay gap", "working pattern"],
    },
    "gender_pay_gap": {
        "cols": ("Year", "UK", "NI"),
        "subtitle_keywords": ["gender pay gap"],
    },
    "working_pattern_pay_gap": {
        "cols": ("Year", "UK", "NI"),
        "subtitle_keywords": ["working pattern pay gap"],
    },
    # Sector × gender hourly earnings timeseries
    "hourly_by_sector_gender": {
        "cols": ("Year", "Male public", "Female public", "Male private", "Female private"),
        "subtitle_keywords": ["sector", "gender"],
    },
    # Snapshot tables — single year, no 'Year' column
    "hours_distribution": {
        "cols": ("Paid hours worked", "Percentage"),
    },
    "uk_regional_pay_ratio": {
        "cols": ("Region", "Ratio"),
    },
    "hourly_by_age_gender": {
        "cols": ("Age group", "Female", "Male"),
        "subtitle_keywords": ["age"],
    },
    "hourly_by_occupation_gender": {
        "cols": ("Occupation", "Female", "Male"),
        "subtitle_keywords": ["occupation"],
    },
    "hourly_by_pattern_gender": {
        "cols": ("Working pattern", "Female", "Male"),
        "subtitle_keywords": ["working pattern"],
        "excludes": ["pay gap", "mean weekly"],
    },
    "mean_hours_by_pattern_gender": {
        "cols": ("Working pattern", "Males", "Females", "All"),
    },
    # Real earnings timeseries — nominal vs real (latest-year base) weekly earnings
    "real_earnings_timeseries": {
        "cols": ("Year", "Nominal earnings", "Real earnings"),
    },
    # Annual % change in nominal vs real weekly earnings by work pattern (snapshot)
    "real_earnings_change_by_pattern": {
        "cols": ("Work pattern", "Nominal Change (%)", "Real change (%)"),
    },
    # Real earnings index by sector (base latest-year=100). Fig 6 sheet has two
    # tables; both share ('Year','Public','Private'). We want the *index* one
    # which appears first; the second is a real-earnings level table.
    "real_earnings_index_by_sector": {
        "cols": ("Year", "Public", "Private"),
        "subtitle_keywords": ["annual index"],
    },
    # Annual % change in weekly earnings by occupation (NI, full-time).
    # Note: NISRA mislabels the column "Industry" but the rows are occupations;
    # disambiguated from Fig 8 by subtitle keyword "by occupation".
    "annual_change_by_occupation": {
        "cols": ("Industry", "Annual change (%)"),
        "subtitle_keywords": ["by occupation"],
    },
    "annual_change_by_industry": {
        "cols": ("Industry", "Annual change (%)"),
        "subtitle_keywords": ["by industry"],
    },
    # Pay distribution timeseries (low/middle/high)
    "pay_distribution_timeseries": {
        "cols": ("Year", "Low-paid jobs", "Middle-paid jobs", "High-paid jobs"),
    },
    # Pay distribution by classification (work pattern / age / industry / occupation)
    "pay_distribution_by_classification": {
        "cols": ("Classification", "Subset", "Low-paid jobs", "Middle-paid jobs", "High-paid jobs"),
    },
}


def _find_linked_sheet(
    file_path: str | Path,
    signature_key: str,
) -> pd.DataFrame:
    """Locate a sheet in the linked tables file by content fingerprint.

    Walks all sheets, finds the header row (first short-valued row after the
    title rows), checks column names against the signature, and optionally
    confirms subtitle keywords. Returns the DataFrame from the matching sheet.

    Args:
        file_path: Path to the ASHE linked tables Excel file.
        signature_key: Key into ``_SHEET_SIGNATURES``.

    Returns:
        Raw DataFrame from the matched sheet (header not yet renamed).

    Raises:
        NISRADataNotFoundError: If no sheet matches the signature.
    """
    import openpyxl

    sig = _SHEET_SIGNATURES[signature_key]
    required_cols = tuple(c.lower() for c in sig["cols"])
    keywords = [k.lower() for k in sig.get("subtitle_keywords", [])]
    excludes = [e.lower() for e in sig.get("excludes", [])]

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
            if len(rows) < 3:
                continue

            # Title is row 0, subtitle row 1, then possibly a note, then headers
            subtitle = str(rows[1][0] or "").lower()

            # Check keyword/exclude filters on subtitle
            if keywords and not any(k in subtitle for k in keywords):
                continue
            if excludes and any(e in subtitle for e in excludes):
                continue

            # Find header row: first row where all required cols appear
            # (case-insensitive, ignoring None cells)
            for i, row in enumerate(rows):
                row_vals = tuple(str(v).lower() for v in row if v is not None)
                if all(rc in row_vals for rc in required_cols):
                    # Found the header — read everything below it, trimmed to signature width
                    n_sig = len(sig["cols"])
                    data_rows = [r[:n_sig] for r in rows[i + 1 :]]
                    header = [str(v) if v is not None else f"_col{j}" for j, v in enumerate(row[:n_sig])]
                    df = pd.DataFrame(data_rows, columns=header)
                    # Drop entirely-None rows
                    df = df.dropna(how="all")
                    logger.debug(f"Matched signature '{signature_key}' to sheet '{sheet_name}' in {file_path}")
                    wb.close()
                    return df
    finally:
        wb.close()

    raise NISRADataNotFoundError(
        f"No sheet matching signature '{signature_key}' found in {file_path}. Expected columns: {sig['cols']}"
    )


def _get_linked_file(year: int | None = None, force_refresh: bool = False) -> Path:
    """Download the linked tables file, returning path. Uses latest year if not specified."""
    if year is None:
        _, year = get_latest_ashe_publication_url()
    return download_file(get_ashe_file_url(year, "linked"), cache_ttl_hours=90 * 24, force_refresh=force_refresh)


# ---------------------------------------------------------------------------
# Parse functions — each calls _find_linked_sheet with its signature key
# ---------------------------------------------------------------------------


def parse_ashe_gender_pay_gap(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE gender pay gap timeseries (NI and UK), any publication year.

    Extracts the NI and UK all-employee gender pay gap from 2005 to present.
    The gap is defined as the difference between male and female median hourly
    earnings as a percentage of male median hourly earnings (all employees,
    excluding overtime).

    Note: methodological changes occurred in 2006, 2011 and 2021 — these are
    annotated in NISRA publications and should be considered when interpreting
    trend breaks.

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - year: int
        - location: str ('Northern Ireland' or 'United Kingdom')
        - gender_pay_gap_pct: float — GPG as % of male earnings (positive = men paid more)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_gender_pay_gap(path)
        >>> sorted(df.columns.tolist())
        ['gender_pay_gap_pct', 'location', 'year']
    """
    df = _find_linked_sheet(file_path, "gender_pay_gap")
    df.columns = ["year", "UK", "NI"]
    df = df.dropna(subset=["year"])
    df["year"] = pd.to_numeric(df["year"], errors="coerce").dropna().astype(int)
    df = df[df["year"].notna()]

    records = []
    for _, row in df.iterrows():
        if pd.isna(row["year"]):
            continue
        records.append({"year": int(row["year"]), "location": "United Kingdom", "gender_pay_gap_pct": float(row["UK"])})
        records.append(
            {"year": int(row["year"]), "location": "Northern Ireland", "gender_pay_gap_pct": float(row["NI"])}
        )

    result = pd.DataFrame(records).dropna().sort_values(["year", "location"]).reset_index(drop=True)
    result["year"] = result["year"].astype(int)
    logger.info(f"Parsed {len(result)} gender pay gap records ({result['year'].min()}–{result['year'].max()})")
    return result


def parse_ashe_hourly_earnings_by_sector_gender(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE hourly earnings by sector and gender timeseries.

    Identified by column signature ['Year', 'Male public', 'Female public',
    'Male private', 'Female private'] — stable across all publication years.

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - year: int
        - sector: str ('Public' or 'Private')
        - sex: str ('Male' or 'Female')
        - median_hourly_earnings: float (£, excluding overtime)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_hourly_earnings_by_sector_gender(path)
        >>> sorted(df.columns.tolist())
        ['median_hourly_earnings', 'sector', 'sex', 'year']
    """
    df = _find_linked_sheet(file_path, "hourly_by_sector_gender")
    df.columns = ["year", "Male public", "Female public", "Male private", "Female private"]
    df = df.dropna(subset=["year"])
    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df = df[df["year"].notna()]
    df["year"] = df["year"].astype(int)

    records = []
    for _, row in df.iterrows():
        for col, sector, sex in [
            ("Male public", "Public", "Male"),
            ("Female public", "Public", "Female"),
            ("Male private", "Private", "Male"),
            ("Female private", "Private", "Female"),
        ]:
            val = pd.to_numeric(row[col], errors="coerce")
            if pd.notna(val):
                records.append(
                    {"year": int(row["year"]), "sector": sector, "sex": sex, "median_hourly_earnings": float(val)}
                )

    result = pd.DataFrame(records).sort_values(["year", "sector", "sex"]).reset_index(drop=True)
    logger.info(f"Parsed {len(result)} sector/gender earnings records ({result['year'].min()}–{result['year'].max()})")
    return result


def parse_ashe_hourly_earnings_by_age_gender(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE hourly earnings by age group and gender, latest year snapshot.

    Identified by column signature ['Age group', 'Female', 'Male'] with subtitle
    containing 'age'. Present in all publication years, though in different figure slots.

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - age_group: str (e.g. '18-21', '22-29', '30-39', '40-49', '50-59', '60+')
        - sex: str ('Male' or 'Female')
        - median_hourly_earnings: float (£, excluding overtime)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_hourly_earnings_by_age_gender(path)
        >>> sorted(df.columns.tolist())
        ['age_group', 'median_hourly_earnings', 'sex']
    """
    df = _find_linked_sheet(file_path, "hourly_by_age_gender")
    df.columns = ["age_group", "Female", "Male"]
    df = df.dropna(subset=["age_group"])

    records = []
    for _, row in df.iterrows():
        for sex in ("Female", "Male"):
            val = pd.to_numeric(row[sex], errors="coerce")
            if pd.notna(val):
                records.append({"age_group": str(row["age_group"]), "sex": sex, "median_hourly_earnings": float(val)})

    result = pd.DataFrame(records)
    logger.info(f"Parsed {len(result)} age/gender earnings records ({len(df)} age bands)")
    return result


def parse_ashe_hourly_earnings_by_occupation_gender(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE hourly earnings by occupation and gender, latest year snapshot.

    Identified by column signature ['Occupation', 'Female', 'Male'] with subtitle
    containing 'occupation'. Present in all publication years.

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - occupation: str (SOC major group label)
        - sex: str ('Male' or 'Female')
        - median_hourly_earnings: float (£, excluding overtime)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_hourly_earnings_by_occupation_gender(path)
        >>> sorted(df.columns.tolist())
        ['median_hourly_earnings', 'occupation', 'sex']
    """
    df = _find_linked_sheet(file_path, "hourly_by_occupation_gender")
    df.columns = ["occupation", "Female", "Male"]
    df = df.dropna(subset=["occupation"])

    records = []
    for _, row in df.iterrows():
        for sex in ("Female", "Male"):
            val = pd.to_numeric(row[sex], errors="coerce")
            if pd.notna(val):
                records.append({"occupation": str(row["occupation"]), "sex": sex, "median_hourly_earnings": float(val)})

    result = pd.DataFrame(records)
    logger.info(f"Parsed {len(result)} occupation/gender earnings records ({len(df)} occupation groups)")
    return result


def parse_ashe_hourly_earnings_by_pattern_gender(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE hourly earnings by working pattern and gender, latest year snapshot.

    Identified by column signature ['Working pattern', 'Female', 'Male'] with
    subtitle containing 'working pattern' (excluding pay gap / hours tables which
    share similar columns). Present in all publication years.

    Note: part-time females earn *more* per hour than part-time males in NI —
    a reversal of the full-time pattern, documented across 2022–2025.

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - work_pattern: str ('Full-time', 'Part-time', 'All Employees')
        - sex: str ('Male' or 'Female')
        - median_hourly_earnings: float (£, excluding overtime)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_hourly_earnings_by_pattern_gender(path)
        >>> sorted(df.columns.tolist())
        ['median_hourly_earnings', 'sex', 'work_pattern']
    """
    df = _find_linked_sheet(file_path, "hourly_by_pattern_gender")
    df.columns = ["work_pattern", "Female", "Male"]
    df = df.dropna(subset=["work_pattern"])

    records = []
    for _, row in df.iterrows():
        for sex in ("Female", "Male"):
            val = pd.to_numeric(row[sex], errors="coerce")
            if pd.notna(val):
                records.append(
                    {"work_pattern": str(row["work_pattern"]), "sex": sex, "median_hourly_earnings": float(val)}
                )

    result = pd.DataFrame(records)
    logger.info(f"Parsed {len(result)} work pattern/gender earnings records")
    return result


def parse_ashe_ni_uk_earnings_comparison(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE NI vs UK full-time weekly earnings timeseries.

    Identified by column signature ['Year', 'UK', 'NI'] with subtitle containing
    'weekly' and 'full-time'. Stable across all publication years (always Figure 1).

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - year: int
        - location: str ('NI' or 'UK')
        - median_weekly_earnings_fulltime: float (£)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_ni_uk_earnings_comparison(path)
        >>> sorted(df.columns.tolist())
        ['location', 'median_weekly_earnings_fulltime', 'year']
    """
    df = _find_linked_sheet(file_path, "ni_uk_weekly_earnings")
    df.columns = ["year", "UK", "NI"]
    df = df.dropna(subset=["year"])
    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df = df[df["year"].notna()]
    df["year"] = df["year"].astype(int)
    melted = df.melt(id_vars="year", var_name="location", value_name="median_weekly_earnings_fulltime")
    melted["median_weekly_earnings_fulltime"] = pd.to_numeric(
        melted["median_weekly_earnings_fulltime"], errors="coerce"
    )
    return melted.dropna(subset=["median_weekly_earnings_fulltime"]).reset_index(drop=True)


def parse_ashe_uk_regional_pay_ratio(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE high-to-low pay ratio by UK region, latest year snapshot.

    Identified by column signature ['Region', 'Ratio']. Present in all years
    but in different figure slots (Figure 14 in 2022, Figure 16 in 2023,
    Figure 13 in 2024–2025).

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - region: str (UK region name)
        - ratio: float (high-paid / low-paid jobs ratio)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_uk_regional_pay_ratio(path)
        >>> sorted(df.columns.tolist())
        ['ratio', 'region']
    """
    df = _find_linked_sheet(file_path, "uk_regional_pay_ratio")
    df.columns = ["region", "ratio"]
    df = df.dropna(subset=["region"])
    df["ratio"] = pd.to_numeric(df["ratio"], errors="coerce")
    return df.dropna(subset=["ratio"]).reset_index(drop=True)


def parse_ashe_hours_distribution(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE distribution of total weekly paid hours, NI, latest year snapshot.

    Identified by column signature ['Paid hours worked', 'Percentage']. Present
    in all years but in different figure slots (Figure 3 in 2022–2023,
    Figure 9 in 2024–2025).

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - paid_hours_worked: int (hours 0–80)
        - percentage: float (% of employees)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_hours_distribution(path)
        >>> sorted(df.columns.tolist())
        ['paid_hours_worked', 'percentage']
    """
    df = _find_linked_sheet(file_path, "hours_distribution")
    df.columns = ["paid_hours_worked", "percentage"]
    df["paid_hours_worked"] = pd.to_numeric(df["paid_hours_worked"], errors="coerce")
    df["percentage"] = pd.to_numeric(df["percentage"], errors="coerce")
    return df.dropna().reset_index(drop=True)


def parse_ashe_working_pattern_pay_gap(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE working pattern pay gap timeseries, NI vs UK.

    Identified by column signature ['Year', 'UK', 'NI'] with subtitle containing
    'working pattern pay gap'. Present from 2023 onwards (Figure 23 in 2023,
    Figure 19 in 2024–2025).

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - year: int
        - location: str ('NI' or 'UK')
        - working_pattern_pay_gap_pct: float (%)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_working_pattern_pay_gap(path)
        >>> sorted(df.columns.tolist())
        ['location', 'working_pattern_pay_gap_pct', 'year']
    """
    df = _find_linked_sheet(file_path, "working_pattern_pay_gap")
    df.columns = ["year", "UK", "NI"]
    df = df.dropna(subset=["year"])
    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df = df[df["year"].notna()]
    df["year"] = df["year"].astype(int)
    melted = df.melt(id_vars="year", var_name="location", value_name="working_pattern_pay_gap_pct")
    melted["working_pattern_pay_gap_pct"] = pd.to_numeric(melted["working_pattern_pay_gap_pct"], errors="coerce")
    return melted.dropna(subset=["working_pattern_pay_gap_pct"]).reset_index(drop=True)


def parse_ashe_mean_hours_by_pattern_gender(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE mean weekly paid hours by work pattern and gender, NI, latest year.

    Identified by column signature ['Working pattern', 'Males', 'Females', 'All'].
    Present in all years (Figure 21 in 2022–2023, Figure 20 in 2024–2025).

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - work_pattern: str ('Part-time', 'Full-time', 'All Employees')
        - male_mean_hours: float
        - female_mean_hours: float
        - all_mean_hours: float

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_mean_hours_by_pattern_gender(path)
        >>> 'work_pattern' in df.columns
        True
    """
    df = _find_linked_sheet(file_path, "mean_hours_by_pattern_gender")
    df.columns = ["work_pattern", "male_mean_hours", "female_mean_hours", "all_mean_hours"]
    df = df.dropna(subset=["work_pattern"])
    for col in ("male_mean_hours", "female_mean_hours", "all_mean_hours"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df.reset_index(drop=True)


def get_gender_pay_gap(force_refresh: bool = False) -> pd.DataFrame:
    """Get ASHE gender pay gap timeseries for NI and the UK.

    Returns the population-level GPG derived from NISRA's ASHE survey — the
    difference between male and female median hourly earnings as a percentage
    of male earnings, for all employees.

    This is survey-based (HMRC PAYE sample) and covers the whole NI economy,
    complementing the mandatory employer-reported GPG data available via
    ``bolster.data_sources.gender_pay_gap`` (which covers named employers with
    250+ staff only).

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns:

        - year: int (2005–present)
        - location: str ('Northern Ireland' or 'United Kingdom')
        - gender_pay_gap_pct: float

    Example:
        >>> df = get_gender_pay_gap()
        >>> 'gender_pay_gap_pct' in df.columns
        True
    """
    return parse_ashe_gender_pay_gap(_get_linked_file(force_refresh=force_refresh))


def get_hourly_earnings_by_sector_gender(force_refresh: bool = False) -> pd.DataFrame:
    """Get ASHE hourly earnings by sector and gender timeseries for NI.

    Returns median gross hourly earnings (excl. overtime) for NI employees by
    public/private sector and sex, from 2005 to present.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns:

        - year: int (2005–present)
        - sector: str ('Public' or 'Private')
        - sex: str ('Male' or 'Female')
        - median_hourly_earnings: float (£)

    Example:
        >>> df = get_hourly_earnings_by_sector_gender()
        >>> 'median_hourly_earnings' in df.columns
        True
    """
    return parse_ashe_hourly_earnings_by_sector_gender(_get_linked_file(force_refresh=force_refresh))


def get_hourly_earnings_by_age_gender(force_refresh: bool = False) -> pd.DataFrame:
    """Get ASHE hourly earnings by age group and gender for NI, latest year snapshot.

    Returns median gross hourly earnings (excl. overtime) for NI employees by
    age band and sex.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns:

        - age_group: str
        - sex: str ('Male' or 'Female')
        - median_hourly_earnings: float (£)

    Example:
        >>> df = get_hourly_earnings_by_age_gender()
        >>> 'median_hourly_earnings' in df.columns
        True
    """
    return parse_ashe_hourly_earnings_by_age_gender(_get_linked_file(force_refresh=force_refresh))


def get_hourly_earnings_by_occupation_gender(force_refresh: bool = False) -> pd.DataFrame:
    """Get ASHE hourly earnings by occupation and gender for NI, latest year snapshot.

    Returns median gross hourly earnings (excl. overtime) for NI employees by
    SOC major occupation group and sex.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns:

        - occupation: str
        - sex: str ('Male' or 'Female')
        - median_hourly_earnings: float (£)

    Example:
        >>> df = get_hourly_earnings_by_occupation_gender()
        >>> 'median_hourly_earnings' in df.columns
        True
    """
    return parse_ashe_hourly_earnings_by_occupation_gender(_get_linked_file(force_refresh=force_refresh))


def get_hourly_earnings_by_pattern_gender(force_refresh: bool = False) -> pd.DataFrame:
    """Get ASHE hourly earnings by working pattern and gender for NI, latest year snapshot.

    Returns median gross hourly earnings (excl. overtime) for NI employees by
    full-time/part-time and sex.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns:

        - work_pattern: str ('Full-time', 'Part-time', 'All Employees')
        - sex: str ('Male' or 'Female')
        - median_hourly_earnings: float (£)

    Example:
        >>> df = get_hourly_earnings_by_pattern_gender()
        >>> 'median_hourly_earnings' in df.columns
        True
    """
    return parse_ashe_hourly_earnings_by_pattern_gender(_get_linked_file(force_refresh=force_refresh))


def get_ni_uk_earnings_comparison(force_refresh: bool = False) -> pd.DataFrame:
    """Get NI vs UK full-time weekly earnings timeseries.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns: year, location ('NI'/'UK'), median_weekly_earnings_fulltime

    Example:
        >>> df = get_ni_uk_earnings_comparison()
        >>> 'median_weekly_earnings_fulltime' in df.columns
        True
    """
    return parse_ashe_ni_uk_earnings_comparison(_get_linked_file(force_refresh=force_refresh))


def get_uk_regional_pay_ratio(force_refresh: bool = False) -> pd.DataFrame:
    """Get high-to-low pay ratio by UK region, latest year snapshot.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns: region, ratio

    Example:
        >>> df = get_uk_regional_pay_ratio()
        >>> 'ratio' in df.columns
        True
    """
    return parse_ashe_uk_regional_pay_ratio(_get_linked_file(force_refresh=force_refresh))


def get_hours_distribution(force_refresh: bool = False) -> pd.DataFrame:
    """Get distribution of total weekly paid hours for NI employees, latest year snapshot.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns: paid_hours_worked, percentage

    Example:
        >>> df = get_hours_distribution()
        >>> 'percentage' in df.columns
        True
    """
    return parse_ashe_hours_distribution(_get_linked_file(force_refresh=force_refresh))


def get_working_pattern_pay_gap(force_refresh: bool = False) -> pd.DataFrame:
    """Get working pattern pay gap timeseries for NI vs UK.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns: year, location ('NI'/'UK'), working_pattern_pay_gap_pct

    Example:
        >>> df = get_working_pattern_pay_gap()
        >>> 'working_pattern_pay_gap_pct' in df.columns
        True
    """
    return parse_ashe_working_pattern_pay_gap(_get_linked_file(force_refresh=force_refresh))


def get_mean_hours_by_pattern_gender(force_refresh: bool = False) -> pd.DataFrame:
    """Get mean weekly paid hours by work pattern and gender for NI, latest year snapshot.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns: work_pattern, male_mean_hours, female_mean_hours, all_mean_hours

    Example:
        >>> df = get_mean_hours_by_pattern_gender()
        >>> 'male_mean_hours' in df.columns
        True
    """
    return parse_ashe_mean_hours_by_pattern_gender(_get_linked_file(force_refresh=force_refresh))


# ---------------------------------------------------------------------------
# Figures 2, 3, 6 — real earnings (issue #1743)
# ---------------------------------------------------------------------------


def parse_ashe_real_earnings(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE Figure 2: nominal vs real weekly earnings timeseries for NI.

    Identified by column signature ['Year', 'Nominal earnings', 'Real earnings'].
    Real earnings are inflation-adjusted to the latest publication year using
    NISRA's deflator. The series covers full-time employees, April 2005 onwards.

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - year: int
        - nominal_weekly_earnings: float (£, in nominal terms)
        - real_weekly_earnings: float (£, inflation-adjusted to latest year)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_real_earnings(path)
        >>> sorted(df.columns.tolist())
        ['nominal_weekly_earnings', 'real_weekly_earnings', 'year']
    """
    df = _find_linked_sheet(file_path, "real_earnings_timeseries")
    df.columns = ["year", "nominal_weekly_earnings", "real_weekly_earnings"]
    df = df.dropna(subset=["year"])
    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df = df[df["year"].notna()].copy()
    df["year"] = df["year"].astype(int)
    for col in ("nominal_weekly_earnings", "real_weekly_earnings"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    result = df.dropna().reset_index(drop=True)
    logger.info(f"Parsed {len(result)} real earnings records ({result['year'].min()}–{result['year'].max()})")
    return result


def parse_ashe_real_earnings_change_by_pattern(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE Figure 3: annual % change in weekly earnings by work pattern (NI).

    Snapshot for the latest reference year showing nominal vs real (inflation-
    adjusted) annual change for each working pattern.

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - work_pattern: str ('Part-time', 'Full-time', 'All employees')
        - nominal_change_pct: float (annual % change in nominal terms)
        - real_change_pct: float (annual % change in real terms)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_real_earnings_change_by_pattern(path)
        >>> sorted(df.columns.tolist())
        ['nominal_change_pct', 'real_change_pct', 'work_pattern']
    """
    df = _find_linked_sheet(file_path, "real_earnings_change_by_pattern")
    df.columns = ["work_pattern", "nominal_change_pct", "real_change_pct"]
    df = df.dropna(subset=["work_pattern"]).copy()
    for col in ("nominal_change_pct", "real_change_pct"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df.dropna().reset_index(drop=True)


def parse_ashe_real_earnings_index_by_sector(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE Figure 6: real earnings index (2019=100) by sector for NI.

    Identified by column signature ['Year', 'Public', 'Private'] with subtitle
    containing 'annual index'. Indexed real (inflation-adjusted) median weekly
    earnings for full-time employees, base year = six years before the latest
    publication year (typically 2019 in the 2025 publication).

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - year: int
        - sector: str ('Public' or 'Private')
        - real_earnings_index: float (index, base year = 100)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_real_earnings_index_by_sector(path)
        >>> sorted(df.columns.tolist())
        ['real_earnings_index', 'sector', 'year']
    """
    # Figure 6 has *two* tables stacked in one sheet sharing the same column
    # signature (the 2019=100 index and a 2005-onward real-earnings level
    # series). We use the scanner to locate the sheet+header, then walk the
    # sheet rows directly to stop at the first all-None separator row.
    import openpyxl

    # Re-locate the sheet by signature so we don't hard-code its name
    sig = _SHEET_SIGNATURES["real_earnings_index_by_sector"]
    required_cols = tuple(c.lower() for c in sig["cols"])
    keywords = [k.lower() for k in sig.get("subtitle_keywords", [])]
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    matched_sheet = None
    header_idx = None
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            non_empty = [r for r in rows if any(v is not None for v in r)]
            if len(non_empty) < 4:
                continue
            subtitle = str(non_empty[1][0] or "").lower()
            if keywords and not any(k in subtitle for k in keywords):
                continue
            for i, row in enumerate(rows):
                row_vals = tuple(str(v).lower() for v in row if v is not None)
                if all(rc in row_vals for rc in required_cols):
                    matched_sheet = sheet_name
                    header_idx = i
                    break
            if matched_sheet:
                break

        if matched_sheet is None:
            raise NISRADataNotFoundError("Could not find real earnings index sheet (Figure 6)")

        ws = wb[matched_sheet]
        rows = list(ws.iter_rows(values_only=True))
        # Read data rows from just after the header until first all-None row
        data = []
        for row in rows[header_idx + 1 :]:
            if all(v is None for v in row):
                break
            data.append(row[:3])
    finally:
        wb.close()

    df = pd.DataFrame(data, columns=["year", "Public", "Private"])
    df = df.dropna(subset=["year"]).copy()
    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df = df[df["year"].notna()].copy()
    df["year"] = df["year"].astype(int)
    for col in ("Public", "Private"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    melted = df.melt(id_vars="year", var_name="sector", value_name="real_earnings_index")
    return melted.dropna(subset=["real_earnings_index"]).sort_values(["year", "sector"]).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Figures 7, 8 — earnings by occupation and industry (issue #1744)
# ---------------------------------------------------------------------------


def parse_ashe_annual_change_by_occupation(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE Figure 7: annual % change in weekly earnings by occupation (NI, FT).

    Full-time employees only. NISRA labels the column "Industry" in the source
    spreadsheet, but the rows are SOC major occupation groups — disambiguated
    from Figure 8 by subtitle keyword 'by occupation'.

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - occupation: str (SOC major group label)
        - annual_change_pct: float (% change in median gross weekly earnings)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_annual_change_by_occupation(path)
        >>> sorted(df.columns.tolist())
        ['annual_change_pct', 'occupation']
    """
    df = _find_linked_sheet(file_path, "annual_change_by_occupation")
    df.columns = ["occupation", "annual_change_pct"]
    df = df.dropna(subset=["occupation"]).copy()
    df["occupation"] = df["occupation"].astype(str).str.strip()
    df["annual_change_pct"] = pd.to_numeric(df["annual_change_pct"], errors="coerce")
    return df.dropna().reset_index(drop=True)


def parse_ashe_annual_change_by_industry(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE Figure 8: annual % change in weekly earnings by industry (NI, FT).

    Full-time employees by SIC industry section. Disambiguated from Figure 7
    by subtitle keyword 'by industry'.

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - industry: str (SIC section label)
        - annual_change_pct: float (% change in median gross weekly earnings)

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_annual_change_by_industry(path)
        >>> sorted(df.columns.tolist())
        ['annual_change_pct', 'industry']
    """
    df = _find_linked_sheet(file_path, "annual_change_by_industry")
    df.columns = ["industry", "annual_change_pct"]
    df = df.dropna(subset=["industry"]).copy()
    df["industry"] = df["industry"].astype(str).str.strip()
    df["annual_change_pct"] = pd.to_numeric(df["annual_change_pct"], errors="coerce")
    return df.dropna().reset_index(drop=True)


# ---------------------------------------------------------------------------
# Figures 11, 12, 13 — pay distribution (issue #1745)
# Note: Fig 13 (regional pay ratio) is already exposed via
# ``get_uk_regional_pay_ratio`` higher up in this module.
# ---------------------------------------------------------------------------


def parse_ashe_pay_distribution_timeseries(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE Figure 11: proportion of low/middle/high-paid employee jobs in NI.

    Identified by column signature ['Year', 'Low-paid jobs', 'Middle-paid jobs',
    'High-paid jobs']. Covers April 2005 to the latest publication year.

    NISRA defines pay bands relative to UK median hourly earnings:
    low-paid = below two-thirds, high-paid = above 1.5x, middle-paid = the rest.

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - year: int
        - low_paid_pct: float
        - middle_paid_pct: float
        - high_paid_pct: float

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_pay_distribution_timeseries(path)
        >>> sorted(df.columns.tolist())
        ['high_paid_pct', 'low_paid_pct', 'middle_paid_pct', 'year']
    """
    df = _find_linked_sheet(file_path, "pay_distribution_timeseries")
    df.columns = ["year", "low_paid_pct", "middle_paid_pct", "high_paid_pct"]
    df = df.dropna(subset=["year"]).copy()
    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df = df[df["year"].notna()].copy()
    df["year"] = df["year"].astype(int)
    for col in ("low_paid_pct", "middle_paid_pct", "high_paid_pct"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df.dropna().sort_values("year").reset_index(drop=True)


def parse_ashe_pay_distribution_by_classification(file_path: str | Path) -> pd.DataFrame:
    """Parse ASHE Figure 12: pay distribution by working pattern, age, industry, occupation.

    Cross-sectional breakdown of low/middle/high-paid proportions for the latest
    reference year, with multiple classification axes stacked in a single sheet.

    Args:
        file_path: Path to the ASHE linked tables Excel file

    Returns:
        DataFrame with columns:

        - classification: str (e.g. 'Working pattern', 'Age group', 'Industry', 'Occupation')
        - subset: str (the specific category, e.g. 'Full-time', '22-29', 'Education')
        - low_paid_pct: float
        - middle_paid_pct: float
        - high_paid_pct: float

    Example:
        >>> _, year = get_latest_ashe_publication_url()
        >>> path = download_file(get_ashe_file_url(year, 'linked'), cache_ttl_hours=90*24)
        >>> df = parse_ashe_pay_distribution_by_classification(path)
        >>> sorted(df.columns.tolist())
        ['classification', 'high_paid_pct', 'low_paid_pct', 'middle_paid_pct', 'subset']
    """
    df = _find_linked_sheet(file_path, "pay_distribution_by_classification")
    df.columns = ["classification", "subset", "low_paid_pct", "middle_paid_pct", "high_paid_pct"]
    df = df.dropna(subset=["classification", "subset"]).copy()
    df["classification"] = df["classification"].astype(str).str.strip()
    df["subset"] = df["subset"].astype(str).str.strip()
    for col in ("low_paid_pct", "middle_paid_pct", "high_paid_pct"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df.dropna().reset_index(drop=True)


# ---------------------------------------------------------------------------
# Public get_* wrappers
# ---------------------------------------------------------------------------


def get_real_earnings(force_refresh: bool = False) -> pd.DataFrame:
    """Get ASHE nominal vs real weekly earnings timeseries for NI (Figure 2).

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns: year, nominal_weekly_earnings, real_weekly_earnings

    Example:
        >>> df = get_real_earnings()
        >>> 'real_weekly_earnings' in df.columns
        True
    """
    return parse_ashe_real_earnings(_get_linked_file(force_refresh=force_refresh))


def get_real_earnings_change_by_pattern(force_refresh: bool = False) -> pd.DataFrame:
    """Get ASHE annual % change in weekly earnings by work pattern (Figure 3).

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns: work_pattern, nominal_change_pct, real_change_pct

    Example:
        >>> df = get_real_earnings_change_by_pattern()
        >>> 'real_change_pct' in df.columns
        True
    """
    return parse_ashe_real_earnings_change_by_pattern(_get_linked_file(force_refresh=force_refresh))


def get_real_earnings_index_by_sector(force_refresh: bool = False) -> pd.DataFrame:
    """Get ASHE real earnings index by sector for NI (Figure 6).

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns: year, sector, real_earnings_index

    Example:
        >>> df = get_real_earnings_index_by_sector()
        >>> 'real_earnings_index' in df.columns
        True
    """
    return parse_ashe_real_earnings_index_by_sector(_get_linked_file(force_refresh=force_refresh))


def get_annual_change_by_occupation(force_refresh: bool = False) -> pd.DataFrame:
    """Get ASHE annual % change in weekly earnings by occupation, NI (Figure 7).

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns: occupation, annual_change_pct

    Example:
        >>> df = get_annual_change_by_occupation()
        >>> 'annual_change_pct' in df.columns
        True
    """
    return parse_ashe_annual_change_by_occupation(_get_linked_file(force_refresh=force_refresh))


def get_annual_change_by_industry(force_refresh: bool = False) -> pd.DataFrame:
    """Get ASHE annual % change in weekly earnings by industry, NI (Figure 8).

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns: industry, annual_change_pct

    Example:
        >>> df = get_annual_change_by_industry()
        >>> 'annual_change_pct' in df.columns
        True
    """
    return parse_ashe_annual_change_by_industry(_get_linked_file(force_refresh=force_refresh))


def get_pay_distribution_timeseries(force_refresh: bool = False) -> pd.DataFrame:
    """Get ASHE low/middle/high-paid proportion timeseries for NI (Figure 11).

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns: year, low_paid_pct, middle_paid_pct, high_paid_pct

    Example:
        >>> df = get_pay_distribution_timeseries()
        >>> 'low_paid_pct' in df.columns
        True
    """
    return parse_ashe_pay_distribution_timeseries(_get_linked_file(force_refresh=force_refresh))


def get_pay_distribution_by_classification(force_refresh: bool = False) -> pd.DataFrame:
    """Get ASHE pay distribution by classification, NI, latest year (Figure 12).

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns: classification, subset, low_paid_pct, middle_paid_pct, high_paid_pct

    Example:
        >>> df = get_pay_distribution_by_classification()
        >>> 'classification' in df.columns
        True
    """
    return parse_ashe_pay_distribution_by_classification(_get_linked_file(force_refresh=force_refresh))


def validate_ashe_data(df: pd.DataFrame) -> bool:  # pragma: no cover
    """Validate ASHE earnings data integrity.

    Args:
        df: DataFrame from ASHE functions

    Returns:
        True if validation passes, False otherwise
    """
    if df.empty:
        logger.warning("ASHE data is empty")
        return False

    # Check for earnings-related columns
    earnings_indicators = ["earnings", "salary", "pay", "gross", "hourly"]
    has_earnings_data = any(indicator in " ".join(df.columns).lower() for indicator in earnings_indicators)
    if not has_earnings_data:
        logger.warning("No earnings indicators found in ASHE data")
        return False

    return True
