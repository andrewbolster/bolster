"""NISRA Labour Market Statistics Module.

This module provides access to Northern Ireland Statistics and Research Agency (NISRA)
Labour Force Survey (LFS) data.

The Labour Force Survey is a sample survey of households providing information on the
labour force using internationally agreed concepts and definitions. It provides estimates
of employment, unemployment, and economic inactivity.

Data is published quarterly covering 3-month rolling periods (e.g., "July to September 2025").
Some tables (particularly Local Government District breakdowns) are published annually only.

Data Source:
    **Mother Page**: https://www.nisra.gov.uk/statistics/labour-market-and-social-welfare

    This page lists all Labour Market and Social Welfare publications in reverse chronological
    order (newest first). The module automatically scrapes this page to find the latest
    "Quarterly Labour Force Survey Tables" publication, then downloads the Excel file from
    that publication's detail page.

    This ensures the module always retrieves the most recent data without hardcoding dates
    or quarters.

Update Frequency: Quarterly publications covering 3-month rolling periods are released
approximately 6-8 weeks after the reference period ends. Labour Force Survey data is
published four times per year with additional annual publications containing Local Government
District breakdowns. Data is updated as part of NISRA's regular labour market statistics
programme following ONS Labour Force Survey methodology.

Architecture:
    - Automatically discovers latest quarterly LFS publication from NISRA mother page
    - Downloads quarterly Labour Force Survey Excel files
    - Parses multiple tables including:
        * Employment by age band and sex (Table 2.15)
        * Employment by industry sector (Table 2.17)
        * Employment by occupation (Table 2.18)
        * Economic inactivity (Table 2.21)
        * Unemployment rates (Table 2.22)
    - Separately handles annual Local Government District data (Table 1.16a)
      published in dedicated LGD tables file (not in quarterly publications)
    - Returns long-format DataFrames for flexibility
    - Uses standardized Excel parsing utilities from _base.py

Usage:
    >>> from bolster.data_sources.nisra import labour_market
    >>> # Get latest quarterly employment data
    >>> df = labour_market.get_latest_employment()
    >>> # Get economic inactivity data
    >>> df = labour_market.get_latest_economic_inactivity()
    >>> # Get all tables for a specific quarter
    >>> data = labour_market.get_quarterly_data(year=2025, quarter="Jul-Sep")
    >>> # Get annual employment by Local Government District
    >>> lgd_df = labour_market.get_latest_employment_by_lgd()

Example:
    >>> from bolster.data_sources.nisra import labour_market
    >>> emp_df = labour_market.get_latest_employment()
    >>> 'age_group' in emp_df.columns
    True

    >>> lgd_df = labour_market.get_latest_employment_by_lgd()
    >>> 'employment_rate' in lgd_df.columns
    True

Author: Claude Code
Date: 2025-12-21
"""

import logging
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from bolster.data_sources.nisra._base import (
    CACHE_DIR,
    NISRADataNotFoundError,
    download_file,
    find_publication_link,
    safe_float,
    safe_int,
)
from bolster.utils.web import session

logger = logging.getLogger(__name__)

# Cache configuration
LFS_CACHE_DIR = CACHE_DIR / "labour_market"
LFS_CACHE_DIR.mkdir(parents=True, exist_ok=True)

# NISRA LFS publication URL pattern
LFS_BASE_URL = "https://www.nisra.gov.uk"


def _quarter_to_month_names(quarter_str: str) -> tuple[str, str, str]:
    """Convert quarter string to month names.

    Args:
        quarter_str: Quarter identifier like "Jul-Sep" or "Jan-Mar"

    Returns:
        Tuple of (first_month, second_month, third_month)

    Examples:
        >>> _quarter_to_month_names("Jul-Sep")
        ('July', 'August', 'September')
        >>> _quarter_to_month_names("Jan-Mar")
        ('January', 'February', 'March')
    """
    quarter_map = {
        "Jan-Mar": ("January", "February", "March"),
        "Apr-Jun": ("April", "May", "June"),
        "Jul-Sep": ("July", "August", "September"),
        "Oct-Dec": ("October", "November", "December"),
    }

    # Normalize input - remove spaces and hyphens for comparison
    quarter_normalized = quarter_str.strip().replace(" ", "").replace("-", "").lower()

    for key, months in quarter_map.items():
        key_normalized = key.replace("-", "").lower()
        if key_normalized == quarter_normalized:
            return months

    raise ValueError(f"Invalid quarter string: {quarter_str}. Expected format like 'Jul-Sep' or 'Jan-Mar'")


def _build_lfs_url(year: int, quarter: str) -> str:
    """Build the download URL for a quarterly LFS file.

    Args:
        year: Year (e.g., 2025)
        quarter: Quarter string like "Jul-Sep" or "Q3"

    Returns:
        Full URL to the Excel file

    Note:
        The URL pattern is:
        /system/files/statistics/YYYY-MM/lmr-labour-force-survey-quarterly-tables-MONTH-MONTH-YY.xlsx

        Where YYYY-MM is the publication date (usually 2 months after quarter end)
        and MONTH-MONTH-YY is the data period
    """
    # Get month names for the quarter
    first_month, _, third_month = _quarter_to_month_names(quarter)

    # Determine publication month (typically 2 months after quarter end)
    quarter_to_pub_month = {
        "Jan-Mar": "05",  # Published in May
        "Apr-Jun": "08",  # Published in August
        "Jul-Sep": "11",  # Published in November
        "Oct-Dec": "02",  # Published in February (next year)
    }

    # Normalize quarter string - remove spaces and hyphens
    quarter_normalized = quarter.strip().replace(" ", "").replace("-", "").lower()
    pub_month = None
    for key, month in quarter_to_pub_month.items():
        key_normalized = key.replace("-", "").lower()
        if key_normalized == quarter_normalized:
            pub_month = month
            break

    if pub_month is None:
        raise ValueError(f"Could not determine publication month for quarter: {quarter}")

    # For Oct-Dec quarter, publication is in February of next year
    pub_year = year + 1 if pub_month == "02" else year

    # Build filename: lmr-labour-force-survey-quarterly-tables-July-September-25.xlsx
    year_short = str(year)[-2:]  # Last 2 digits
    filename = f"lmr-labour-force-survey-quarterly-tables-{first_month}-{third_month}-{year_short}.xlsx"

    # Build full URL
    return f"{LFS_BASE_URL}/system/files/statistics/{pub_year}-{pub_month}/{filename}"


def download_quarterly_lfs(year: int, quarter: str, force_refresh: bool = False, cache_ttl_days: int = 90) -> Path:
    """Download Labour Force Survey quarterly tables Excel file.

    Args:
        year: Year (e.g., 2025)
        quarter: Quarter string like "Jul-Sep", "Jan-Mar", etc.
        force_refresh: If True, always download fresh data ignoring cache
        cache_ttl_days: Cache time-to-live in days (default: 90 days)

    Returns:
        Path to the downloaded Excel file

    Raises:
        NISRADataNotFoundError: If the file cannot be downloaded

    Example:
        >>> file_path = download_quarterly_lfs(2025, "Jul-Sep")
        >>> file_path.exists()
        True
    """
    # Build download URL
    url = _build_lfs_url(year, quarter)

    logger.info(f"Downloading LFS data from {url}")

    # Convert days to hours for download_file function
    cache_ttl_hours = cache_ttl_days * 24

    try:
        return download_file(url, cache_ttl_hours=cache_ttl_hours, force_refresh=force_refresh)
    except Exception as e:
        raise NISRADataNotFoundError(f"Failed to download LFS quarterly data for {year} {quarter}: {e}") from e


def parse_employment_by_age_sex(file_path: str | Path) -> pd.DataFrame:
    """Parse Table 2.15: Employment by Age Band and Sex.

    Extracts employment numbers and percentages broken down by age group and sex.

    Args:
        file_path: Path to the quarterly LFS Excel file

    Returns:
        DataFrame with columns:
            - quarter_period: Quarter label (e.g., "July to September 2025")
            - age_group: Age band (e.g., "16 to 19", "20 to 24", "25 to 29", etc.)
            - sex: "Male", "Female", or "All Persons"
            - percentage: Percentage of employment in this age group (for this sex)
            - number: Absolute number employed (from Table 2.15b)

    Example:
        >>> url, year, quarter = get_latest_lfs_publication_url()
        >>> path = download_file(url, cache_ttl_hours=90*24)
        >>> df = parse_employment_by_age_sex(path)
        >>> 'age_group' in df.columns
        True
    """
    wb = load_workbook(file_path, data_only=True)

    if "2.15" not in wb.sheetnames:
        raise NISRADataNotFoundError("Table 2.15 (Employment by Age and Sex) not found in file")

    sheet = wb["2.15"]

    # Extract quarter period from title (Row 1 or nearby)
    quarter_period = None
    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        if row[0] and "Percentage in Employment by Age Band" in str(row[0]):
            # Extract date range from title
            match = re.search(r"([A-Z][a-z]+ to [A-Z][a-z]+ \d{4})", str(row[0]))
            if match:
                quarter_period = match.group(1)
                break

    if not quarter_period:
        logger.warning("Could not extract quarter period from Table 2.15 title")
        quarter_period = "Unknown"

    # Find header rows for Table 2.15a (percentages) and Table 2.15b (numbers)
    # Table 2.15a starts around row 6
    # Headers: "Age group", "Male\n(percentage)", "Female\n(percentage)", "All Persons\n(percentage)"

    header_row_a = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=5, max_row=10, values_only=True), start=5):
        if row[0] and "Age group" in str(row[0]):
            header_row_a = row_idx
            break

    if not header_row_a:
        raise NISRADataNotFoundError("Could not find header row for Table 2.15a")

    # Parse Table 2.15a (percentages)
    percentage_records = []
    for row in sheet.iter_rows(min_row=header_row_a + 1, max_row=header_row_a + 15, values_only=True):
        age_group = row[0]

        # Stop at empty rows or footer
        if not age_group or age_group == "All Persons":
            break

        male_pct = safe_float(row[1])
        female_pct = safe_float(row[2])
        all_persons_pct = safe_float(row[3])

        if male_pct is not None:
            percentage_records.append(
                {
                    "quarter_period": quarter_period,
                    "age_group": str(age_group).strip(),
                    "sex": "Male",
                    "percentage": male_pct,
                }
            )

        if female_pct is not None:
            percentage_records.append(
                {
                    "quarter_period": quarter_period,
                    "age_group": str(age_group).strip(),
                    "sex": "Female",
                    "percentage": female_pct,
                }
            )

        if all_persons_pct is not None:
            percentage_records.append(
                {
                    "quarter_period": quarter_period,
                    "age_group": str(age_group).strip(),
                    "sex": "All Persons",
                    "percentage": all_persons_pct,
                }
            )

    df_pct = pd.DataFrame(percentage_records)

    # Find Table 2.15b (numbers) - should be to the right, around column 5
    # Headers: "Sex", "Number"
    header_row_b = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=5, max_row=10, values_only=True), start=5):
        if len(row) > 5 and row[5] and "Sex" in str(row[5]):
            header_row_b = row_idx
            break

    if not header_row_b:
        logger.warning("Could not find Table 2.15b (total numbers by sex)")
        # Return just percentages
        return df_pct

    # Parse Table 2.15b (total numbers) - only has Male, Female, All Persons totals
    # Not broken down by age, so we'll add as a separate record with age_group = "All ages"
    number_records = []
    for row in sheet.iter_rows(min_row=header_row_b + 1, max_row=header_row_b + 5, values_only=True):
        sex = row[5] if len(row) > 5 else None
        number = safe_int(row[6]) if len(row) > 6 else None

        if not sex or number is None:
            continue

        number_records.append(
            {
                "quarter_period": quarter_period,
                "age_group": "All ages",
                "sex": str(sex).strip(),
                "number": number,
            }
        )

    df_num = pd.DataFrame(number_records)

    # Merge percentages with total numbers
    # The total numbers apply to all age groups, so we need to calculate
    # We can merge on sex only for the "All ages" totals

    # Return both as separate info or merged smartly
    # For now, let's add the total numbers as additional rows
    return pd.concat([df_pct, df_num], ignore_index=True)


def parse_economic_inactivity(file_path: str | Path) -> pd.DataFrame:
    """Parse Table 2.21: Economically Inactive by Sex (Time Series).

    Extracts economic inactivity data broken down by sex with historical time series.
    Table 2.21 shows data for the same quarter (e.g., Jul-Sep) across multiple years (2012-2025).

    Args:
        file_path: Path to the quarterly LFS Excel file

    Returns:
        DataFrame with columns:
            - time_period: Quarter label (e.g., "Jul to Sep 2025")
            - sex: "Male", "Female", or "All Persons"
            - economically_inactive_number: Number of people economically inactive
            - economic_inactivity_rate: Percentage economically inactive

    Note:
        Economically inactive persons are those not in employment and not seeking work
        (students, retired, caring for family, long-term sick/disabled, etc.).

        This table provides **historical time series** for the same quarter across years,
        allowing year-over-year comparisons for the same seasonal period.

    Example:
        >>> url, year, quarter = get_latest_lfs_publication_url()
        >>> path = download_file(url, cache_ttl_hours=90*24)
        >>> df = parse_economic_inactivity(path)
        >>> 'economic_inactivity_rate' in df.columns
        True
    """
    wb = load_workbook(file_path, data_only=True)

    if "2.21" not in wb.sheetnames:
        raise NISRADataNotFoundError("Table 2.21 (Economic Inactivity) not found in file")

    sheet = wb["2.21"]

    # Table 2.21 is a time series table with two sub-tables:
    # Table 2.21a: Numbers (Male, Female, All Persons columns)
    # Table 2.21b: Rates (Male, Female, All Persons columns)
    # Rows are time periods (Jul to Sep 2012, Jul to Sep 2013, etc.)

    # Find header for Table 2.21a (columns: Time period, Male, Female, All Persons)
    header_row_a = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=5, max_row=10, values_only=True), start=5):
        if row[0] and "Time period" in str(row[0]):
            header_row_a = row_idx
            break

    if not header_row_a:
        raise NISRADataNotFoundError("Could not find header row for Table 2.21a")

    records = []

    # Parse Table 2.21a (numbers) and Table 2.21b (rates) side-by-side
    # Data rows start after header
    for row in sheet.iter_rows(min_row=header_row_a + 1, max_row=header_row_a + 20, values_only=True):
        time_period_a = row[0]

        # Stop at empty rows
        if not time_period_a:
            break

        # Table 2.21a columns (0-3): Time period, Male, Female, All Persons
        male_number = safe_int(row[1])
        female_number = safe_int(row[2])
        all_persons_number = safe_int(row[3])

        # Table 2.21b columns (5-8): Time period, Male, Female, All Persons
        male_rate = safe_float(row[6]) if len(row) > 6 else None
        female_rate = safe_float(row[7]) if len(row) > 7 else None
        all_persons_rate = safe_float(row[8]) if len(row) > 8 else None

        time_period_str = str(time_period_a).strip()

        # Add record for Male
        if male_number is not None or male_rate is not None:
            records.append(
                {
                    "time_period": time_period_str,
                    "sex": "Male",
                    "economically_inactive_number": male_number,
                    "economic_inactivity_rate": male_rate,
                }
            )

        # Add record for Female
        if female_number is not None or female_rate is not None:
            records.append(
                {
                    "time_period": time_period_str,
                    "sex": "Female",
                    "economically_inactive_number": female_number,
                    "economic_inactivity_rate": female_rate,
                }
            )

        # Add record for All Persons
        if all_persons_number is not None or all_persons_rate is not None:
            records.append(
                {
                    "time_period": time_period_str,
                    "sex": "All Persons",
                    "economically_inactive_number": all_persons_number,
                    "economic_inactivity_rate": all_persons_rate,
                }
            )

    return pd.DataFrame(records)


def get_latest_lfs_publication_url(force_refresh: bool = False) -> tuple[str, str, str]:
    """Find the latest Labour Force Survey quarterly tables publication.

    Scrapes the NISRA Labour Market statistics mother page to find the most recent
    "Quarterly Labour Force Survey Tables" publication.

    Args:
        force_refresh: If True, bypass the page-discovery cache as well as the file cache.

    Returns:
        Tuple of (excel_file_url, year, quarter)
        - excel_file_url: Full URL to the Excel file
        - year: Data year as string (e.g., "2025")
        - quarter: Quarter string (e.g., "Jul-Sep")

    Raises:
        NISRADataNotFoundError: If no publication found

    Example:
        >>> url, year, quarter = get_latest_lfs_publication_url()
        >>> url.startswith('https://')
        True
    """
    excel_url = find_publication_link(
        hub_url="https://www.nisra.gov.uk/statistics/labour-market-and-social-welfare",
        pub_text_contains="Quarterly Labour Force Survey Tables",
        file_href_contains="lmr-labour-force-survey-quarterly-tables",
        force_refresh=force_refresh,
    )

    # Extract quarter and year from filename
    # Pattern: lmr-labour-force-survey-quarterly-tables-July-September-25.xlsx
    filename = excel_url.split("/")[-1]
    match = re.search(
        r"(January|February|March|April|May|June|July|August|September|October|November|December)-"
        r"(January|February|March|April|May|June|July|August|September|October|November|December)-"
        r"(\d{2})",
        filename,
    )

    if not match:
        raise NISRADataNotFoundError(f"Could not extract quarter and year from filename: {filename}")

    month_to_abbrev = {
        "January": "Jan",
        "February": "Feb",
        "March": "Mar",
        "April": "Apr",
        "May": "May",
        "June": "Jun",
        "July": "Jul",
        "August": "Aug",
        "September": "Sep",
        "October": "Oct",
        "November": "Nov",
        "December": "Dec",
    }

    year = f"20{match.group(3)}"
    quarter = f"{month_to_abbrev[match.group(1)]}-{month_to_abbrev[match.group(2)]}"
    logger.info("Extracted quarter: %s %s from %s", quarter, year, filename)
    return (excel_url, year, quarter)


# ============================================================================
# Monthly Labour Market Report (LMR) Functions
# ============================================================================

_LMR_MONTH_NAMES = [
    "january",
    "february",
    "march",
    "april",
    "may",
    "june",
    "july",
    "august",
    "september",
    "october",
    "november",
    "december",
]


def get_latest_monthly_lmr_url(force_refresh: bool = False) -> tuple[str, int, int]:
    """Find the latest monthly Labour Market Report publication URL.

    NISRA publishes a monthly Labour Market Report in addition to the quarterly LFS Tables.
    Monthly reports are grouped by year on collection pages (e.g.,
    ``/publications/labour-market-reports-2026``). This function follows the two-hop
    structure: mother page → yearly collection → latest monthly publication → Excel file.

    The monthly report contains rolling 3-month averages (e.g., "Mar-May 2026") and is
    typically 2–3 months more current than the quarterly LFS Tables file.

    Args:
        force_refresh: If True, bypass the page-discovery cache.

    Returns:
        Tuple of (excel_file_url, year, month_number)
        - excel_file_url: Full URL to the main LFS tables Excel file
        - year: Publication year as int (e.g., 2026)
        - month_number: Publication month as int 1-12 (e.g., 7 for July)

    Raises:
        NISRADataNotFoundError: If no monthly publication found

    Example:
        >>> url, year, month = get_latest_monthly_lmr_url()
        >>> url.startswith('https://')
        True
    """
    from bs4 import BeautifulSoup

    mother_page = "https://www.nisra.gov.uk/statistics/labour-market-and-social-welfare"

    try:
        response = session.get(mother_page, timeout=30, force_refresh=force_refresh)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, "html.parser")

        # Step 1: Find the most recent yearly collection page ("Labour Market Reports - YYYY")
        collection_url = None
        for link in soup.find_all("a", href=True):
            text = link.get_text(strip=True)
            href = link["href"]
            if re.match(r"Labour Market Reports\s*-\s*\d{4}", text) and "/publications/labour-market-reports-" in href:
                collection_url = f"{LFS_BASE_URL}{href}" if href.startswith("/") else href
                logger.info(f"Found LMR collection page: {text} → {collection_url}")
                break

        if not collection_url:  # pragma: no cover
            raise NISRADataNotFoundError("No 'Labour Market Reports - YYYY' collection page found on mother page")

        # Step 2: Find the latest individual monthly report that has a /publications/ page
        # (earlier months may link to datavis only and have no downloadable Excel)
        coll_response = session.get(collection_url, timeout=30, force_refresh=force_refresh)
        coll_response.raise_for_status()
        coll_soup = BeautifulSoup(coll_response.content, "html.parser")

        monthly_links = []
        for link in coll_soup.find_all("a", href=True):
            text = link.get_text(strip=True)
            href = link["href"]
            if (
                re.match(r"Labour Market Report\s*-\s*\w+ \d{4}", text)
                and "/publications/labour-market-report-" in href
            ):
                full_href = f"{LFS_BASE_URL}{href}" if href.startswith("/") else href
                monthly_links.append({"text": text, "url": full_href})

        if not monthly_links:  # pragma: no cover
            raise NISRADataNotFoundError(f"No individual monthly LMR publication links found on {collection_url}")

        # Try each month newest-first (collection page lists oldest-first) until we find Excel
        for pub in reversed(monthly_links):
            pub_response = session.get(pub["url"], timeout=30, force_refresh=force_refresh)
            if not pub_response.ok:  # pragma: no cover
                continue
            pub_soup = BeautifulSoup(pub_response.content, "html.parser")

            excel_url = None
            for link in pub_soup.find_all("a", href=True):
                href = link["href"]
                if (
                    "lmr-labour-force-survey-tables-" in href.lower()
                    and "historical" not in href.lower()
                    and href.endswith(".xlsx")
                ):
                    excel_url = f"{LFS_BASE_URL}{href}" if href.startswith("/") else href
                    break

            if not excel_url:  # pragma: no cover
                logger.debug(f"No Excel file on {pub['url']}, trying next month")
                continue

            # Extract year and month from filename, e.g. lmr-labour-force-survey-tables-july-2026.xlsx
            filename = excel_url.split("/")[-1]
            m = re.search(
                r"lmr-labour-force-survey-tables-(" + "|".join(_LMR_MONTH_NAMES) + r")-(\d{4})\.xlsx",
                filename.lower(),
            )
            if not m:  # pragma: no cover
                logger.debug(f"Could not parse month/year from filename: {filename}")
                continue

            month_name = m.group(1)
            year = int(m.group(2))
            month_number = _LMR_MONTH_NAMES.index(month_name) + 1
            logger.info(f"Found monthly LMR Excel: {month_name} {year} at {excel_url}")
            return (excel_url, year, month_number)

        raise NISRADataNotFoundError(
            "No downloadable Excel file found in any monthly LMR publication"
        )  # pragma: no cover

    except NISRADataNotFoundError:
        raise
    except Exception as e:  # pragma: no cover
        raise NISRADataNotFoundError(f"Failed to fetch monthly LMR publication: {e}") from e


def parse_monthly_lmr_structure(file_path: str | Path) -> pd.DataFrame:
    """Parse Table 2.1a: NI Labour Market Structure from the monthly LMR file.

    Extracts the rolling 3-month aggregate labour market summary (age 16+):
    population, total economically active, total in employment, unemployed,
    and economically inactive. Data is seasonally adjusted.

    Args:
        file_path: Path to the monthly LFS tables Excel file.

    Returns:
        DataFrame with columns:
            - rolling_quarter: Rolling 3-month period label (e.g., "Mar-May 2026")
            - population_16plus: Population aged 16 and over (thousands)
            - economically_active: Total economically active (thousands)
            - in_employment: Total in employment (thousands)
            - unemployed: Unemployed (thousands)
            - economically_inactive: Economically inactive (thousands)

    Example:
        >>> url, year, month = get_latest_monthly_lmr_url()
        >>> from bolster.data_sources.nisra._base import download_file
        >>> path = download_file(url, cache_ttl_hours=30*24)
        >>> df = parse_monthly_lmr_structure(path)
        >>> 'in_employment' in df.columns
        True
    """
    wb = load_workbook(file_path, data_only=True)

    if "2.1" not in wb.sheetnames:
        raise NISRADataNotFoundError("Sheet '2.1' (Labour Market Structure) not found in monthly LMR file")

    sheet = wb["2.1"]

    # Find the header row for Table 2.1a (contains "Rolling monthly quarter")
    header_row = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        if row[0] and "Rolling monthly quarter" in str(row[0]):
            header_row = row_idx
            break

    if not header_row:
        raise NISRADataNotFoundError("Could not find header row in sheet 2.1")

    records = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        period = row[0]
        # Stop at empty rows or annotation rows (e.g., "Change on quarter")
        if not period or not re.match(r"[A-Z][a-z]+-[A-Z][a-z]+ \d{4}", str(period)):
            break

        pop = safe_int(row[1])
        active = safe_int(row[2])
        employed = safe_int(row[3])
        unemployed = safe_int(row[4])
        inactive = safe_int(row[5])

        records.append(
            {
                "rolling_quarter": str(period).strip(),
                "population_16plus": pop,
                "economically_active": active,
                "in_employment": employed,
                "unemployed": unemployed,
                "economically_inactive": inactive,
            }
        )

    if not records:
        raise NISRADataNotFoundError("No data rows found in sheet 2.1")

    return pd.DataFrame(records)


def get_latest_labour_market_overview(force_refresh: bool = False) -> pd.DataFrame:
    """Get the latest monthly Labour Market Report overview data.

    Downloads the most recently published monthly Labour Market Report and returns
    the rolling 3-month labour market structure summary (Table 2.1a). This data is
    published monthly and is typically 2–3 months more current than the quarterly
    LFS Tables.

    Args:
        force_refresh: If True, bypass both the page-discovery and file caches.

    Returns:
        DataFrame with columns:
            - rolling_quarter: Rolling 3-month label (e.g., "Mar-May 2026")
            - population_16plus: Population aged 16+ (thousands, seasonally adjusted)
            - economically_active: Total economically active (thousands)
            - in_employment: Total in employment (thousands)
            - unemployed: Unemployed (thousands)
            - economically_inactive: Economically inactive (thousands)

    Example:
        >>> df = get_latest_labour_market_overview()
        >>> 'in_employment' in df.columns
        True
        >>> len(df) > 0
        True
    """
    excel_url, year, month = get_latest_monthly_lmr_url(force_refresh=force_refresh)
    logger.info(f"Downloading monthly LMR: month {month} of {year}")

    # Monthly reports change monthly; cache for 30 days
    file_path = download_file(excel_url, cache_ttl_hours=30 * 24, force_refresh=force_refresh)
    return parse_monthly_lmr_structure(file_path)


def get_latest_employment(force_refresh: bool = False) -> pd.DataFrame:
    """Get the latest quarterly employment data by age and sex.

    Automatically discovers and downloads the most recent quarterly LFS publication
    from the NISRA website.

    Args:
        force_refresh: If True, always download fresh data ignoring cache

    Returns:
        DataFrame with employment statistics by age group and sex

    Example:
        >>> df = get_latest_employment()
        >>> 'age_group' in df.columns
        True
        >>> young_adults = df[df['age_group'] == '25 to 29']
        >>> len(young_adults) >= 0
        True
    """
    # Discover latest publication from NISRA website
    excel_url, year, quarter = get_latest_lfs_publication_url(force_refresh=force_refresh)

    # Download directly using the discovered URL
    logger.info(f"Downloading latest LFS data: {quarter} {year}")

    # Use download_file directly with the discovered URL
    cache_ttl_hours = 90 * 24  # 90 days for quarterly data
    file_path = download_file(excel_url, cache_ttl_hours=cache_ttl_hours, force_refresh=force_refresh)

    return parse_employment_by_age_sex(file_path)


def get_latest_economic_inactivity(force_refresh: bool = False) -> pd.DataFrame:
    """Get the latest quarterly economic inactivity data.

    Automatically discovers and downloads the most recent quarterly LFS publication
    from the NISRA website.

    Args:
        force_refresh: If True, always download fresh data ignoring cache

    Returns:
        DataFrame with economic inactivity statistics by sex

    Example:
        >>> df = get_latest_economic_inactivity()
        >>> 'sex' in df.columns
        True
    """
    # Discover latest publication from NISRA website
    excel_url, year, quarter = get_latest_lfs_publication_url(force_refresh=force_refresh)

    # Download directly using the discovered URL
    logger.info(f"Downloading latest LFS data: {quarter} {year}")

    # Use download_file directly with the discovered URL
    cache_ttl_hours = 90 * 24  # 90 days for quarterly data
    file_path = download_file(excel_url, cache_ttl_hours=cache_ttl_hours, force_refresh=force_refresh)

    return parse_economic_inactivity(file_path)


def get_quarterly_data(
    year: int, quarter: str, tables: list[str] | None = None, force_refresh: bool = False
) -> dict[str, pd.DataFrame]:
    """Get Labour Force Survey data for a specific quarter.

    Args:
        year: Year (e.g., 2025)
        quarter: Quarter string like "Jul-Sep", "Jan-Mar", etc.
        tables: List of table names to parse (default: ['employment', 'economic_inactivity'])
            Options: 'employment', 'economic_inactivity'
        force_refresh: If True, always download fresh data ignoring cache

    Returns:
        Dictionary mapping table names to DataFrames:
            - 'employment': Employment by age and sex (Table 2.15)
            - 'economic_inactivity': Economic inactivity data (Table 2.21)

    Example:
        >>> data = get_quarterly_data(2025, "Jul-Sep")
        >>> sorted(data.keys())
        ['economic_inactivity', 'employment']
        >>> emp_df = data['employment']
        >>> 'age_group' in emp_df.columns
        True
    """
    if tables is None:
        tables = ["employment", "economic_inactivity"]

    file_path = download_quarterly_lfs(year, quarter, force_refresh=force_refresh)

    result = {}

    if "employment" in tables:
        result["employment"] = parse_employment_by_age_sex(file_path)

    if "economic_inactivity" in tables:
        result["economic_inactivity"] = parse_economic_inactivity(file_path)

    return result


# ============================================================================
# Local Government District Employment Functions
# ============================================================================


def get_latest_lgd_employment_url() -> tuple[str, int]:
    """Get the URL of the latest LFS Local Government District tables publication.

    Uses known URL pattern to construct the file URL. The LGD tables are published
    annually but not listed on the main Labour Market page.

    Returns:
        Tuple of (excel_url, year)

    Raises:
        NISRADataNotFoundError: If unable to find the latest publication

    Example:
        >>> url, year = get_latest_lgd_employment_url()
        >>> url.startswith('https://')
        True
    """
    from datetime import datetime

    logger.info("Fetching latest LFS LGD employment publication URL...")

    # The LGD tables follow a predictable URL pattern
    # We know the publication exists for 2024, try current year first then fall back
    current_year = datetime.now().year
    years_to_try = [current_year, current_year - 1, 2024]  # Try current, last year, and known 2024

    for year in years_to_try:
        # Construct the expected URL
        # Pattern: https://www.nisra.gov.uk/publications/labour-force-survey-tables-local-government-districts-2009-YYYY
        pub_url = f"{LFS_BASE_URL}/publications/labour-force-survey-tables-local-government-districts-2009-{year}"

        try:
            response = session.get(pub_url, timeout=30)
            if response.status_code == 200:
                # Found the publication page, now find the Excel file
                from bs4 import BeautifulSoup

                soup = BeautifulSoup(response.content, "html.parser")

                # Find Excel file link
                for link in soup.find_all("a", href=True):
                    href = link["href"]
                    if ".xlsx" in href.lower() and "labour" in href.lower():
                        excel_url = href
                        if not excel_url.startswith("http"):
                            excel_url = f"{LFS_BASE_URL}{excel_url}"

                        logger.info(f"Found latest LFS LGD tables: {year} at {excel_url}")
                        return excel_url, year

        except Exception:
            continue

    # If all attempts fail, return the known 2024 file URL as fallback
    fallback_url = (
        "https://www.nisra.gov.uk/system/files/statistics/2025-07/"
        "Labour%20Force%20Survey%20-%20Tables%20for%20Local%20Government%20Districts%202009%20to%202024.xlsx"
    )
    logger.warning(f"Could not find latest LGD publication via pattern, using known 2024 file: {fallback_url}")
    return fallback_url, 2024


def parse_employment_by_lgd(file_path: str | Path, year: int = None) -> pd.DataFrame:
    """Parse Table 1.16a: Employment by Local Government District (ages 16+).

    Extracts employment statistics for all 11 Northern Ireland LGDs from the annual
    LFS LGD tables file.

    Args:
        file_path: Path to the LFS LGD tables Excel file
        year: Year of the data (if not provided, will be extracted from sheet name)

    Returns:
        DataFrame with columns:
            - year: int
            - lgd: str (Local Government District name)
            - population_16plus: int (thousands)
            - economically_active: int (thousands)
            - in_employment: int (thousands)
            - full_time_employment: int (thousands)
            - part_time_employment: int (thousands)
            - economically_inactive: int (thousands)
            - economic_activity_rate: float (%)
            - employment_rate: float (%)

    Example:
        >>> url, year = get_latest_lgd_employment_url()
        >>> path = download_file(url, cache_ttl_hours=180*24)
        >>> df = parse_employment_by_lgd(path, year=year)
        >>> 'employment_rate' in df.columns
        True
    """
    logger.info(f"Parsing LFS LGD employment from: {file_path}")

    # If year not provided, try to extract from filename or use the latest sheet
    if year is None:
        # Try to extract from filename
        filename_match = re.search(r"(\d{4})\.xlsx", str(file_path))
        if filename_match:
            year = int(filename_match.group(1))
        else:
            # Use the latest year sheet
            wb = load_workbook(file_path, data_only=True)
            # Find numeric sheet names (years)
            year_sheets = [int(s) for s in wb.sheetnames if s.isdigit()]
            if year_sheets:
                year = max(year_sheets)
            else:
                raise ValueError("Cannot determine year from filename or sheet names")

    # Read Table 1.16a (ages 16+)
    # Header is in row 6 (0-indexed), data starts in row 7
    # Skip first 6 rows to use row 6 as header
    df = pd.read_excel(file_path, sheet_name=str(year), skiprows=6, nrows=12)

    # Rename columns to standard names
    df.columns = [
        "lgd",
        "population_16plus",
        "economically_active",
        "in_employment",
        "full_time_employment",
        "part_time_employment",
        "economically_inactive",
        "economic_activity_rate",
        "employment_rate",
        "notes",
    ]

    # Drop notes column and Total row
    df = df.drop(columns=["notes"])
    df = df[df["lgd"] != "Total"].reset_index(drop=True)

    # Add year column
    df["year"] = year

    # Reorder columns
    df = df[
        [
            "year",
            "lgd",
            "population_16plus",
            "economically_active",
            "in_employment",
            "full_time_employment",
            "part_time_employment",
            "economically_inactive",
            "economic_activity_rate",
            "employment_rate",
        ]
    ]

    logger.info(f"Parsed {len(df)} LGD employment records for {year}")
    return df


def get_latest_employment_by_lgd(force_refresh: bool = False) -> pd.DataFrame:
    """Get the latest Employment by Local Government District data.

    Downloads and parses the most recent annual LFS LGD tables publication.
    Results are cached for 180 days unless force_refresh=True.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with employment statistics for all 11 NI Local Government Districts

    Example:
        >>> df = get_latest_employment_by_lgd()
        >>> 'employment_rate' in df.columns
        True
        >>> len(df) > 0
        True
    """
    excel_url, year = get_latest_lgd_employment_url()

    # Cache for 180 days (published annually)
    file_path = download_file(excel_url, cache_ttl_hours=180 * 24, force_refresh=force_refresh)

    return parse_employment_by_lgd(file_path, year=year)


def validate_labour_market_data(df: pd.DataFrame) -> bool:  # pragma: no cover
    """Validate labour market data integrity.

    Args:
        df: DataFrame from labour market functions

    Returns:
        True if validation passes, False otherwise
    """
    if df.empty:
        logger.warning("Labour market data is empty")
        return False

    # Check for employment-related columns
    employment_indicators = ["employed", "unemployed", "rate", "count", "percentage"]
    has_employment_data = any(indicator in " ".join(df.columns).lower() for indicator in employment_indicators)

    if not has_employment_data:
        logger.warning("No employment indicators found in labour market data")
        return False

    # Check for reasonable employment rates/percentages
    rate_cols = [col for col in df.columns if "rate" in col.lower() or "percentage" in col.lower()]
    for col in rate_cols:
        if (
            col in df.columns
            and df[col].dtype in ["float64", "int64"]
            and ((df[col] < 0).any() or (df[col] > 100).any())
        ):
            logger.warning(f"Employment rates out of range in column {col}")
            return False

    return True
