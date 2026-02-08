"""NISRA Monthly Occupancy Statistics Data Source.

Provides access to monthly hotel and accommodation occupancy data for Northern Ireland.

Data includes:

- Hotel room and bed occupancy rates from 2011 to present
- Rooms and beds sold monthly (hotels)
- Small Service Accommodation (SSA) occupancy from 2013 to present (B&Bs, guest houses, and similar establishments)
- Rooms and beds sold monthly (SSA)

The survey provides indicative monthly occupancy rates that are revised and
finalised in the Annual Publication.

Data Source:
    **Mother Page**: https://www.nisra.gov.uk/statistics/tourism/occupancy-surveys

    This page lists all occupancy survey publications. The module automatically
    scrapes this page to find the latest "Hotel Occupancy" or "Small Service
    Accommodation" publication, then downloads the Excel file.

Update Frequency: Monthly (published around the 15th of the following month)
Geographic Coverage: Northern Ireland
Reference Date: Month of survey

Example:
    >>> from bolster.data_sources.nisra.tourism import occupancy
    >>> # Get latest hotel occupancy rates
    >>> df = occupancy.get_latest_hotel_occupancy()
    >>> print(df.head())

    >>> # Get rooms/beds sold
    >>> df_sold = occupancy.get_latest_rooms_beds_sold()
    >>> print(f"Total rooms sold in 2024: {df_sold[df_sold['year']==2024]['rooms_sold'].sum():,.0f}")

    >>> # Get SSA (B&B/guest house) occupancy rates
    >>> df_ssa = occupancy.get_latest_ssa_occupancy()
    >>> print(df_ssa.head())

    >>> # Compare hotel vs SSA occupancy
    >>> df_combined = occupancy.get_combined_occupancy()
    >>> print(df_combined.groupby('accommodation_type')['room_occupancy'].mean())
"""

import logging
import re
from pathlib import Path
from typing import Literal, Tuple, Union

import pandas as pd

from bolster.utils.web import session

from .._base import NISRADataNotFoundError, NISRAValidationError, download_file

logger = logging.getLogger(__name__)

# Base URL for occupancy statistics
OCCUPANCY_BASE_URL = "https://www.nisra.gov.uk/statistics/tourism/occupancy-surveys"


def get_latest_hotel_occupancy_publication_url() -> Tuple[str, str]:
    """Scrape NISRA occupancy surveys page to find the latest hotel occupancy file.

    Navigates the publication structure:
    1. Scrapes mother page for latest hotel occupancy publication
    2. Follows link to publication detail page
    3. Finds hotel occupancy Excel file

    Returns:
        Tuple of (excel_file_url, publication_date)

    Raises:
        NISRADataNotFoundError: If publication or file not found
    """
    from bs4 import BeautifulSoup

    mother_page = OCCUPANCY_BASE_URL

    try:
        # Use shared session with retry logic for resilient requests
        response = session.get(mother_page, timeout=30)
        response.raise_for_status()
    except Exception as e:
        raise NISRADataNotFoundError(f"Failed to fetch occupancy surveys page: {e}")

    soup = BeautifulSoup(response.content, "html.parser")

    # Find latest hotel occupancy publication
    # Pattern: "December 2024 hotel occupancy survey publications" or similar
    pub_link = None
    pub_date = None

    for link in soup.find_all("a", href=True):
        link_text = link.get_text(strip=True)

        # Match hotel occupancy publications
        if "hotel occupancy" in link_text.lower() and "publications" in link["href"]:
            href = link["href"]

            if href.startswith("/"):
                href = f"https://www.nisra.gov.uk{href}"

            # Extract month/year from link text if available
            # Pattern: "December 2024 hotel occupancy..."
            date_match = re.search(r"([A-Z][a-z]+)\s+(\d{4})", link_text)
            if date_match:
                pub_date = f"{date_match.group(1)} {date_match.group(2)}"

            # Take first match (should be newest due to reverse chronological order)
            pub_link = href
            logger.info(f"Found hotel occupancy publication: {link_text}")
            break

    if not pub_link:
        raise NISRADataNotFoundError("Could not find hotel occupancy publication on mother page")

    # Scrape the publication page for Excel file
    try:
        pub_response = session.get(pub_link, timeout=30)
        pub_response.raise_for_status()
    except Exception as e:
        raise NISRADataNotFoundError(f"Failed to fetch publication page: {e}")

    pub_soup = BeautifulSoup(pub_response.content, "html.parser")

    # Find hotel occupancy Excel file
    # Pattern: "2024-Hotel-December-excel-(online).xls" or similar
    excel_url = None

    for link in pub_soup.find_all("a", href=True):
        href = link["href"]

        # Match Excel files (.xls or .xlsx)
        if "Hotel" in href and (href.endswith(".xls") or href.endswith(".xlsx")):
            if href.startswith("/"):
                href = f"https://www.nisra.gov.uk{href}"

            excel_url = href
            logger.info(f"Found hotel occupancy Excel file: {href}")
            break

    if not excel_url:
        raise NISRADataNotFoundError("Could not find hotel occupancy Excel file on publication page")

    return excel_url, pub_date or "Unknown"


def get_latest_ssa_occupancy_publication_url() -> Tuple[str, str]:
    """Scrape NISRA occupancy surveys page to find the latest SSA file.

    SSA = Small Service Accommodation (B&Bs, guest houses, etc.)

    Navigates the publication structure:
    1. Scrapes mother page for latest SSA occupancy publication
    2. Follows link to publication detail page
    3. Finds SSA occupancy Excel file

    Returns:
        Tuple of (excel_file_url, publication_date)

    Raises:
        NISRADataNotFoundError: If publication or file not found
    """
    from bs4 import BeautifulSoup

    mother_page = OCCUPANCY_BASE_URL

    try:
        # Use shared session with retry logic for resilient requests
        response = session.get(mother_page, timeout=30)
        response.raise_for_status()
    except Exception as e:
        raise NISRADataNotFoundError(f"Failed to fetch occupancy surveys page: {e}")

    soup = BeautifulSoup(response.content, "html.parser")

    # Find latest SSA occupancy publication
    # Pattern: "November 2025 Small Service Accommodation occupancy" or similar
    pub_link = None
    pub_date = None

    for link in soup.find_all("a", href=True):
        link_text = link.get_text(strip=True)

        # Match SSA publications
        if "small service accommodation" in link_text.lower() and "publications" in link["href"]:
            href = link["href"]

            if href.startswith("/"):
                href = f"https://www.nisra.gov.uk{href}"

            # Extract month/year from link text if available
            date_match = re.search(r"([A-Z][a-z]+)\s+(\d{4})", link_text)
            if date_match:
                pub_date = f"{date_match.group(1)} {date_match.group(2)}"

            # Take first match (should be newest due to reverse chronological order)
            pub_link = href
            logger.info(f"Found SSA occupancy publication: {link_text}")
            break

    if not pub_link:
        raise NISRADataNotFoundError("Could not find SSA occupancy publication on mother page")

    # Scrape the publication page for Excel file
    try:
        pub_response = session.get(pub_link, timeout=30)
        pub_response.raise_for_status()
    except Exception as e:
        raise NISRADataNotFoundError(f"Failed to fetch SSA publication page: {e}")

    pub_soup = BeautifulSoup(pub_response.content, "html.parser")

    # Find SSA occupancy Excel file
    # Pattern: "2025-Small Service-November-excel-(online).xls" or similar
    excel_url = None

    for link in pub_soup.find_all("a", href=True):
        href = link["href"]

        # Match Excel files (.xls or .xlsx) - SSA files have "Small" or "Service" in name
        if ("Small" in href or "Service" in href) and (href.endswith(".xls") or href.endswith(".xlsx")):
            if href.startswith("/"):
                href = f"https://www.nisra.gov.uk{href}"

            excel_url = href
            logger.info(f"Found SSA occupancy Excel file: {href}")
            break

    if not excel_url:
        raise NISRADataNotFoundError("Could not find SSA occupancy Excel file on publication page")

    return excel_url, pub_date or "Unknown"


def _find_table_by_title(file_path: Union[str, Path], title_contains: str) -> str:
    """Find the sheet name that contains the specified title.

    Args:
        file_path: Path to the Excel file
        title_contains: String to search for in the first row of each table

    Returns:
        Sheet name matching the criteria

    Raises:
        NISRAValidationError: If no matching table is found
    """
    import openpyxl

    file_path = Path(file_path)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().startswith("Table"):
            ws = wb[sheet_name]
            # Check first cell for title
            first_cell = ws.cell(1, 1).value
            if first_cell and title_contains.lower() in str(first_cell).lower():
                wb.close()
                return sheet_name

    wb.close()
    raise NISRAValidationError(f"Could not find table with title containing '{title_contains}'")


def parse_hotel_occupancy_rates(file_path: Union[str, Path]) -> pd.DataFrame:
    """Parse NISRA hotel occupancy rates from Excel file (Table 1).

    Args:
        file_path: Path to the hotel occupancy Excel file

    Returns:
        DataFrame with columns:
            - date: datetime (first day of month)
            - year: int
            - month: str (month name)
            - room_occupancy: float (room occupancy rate, 0-1)
            - bed_occupancy: float (bed occupancy rate, 0-1)

    Raises:
        NISRAValidationError: If file structure is unexpected
    """
    file_path = Path(file_path)

    try:
        # Read Table 1 - Monthly occupancy rates
        # Row 0-7: Notes and metadata
        # Row 8: "<< link back to contents >>"
        # Row 9: Headers (Month, YYYY Room occupancy, YYYY Bed occupancy, ...)
        # Row 10+: Data (January, February, ...)
        df_raw = pd.read_excel(
            file_path,
            sheet_name="Table 1",
            header=None,
            skiprows=9,  # Skip to header row
            nrows=13,  # Read header + 12 months
        )
    except Exception as e:
        raise NISRAValidationError(f"Failed to read hotel occupancy file: {e}")

    # First row contains headers
    headers = df_raw.iloc[0].tolist()
    df_raw = df_raw.iloc[1:].reset_index(drop=True)  # Drop header row
    df_raw.columns = headers

    # First column should be "Month"
    month_col = headers[0]

    # Identify year columns - they have pattern "YYYY Room occupancy" or "YYYY Bed occupancy"
    year_room_cols = {}
    year_bed_cols = {}

    for col in headers[1:]:
        col_str = str(col)
        room_match = re.match(r"(\d{4})\s*Room\s*occupancy", col_str, re.IGNORECASE)
        bed_match = re.match(r"(\d{4})\s*Bed\s*occupancy", col_str, re.IGNORECASE)

        if room_match:
            year_room_cols[int(room_match.group(1))] = col
        elif bed_match:
            year_bed_cols[int(bed_match.group(1))] = col

    # Build long-format data
    records = []
    month_map = {
        "January": 1,
        "February": 2,
        "March": 3,
        "April": 4,
        "May": 5,
        "June": 6,
        "July": 7,
        "August": 8,
        "September": 9,
        "October": 10,
        "November": 11,
        "December": 12,
    }

    for _, row in df_raw.iterrows():
        month_name = str(row[month_col]).strip()
        if month_name not in month_map:
            continue

        month_num = month_map[month_name]

        # Get data for each year
        for year in sorted(set(year_room_cols.keys()) | set(year_bed_cols.keys())):
            room_occ = None
            bed_occ = None

            if year in year_room_cols:
                val = row[year_room_cols[year]]
                if pd.notna(val) and val != 0 and val != "c":
                    room_occ = float(val)

            if year in year_bed_cols:
                val = row[year_bed_cols[year]]
                if pd.notna(val) and val != 0 and val != "c":
                    bed_occ = float(val)

            # Only add if we have at least one value
            if room_occ is not None or bed_occ is not None:
                records.append(
                    {
                        "year": year,
                        "month": month_name,
                        "month_num": month_num,
                        "room_occupancy": room_occ,
                        "bed_occupancy": bed_occ,
                    }
                )

    df = pd.DataFrame(records)

    # Create datetime column
    df["date"] = pd.to_datetime({"year": df["year"], "month": df["month_num"], "day": 1})

    # Select and reorder columns
    result = df[["date", "year", "month", "room_occupancy", "bed_occupancy"]].copy()

    # Sort by date
    result = result.sort_values("date").reset_index(drop=True)

    # Log summary
    logger.info(
        f"Parsed {len(result)} monthly hotel occupancy records "
        f"({result['date'].min().strftime('%Y-%m')} to {result['date'].max().strftime('%Y-%m')})"
    )

    return result


def parse_rooms_beds_sold(file_path: Union[str, Path]) -> pd.DataFrame:
    """Parse NISRA hotel rooms and beds sold from Excel file (Table 3).

    Args:
        file_path: Path to the hotel occupancy Excel file

    Returns:
        DataFrame with columns:
            - date: datetime (first day of month)
            - year: int
            - month: str (month name)
            - rooms_sold: float (number of rooms sold)
            - beds_sold: float (number of beds sold)

    Raises:
        NISRAValidationError: If file structure is unexpected
    """
    file_path = Path(file_path)

    try:
        # Read Table 3 - Rooms and beds sold
        # Row 0-7: Notes and metadata
        # Row 8: "<< link to methodology >>"
        # Row 9: "<< link back to contents >>"
        # Row 10: Headers (Month, YYYY Rooms sold, YYYY Beds sold, ...)
        # Row 11+: Data (January, February, ...)
        df_raw = pd.read_excel(
            file_path,
            sheet_name="Table 3",
            header=None,
            skiprows=10,  # Skip to header row
            nrows=13,  # Read header + 12 months
        )
    except Exception as e:
        raise NISRAValidationError(f"Failed to read hotel occupancy file: {e}")

    # First row contains headers
    headers = df_raw.iloc[0].tolist()
    df_raw = df_raw.iloc[1:].reset_index(drop=True)  # Drop header row
    df_raw.columns = headers

    # First column should be "Month"
    month_col = headers[0]

    # Identify year columns
    year_rooms_cols = {}
    year_beds_cols = {}

    for col in headers[1:]:
        col_str = str(col)
        rooms_match = re.match(r"(\d{4})\s*Rooms\s*sold", col_str, re.IGNORECASE)
        beds_match = re.match(r"(\d{4})\s*Beds\s*sold", col_str, re.IGNORECASE)

        if rooms_match:
            year_rooms_cols[int(rooms_match.group(1))] = col
        elif beds_match:
            year_beds_cols[int(beds_match.group(1))] = col

    # Build long-format data
    records = []
    month_map = {
        "January": 1,
        "February": 2,
        "March": 3,
        "April": 4,
        "May": 5,
        "June": 6,
        "July": 7,
        "August": 8,
        "September": 9,
        "October": 10,
        "November": 11,
        "December": 12,
    }

    for _, row in df_raw.iterrows():
        month_name = str(row[month_col]).strip()
        if month_name not in month_map:
            continue

        month_num = month_map[month_name]

        # Get data for each year
        for year in sorted(set(year_rooms_cols.keys()) | set(year_beds_cols.keys())):
            rooms_sold = None
            beds_sold = None

            if year in year_rooms_cols:
                val = row[year_rooms_cols[year]]
                if pd.notna(val) and val != 0 and val != "*" and val != "c":
                    rooms_sold = float(val)

            if year in year_beds_cols:
                val = row[year_beds_cols[year]]
                if pd.notna(val) and val != 0 and val != "*" and val != "c":
                    beds_sold = float(val)

            # Only add if we have at least one value
            if rooms_sold is not None or beds_sold is not None:
                records.append(
                    {
                        "year": year,
                        "month": month_name,
                        "month_num": month_num,
                        "rooms_sold": rooms_sold,
                        "beds_sold": beds_sold,
                    }
                )

    df = pd.DataFrame(records)

    # Create datetime column
    df["date"] = pd.to_datetime({"year": df["year"], "month": df["month_num"], "day": 1})

    # Select and reorder columns
    result = df[["date", "year", "month", "rooms_sold", "beds_sold"]].copy()

    # Sort by date
    result = result.sort_values("date").reset_index(drop=True)

    logger.info(
        f"Parsed {len(result)} monthly rooms/beds sold records "
        f"({result['date'].min().strftime('%Y-%m')} to {result['date'].max().strftime('%Y-%m')})"
    )

    return result


def get_latest_hotel_occupancy(force_refresh: bool = False) -> pd.DataFrame:
    """Get the latest monthly hotel occupancy rates data.

    Automatically discovers and downloads the most recent hotel occupancy
    data from the NISRA website.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns:
            - date: datetime (first day of month)
            - year: int
            - month: str (month name)
            - room_occupancy: float (room occupancy rate, 0-1)
            - bed_occupancy: float (bed occupancy rate, 0-1)

    Raises:
        NISRADataNotFoundError: If latest publication cannot be found
        NISRAValidationError: If file structure is unexpected

    Example:
        >>> df = get_latest_hotel_occupancy()
        >>> # Get 2024 average occupancy
        >>> avg_2024 = df[df['year'] == 2024]['room_occupancy'].mean()
        >>> print(f"2024 average room occupancy: {avg_2024:.1%}")
    """
    excel_url, pub_date = get_latest_hotel_occupancy_publication_url()

    logger.info(f"Downloading hotel occupancy data ({pub_date}) from: {excel_url}")

    # Cache for 30 days (monthly data)
    cache_ttl_hours = 30 * 24
    file_path = download_file(excel_url, cache_ttl_hours=cache_ttl_hours, force_refresh=force_refresh)

    return parse_hotel_occupancy_rates(file_path)


def get_latest_rooms_beds_sold(force_refresh: bool = False) -> pd.DataFrame:
    """Get the latest monthly rooms and beds sold data.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns:
            - date: datetime (first day of month)
            - year: int
            - month: str (month name)
            - rooms_sold: float (number of rooms sold)
            - beds_sold: float (number of beds sold)

    Example:
        >>> df = get_latest_rooms_beds_sold()
        >>> total_2024 = df[df['year'] == 2024]['rooms_sold'].sum()
        >>> print(f"Total rooms sold in 2024: {total_2024:,.0f}")
    """
    excel_url, pub_date = get_latest_hotel_occupancy_publication_url()

    logger.info(f"Downloading rooms/beds sold data ({pub_date}) from: {excel_url}")

    cache_ttl_hours = 30 * 24
    file_path = download_file(excel_url, cache_ttl_hours=cache_ttl_hours, force_refresh=force_refresh)

    return parse_rooms_beds_sold(file_path)


# ============================================================================
# Small Service Accommodation (SSA) Functions
# ============================================================================


def parse_ssa_occupancy_rates(file_path: Union[str, Path]) -> pd.DataFrame:
    """Parse NISRA SSA occupancy rates from Excel file (Table 1).

    SSA = Small Service Accommodation (B&Bs, guest houses, etc.)

    Args:
        file_path: Path to the SSA occupancy Excel file

    Returns:
        DataFrame with columns:
            - date: datetime (first day of month)
            - year: int
            - month: str (month name)
            - room_occupancy: float (room occupancy rate, 0-1)
            - bed_occupancy: float (bed occupancy rate, 0-1)

    Raises:
        NISRAValidationError: If file structure is unexpected
    """
    file_path = Path(file_path)

    try:
        # SSA files have trailing space on some sheet names
        # Try both "Table 1" and "Table 1 "
        sheet_name = None
        xl = pd.ExcelFile(file_path)
        for name in xl.sheet_names:
            if name.strip() == "Table 1":
                sheet_name = name
                break

        if not sheet_name:
            raise NISRAValidationError("Could not find Table 1 in SSA file")

        # Read Table 1 - Monthly occupancy rates
        # Structure is same as hotel: rows 0-9 metadata, row 10 headers, rows 11+ data
        df_raw = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,
            skiprows=10,  # Skip to header row
            nrows=13,  # Read header + 12 months
        )
    except NISRAValidationError:
        raise
    except Exception as e:
        raise NISRAValidationError(f"Failed to read SSA occupancy file: {e}")

    # First row contains headers
    headers = df_raw.iloc[0].tolist()
    df_raw = df_raw.iloc[1:].reset_index(drop=True)  # Drop header row
    df_raw.columns = headers

    # First column should be "Month"
    month_col = headers[0]

    # Identify year columns - same pattern as hotel
    year_room_cols = {}
    year_bed_cols = {}

    for col in headers[1:]:
        col_str = str(col)
        room_match = re.match(r"(\d{4})\s*Room\s*occupancy", col_str, re.IGNORECASE)
        bed_match = re.match(r"(\d{4})\s*Bed\s*occupancy", col_str, re.IGNORECASE)

        if room_match:
            year_room_cols[int(room_match.group(1))] = col
        elif bed_match:
            year_bed_cols[int(bed_match.group(1))] = col

    # Build long-format data
    records = []
    month_map = {
        "January": 1,
        "February": 2,
        "March": 3,
        "April": 4,
        "May": 5,
        "June": 6,
        "July": 7,
        "August": 8,
        "September": 9,
        "October": 10,
        "November": 11,
        "December": 12,
    }

    for _, row in df_raw.iterrows():
        month_name = str(row[month_col]).strip()
        if month_name not in month_map:
            continue

        month_num = month_map[month_name]

        # Get data for each year
        for year in sorted(set(year_room_cols.keys()) | set(year_bed_cols.keys())):
            room_occ = None
            bed_occ = None

            if year in year_room_cols:
                val = row[year_room_cols[year]]
                if pd.notna(val) and val != 0 and val != "c":
                    room_occ = float(val)

            if year in year_bed_cols:
                val = row[year_bed_cols[year]]
                if pd.notna(val) and val != 0 and val != "c":
                    bed_occ = float(val)

            # Only add if we have at least one value
            if room_occ is not None or bed_occ is not None:
                records.append(
                    {
                        "year": year,
                        "month": month_name,
                        "month_num": month_num,
                        "room_occupancy": room_occ,
                        "bed_occupancy": bed_occ,
                    }
                )

    df = pd.DataFrame(records)

    # Create datetime column
    df["date"] = pd.to_datetime({"year": df["year"], "month": df["month_num"], "day": 1})

    # Select and reorder columns
    result = df[["date", "year", "month", "room_occupancy", "bed_occupancy"]].copy()

    # Sort by date
    result = result.sort_values("date").reset_index(drop=True)

    logger.info(
        f"Parsed {len(result)} monthly SSA occupancy records "
        f"({result['date'].min().strftime('%Y-%m')} to {result['date'].max().strftime('%Y-%m')})"
    )

    return result


def parse_ssa_rooms_beds_sold(file_path: Union[str, Path]) -> pd.DataFrame:
    """Parse NISRA SSA rooms and beds sold from Excel file (Table 2).

    Note: SSA uses Table 2 for rooms/beds sold, while Hotel uses Table 3.

    Args:
        file_path: Path to the SSA occupancy Excel file

    Returns:
        DataFrame with columns:
            - date: datetime (first day of month)
            - year: int
            - month: str (month name)
            - rooms_sold: float (number of rooms sold)
            - beds_sold: float (number of beds sold)

    Raises:
        NISRAValidationError: If file structure is unexpected
    """
    file_path = Path(file_path)

    try:
        # Read Table 2 - Rooms and beds sold
        df_raw = pd.read_excel(
            file_path,
            sheet_name="Table 2",
            header=None,
            skiprows=10,  # Skip to header row
            nrows=13,  # Read header + 12 months
        )
    except Exception as e:
        raise NISRAValidationError(f"Failed to read SSA occupancy file: {e}")

    # First row contains headers
    headers = df_raw.iloc[0].tolist()
    df_raw = df_raw.iloc[1:].reset_index(drop=True)  # Drop header row
    df_raw.columns = headers

    # First column should be "Month"
    month_col = headers[0]

    # Identify year columns
    year_rooms_cols = {}
    year_beds_cols = {}

    for col in headers[1:]:
        col_str = str(col)
        rooms_match = re.match(r"(\d{4})\s*Rooms\s*sold", col_str, re.IGNORECASE)
        beds_match = re.match(r"(\d{4})\s*Beds\s*sold", col_str, re.IGNORECASE)

        if rooms_match:
            year_rooms_cols[int(rooms_match.group(1))] = col
        elif beds_match:
            year_beds_cols[int(beds_match.group(1))] = col

    # Build long-format data
    records = []
    month_map = {
        "January": 1,
        "February": 2,
        "March": 3,
        "April": 4,
        "May": 5,
        "June": 6,
        "July": 7,
        "August": 8,
        "September": 9,
        "October": 10,
        "November": 11,
        "December": 12,
    }

    for _, row in df_raw.iterrows():
        month_name = str(row[month_col]).strip()
        if month_name not in month_map:
            continue

        month_num = month_map[month_name]

        # Get data for each year
        for year in sorted(set(year_rooms_cols.keys()) | set(year_beds_cols.keys())):
            rooms_sold = None
            beds_sold = None

            if year in year_rooms_cols:
                val = row[year_rooms_cols[year]]
                if pd.notna(val) and val != 0 and val != "*" and val != "c":
                    rooms_sold = float(val)

            if year in year_beds_cols:
                val = row[year_beds_cols[year]]
                if pd.notna(val) and val != 0 and val != "*" and val != "c":
                    beds_sold = float(val)

            # Only add if we have at least one value
            if rooms_sold is not None or beds_sold is not None:
                records.append(
                    {
                        "year": year,
                        "month": month_name,
                        "month_num": month_num,
                        "rooms_sold": rooms_sold,
                        "beds_sold": beds_sold,
                    }
                )

    df = pd.DataFrame(records)

    # Create datetime column
    df["date"] = pd.to_datetime({"year": df["year"], "month": df["month_num"], "day": 1})

    # Select and reorder columns
    result = df[["date", "year", "month", "rooms_sold", "beds_sold"]].copy()

    # Sort by date
    result = result.sort_values("date").reset_index(drop=True)

    logger.info(
        f"Parsed {len(result)} monthly SSA rooms/beds sold records "
        f"({result['date'].min().strftime('%Y-%m')} to {result['date'].max().strftime('%Y-%m')})"
    )

    return result


def get_latest_ssa_occupancy(force_refresh: bool = False) -> pd.DataFrame:
    """Get the latest monthly SSA occupancy rates data.

    SSA = Small Service Accommodation (B&Bs, guest houses, etc.)

    Automatically discovers and downloads the most recent SSA occupancy
    data from the NISRA website.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns:
            - date: datetime (first day of month)
            - year: int
            - month: str (month name)
            - room_occupancy: float (room occupancy rate, 0-1)
            - bed_occupancy: float (bed occupancy rate, 0-1)

    Raises:
        NISRADataNotFoundError: If latest publication cannot be found
        NISRAValidationError: If file structure is unexpected

    Example:
        >>> df = get_latest_ssa_occupancy()
        >>> # Get 2024 average occupancy for B&Bs
        >>> avg_2024 = df[df['year'] == 2024]['room_occupancy'].mean()
        >>> print(f"2024 average SSA room occupancy: {avg_2024:.1%}")
    """
    excel_url, pub_date = get_latest_ssa_occupancy_publication_url()

    logger.info(f"Downloading SSA occupancy data ({pub_date}) from: {excel_url}")

    # Cache for 30 days (monthly data)
    cache_ttl_hours = 30 * 24
    file_path = download_file(excel_url, cache_ttl_hours=cache_ttl_hours, force_refresh=force_refresh)

    return parse_ssa_occupancy_rates(file_path)


def get_latest_ssa_rooms_beds_sold(force_refresh: bool = False) -> pd.DataFrame:
    """Get the latest monthly SSA rooms and beds sold data.

    SSA = Small Service Accommodation (B&Bs, guest houses, etc.)

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns:
            - date: datetime (first day of month)
            - year: int
            - month: str (month name)
            - rooms_sold: float (number of rooms sold)
            - beds_sold: float (number of beds sold)

    Example:
        >>> df = get_latest_ssa_rooms_beds_sold()
        >>> total_2024 = df[df['year'] == 2024]['rooms_sold'].sum()
        >>> print(f"Total SSA rooms sold in 2024: {total_2024:,.0f}")
    """
    excel_url, pub_date = get_latest_ssa_occupancy_publication_url()

    logger.info(f"Downloading SSA rooms/beds sold data ({pub_date}) from: {excel_url}")

    cache_ttl_hours = 30 * 24
    file_path = download_file(excel_url, cache_ttl_hours=cache_ttl_hours, force_refresh=force_refresh)

    return parse_ssa_rooms_beds_sold(file_path)


# ============================================================================
# Combined Accommodation Functions
# ============================================================================


def get_combined_occupancy(force_refresh: bool = False) -> pd.DataFrame:
    """Get combined hotel and SSA occupancy data with accommodation type column.

    This function fetches both hotel and SSA occupancy data and combines them
    into a single DataFrame with an 'accommodation_type' column to distinguish
    between the two accommodation types.

    Args:
        force_refresh: If True, bypass cache and download fresh data

    Returns:
        DataFrame with columns:
            - date: datetime (first day of month)
            - year: int
            - month: str (month name)
            - room_occupancy: float (room occupancy rate, 0-1)
            - bed_occupancy: float (bed occupancy rate, 0-1)
            - accommodation_type: str ('hotel' or 'ssa')

    Example:
        >>> df = get_combined_occupancy()
        >>> # Compare hotel vs SSA occupancy by year
        >>> df.groupby(['year', 'accommodation_type'])['room_occupancy'].mean()
    """
    hotel_df = get_latest_hotel_occupancy(force_refresh=force_refresh)
    hotel_df["accommodation_type"] = "hotel"

    ssa_df = get_latest_ssa_occupancy(force_refresh=force_refresh)
    ssa_df["accommodation_type"] = "ssa"

    combined = pd.concat([hotel_df, ssa_df], ignore_index=True)
    combined = combined.sort_values(["date", "accommodation_type"]).reset_index(drop=True)

    logger.info(f"Combined {len(hotel_df)} hotel + {len(ssa_df)} SSA = {len(combined)} total occupancy records")

    return combined


def compare_accommodation_types(
    df: pd.DataFrame, metric: Literal["room_occupancy", "bed_occupancy"] = "room_occupancy"
) -> pd.DataFrame:
    """Compare occupancy between hotel and SSA by year.

    Args:
        df: DataFrame from get_combined_occupancy()
        metric: Which occupancy metric to compare

    Returns:
        DataFrame with columns:
            - year: int
            - hotel_{metric}: float
            - ssa_{metric}: float
            - difference: float (hotel - ssa)
            - ratio: float (hotel / ssa)

    Example:
        >>> df = get_combined_occupancy()
        >>> comparison = compare_accommodation_types(df)
        >>> # Hotels typically have higher occupancy than B&Bs
        >>> print(comparison[['year', 'hotel_room_occupancy', 'ssa_room_occupancy', 'difference']])
    """
    pivot = df.pivot_table(
        index="year",
        columns="accommodation_type",
        values=metric,
        aggfunc="mean",
    ).reset_index()

    pivot.columns.name = None
    pivot = pivot.rename(columns={"hotel": f"hotel_{metric}", "ssa": f"ssa_{metric}"})

    # Calculate comparison metrics
    pivot["difference"] = pivot[f"hotel_{metric}"] - pivot[f"ssa_{metric}"]
    pivot["ratio"] = pivot[f"hotel_{metric}"] / pivot[f"ssa_{metric}"]

    return pivot


def get_occupancy_by_year(df: pd.DataFrame, year: int) -> pd.DataFrame:
    """Filter occupancy data for a specific year.

    Args:
        df: DataFrame from get_latest_hotel_occupancy()
        year: Year to filter

    Returns:
        Filtered DataFrame
    """
    return df[df["year"] == year].reset_index(drop=True)


def get_occupancy_summary_by_year(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate annual occupancy averages and statistics.

    Args:
        df: DataFrame from get_latest_hotel_occupancy()

    Returns:
        DataFrame with columns:
            - year: int
            - avg_room_occupancy: float
            - avg_bed_occupancy: float
            - months_reported: int
    """
    summary = (
        df.groupby("year")
        .agg(
            avg_room_occupancy=("room_occupancy", "mean"),
            avg_bed_occupancy=("bed_occupancy", "mean"),
            months_reported=("room_occupancy", lambda x: x.notna().sum()),
        )
        .reset_index()
    )

    return summary


def get_seasonal_patterns(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate average occupancy by month across all years.

    Args:
        df: DataFrame from get_latest_hotel_occupancy()

    Returns:
        DataFrame with columns:
            - month: str (month name)
            - avg_room_occupancy: float
            - avg_bed_occupancy: float

    Example:
        >>> df = get_latest_hotel_occupancy()
        >>> seasonal = get_seasonal_patterns(df)
        >>> # Find peak season
        >>> peak = seasonal.loc[seasonal['avg_room_occupancy'].idxmax()]
        >>> print(f"Peak month: {peak['month']} ({peak['avg_room_occupancy']:.1%})")
    """
    month_order = [
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    ]

    summary = (
        df.groupby("month")
        .agg(
            avg_room_occupancy=("room_occupancy", "mean"),
            avg_bed_occupancy=("bed_occupancy", "mean"),
        )
        .reset_index()
    )

    # Sort by month order
    summary["month"] = pd.Categorical(summary["month"], categories=month_order, ordered=True)
    summary = summary.sort_values("month").reset_index(drop=True)

    return summary


def validate_occupancy_data(df: pd.DataFrame) -> bool:
    """Validate tourism occupancy data integrity.

    Args:
        df: DataFrame from get_latest_occupancy_data

    Returns:
        True if validation passes, False otherwise
    """
    if df.empty:
        logger.warning("Occupancy data is empty")
        return False

    required_cols = {"month", "accommodation_type"}
    if not required_cols.issubset(df.columns):
        missing = required_cols - set(df.columns)
        logger.warning(f"Missing required occupancy columns: {missing}")
        return False

    # Check for reasonable occupancy percentages
    percentage_cols = [col for col in df.columns if "occupancy" in col.lower() and col != "accommodation_type"]
    for col in percentage_cols:
        if col in df.columns:
            if (df[col] < 0).any() or (df[col] > 100).any():
                logger.warning(f"Occupancy percentages out of range in column {col}")
                return False

    return True
